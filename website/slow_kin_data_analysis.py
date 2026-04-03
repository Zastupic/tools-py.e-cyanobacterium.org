from flask import Blueprint, render_template, request, jsonify, send_file
import os, io, time, base64
import pandas as pd
import numpy as np
try:
    from scipy.optimize import curve_fit as _scipy_curve_fit
    _SCIPY_OK = True
except ImportError:
    _SCIPY_OK = False
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from . import UPLOAD_FOLDER
from werkzeug.utils import secure_filename

slow_kin_data_analysis = Blueprint('slow_kin_data_analysis', __name__)

# ─── AquaPen protocol definitions ────────────────────────────────────────────

# Timing arrays for each protocol (microseconds)
# Each protocol: dict with keys 'fm_time', 'ft_time', 'npq_time', 'fm_labels', 'ft_labels',
#                'fv_labels', 'npq_labels', 'qp_labels', 'qy_labels', 'etr_labels',
#                'n_light', 'n_dark', 'detect_fn'

AQUAPEN_PROTOCOLS = {
    'NPQ1': {
        # 5 light steps (L1–L4, Lss) + 3 dark recovery (D1–D3)
        'fm_time':   [1422801,19364701,31261001,43157301,55053601,66949901,83876601,109879301,135882001],
        'ft_time':   [207601,18564701,30461001,42357301,54253601,66149901,83076601,109079301,135082001],
        'npq_time':  [19364701,31261001,43157301,55053601,66949901,83876601,109879301,135882001],
        'fm_labels': ['Fm','Fm_L1','Fm_L2','Fm_L3','Fm_L4','Fm_Lss','Fm_D1','Fm_D2','Fm_D3'],
        'ft_labels': ['F0','Ft_L1','Ft_L2','Ft_L3','Ft_L4','Ft_Lss','Ft_D1','Ft_D2','Ft_D3'],
        'fv_labels': ['Fv','Fv_L1','Fv_L2','Fv_L3','Fv_L4','Fv_Lss','Fv_D1','Fv_D2','Fv_D3'],
        'npq_labels':['NPQ_L1','NPQ_L2','NPQ_L3','NPQ_L4','NPQ_Lss','NPQ_D1','NPQ_D2','NPQ_D3'],
        'qp_labels': ['QP_L1','QP_L2','QP_L3','QP_L4','QP_Lss','QP_D1','QP_D2','QP_D3'],
        'qy_labels': ['QY_max (Fv/Fm)','QY_L1','QY_L2','QY_L3','QY_L4','QY_Lss','QY_D1','QY_D2','QY_D3'],
        'etr_labels':['ETR_Fv/Fm','ETR_L1','ETR_L2','ETR_L3','ETR_L4','ETR_Lss','ETR_D1','ETR_D2','ETR_D3'],
        'last_fm_dark': 'Fm_D3',
        'last_npq_dark': 'NPQ_D3',
        'last_qp_dark': 'Qp_D3',
        'last_qy_dark': 'QY_D3',
        'detect': lambda s: (
            s.astype(str).str.contains('NPQ_L4').any() and
            not s.astype(str).str.contains('NPQ_L9').any()
        ),
    },
    'NPQ2': {
        # 10 light steps (L1–L9, Lss) + 7 dark (D1–D7)
        'fm_time':   [1422801,32425501,53314201,74202901,95091601,115980301,136869001,157757701,178646401,199535101,220423801,243327701,304174001,365020301,425866601,486712901,547559201,608405501],
        'ft_time':   [207601,31625501,52514201,73402901,94291601,115180301,136069001,156957701,177846401,198735101,219623801,242527701,303374001,364220301,425066601,485912901,546759201,607605501],
        'npq_time':  [32425501,53314201,74202901,95091601,115980301,136869001,157757701,178646401,199535101,220423801,243327701,304174001,365020301,425866601,486712901,547559201,608405501],
        'fm_labels': ['Fm','Fm_L1','Fm_L2','Fm_L3','Fm_L4','Fm_L5','Fm_L6','Fm_L7','Fm_L8','Fm_L9','Fm_Lss','Fm_D1','Fm_D2','Fm_D3','Fm_D4','Fm_D5','Fm_D6','Fm_D7'],
        'ft_labels': ['Ft','Ft_L1','Ft_L2','Ft_L3','Ft_L4','Ft_L5','Ft_L6','Ft_L7','Ft_L8','Ft_L9','Ft_Lss','Ft_D1','Ft_D2','Ft_D3','Ft_D4','Ft_D5','Ft_D6','Ft_D7'],
        'fv_labels': ['Fv','Fv_L1','Fv_L2','Fv_L3','Fv_L4','Fv_L5','Fv_L6','Fv_L7','Fv_L8','Fv_L9','Fv_Lss','Fv_D1','Fv_D2','Fv_D3','Fv_D4','Fv_D5','Fv_D6','Fv_D7'],
        'npq_labels':['NPQ_L1','NPQ_L2','NPQ_L3','NPQ_L4','NPQ_L5','NPQ_L6','NPQ_L7','NPQ_L8','NPQ_L9','NPQ_Lss','NPQ_D1','NPQ_D2','NPQ_D3','NPQ_D4','NPQ_D5','NPQ_D6','NPQ_D7'],
        'qp_labels': ['QP_L1','QP_L2','QP_L3','QP_L4','QP_L5','QP_L6','QP_L7','QP_L8','QP_L9','QP_Lss','QP_D1','QP_D2','QP_D3','QP_D4','QP_D5','QP_D6','QP_D7'],
        'qy_labels': ['QY_max (Fv/Fm)','QY_L1','QY_L2','QY_L3','QY_L4','QY_L5','QY_L6','QY_L7','QY_L8','QY_L9','QY_Lss','QY_D1','QY_D2','QY_D3','QY_D4','QY_D5','QY_D6','QY_D7'],
        'etr_labels':['ETR_Fv/Fm','ETR_L1','ETR_L2','ETR_L3','ETR_L4','ETR_L5','ETR_L6','ETR_L7','ETR_L8','ETR_L9','ETR_Lss','ETR_D1','ETR_D2','ETR_D3','ETR_D4','ETR_D5','ETR_D6','ETR_D7'],
        'last_fm_dark': 'Fm_D7',
        'last_npq_dark': 'NPQ_D7',
        'last_qp_dark': 'Qp_D7',
        'last_qy_dark': 'QY_D7',
        'detect': lambda s: (
            s.astype(str).str.contains('NPQ_L9').any() and
            s.astype(str).str.contains('NPQ_D7').any()
        ),
    },
    'NPQ3': {
        # 10 light steps (L1–L9, Lss) + 2 dark (D1–D2)
        'fm_time':   [1422801,32425501,53314201,74202901,95091601,115980301,136869001,157757701,178646401,199535101,220423801,243327701,264224001],
        'ft_time':   [207601,31625501,52514201,73402901,94291601,115180301,136069001,156957701,177846401,198735101,219623801,242527701,263424001],
        'npq_time':  [32425501,53314201,74202901,95091601,115980301,136869001,157757701,178646401,199535101,220423801,243327701,264224001],
        'fm_labels': ['Fm','Fm_L1','Fm_L2','Fm_L3','Fm_L4','Fm_L5','Fm_L6','Fm_L7','Fm_L8','Fm_L9','Fm_Lss','Fm_D1','Fm_D2'],
        'ft_labels': ['Ft','Ft_L1','Ft_L2','Ft_L3','Ft_L4','Ft_L5','Ft_L6','Ft_L7','Ft_L8','Ft_L9','Ft_Lss','Ft_D1','Ft_D2'],
        'fv_labels': ['Fv','Fv_L1','Fv_L2','Fv_L3','Fv_L4','Fv_L5','Fv_L6','Fv_L7','Fv_L8','Fv_L9','Fv_Lss','Fv_D1','Fv_D2'],
        'npq_labels':['NPQ_L1','NPQ_L2','NPQ_L3','NPQ_L4','NPQ_L5','NPQ_L6','NPQ_L7','NPQ_L8','NPQ_L9','NPQ_Lss','NPQ_D1','NPQ_D2'],
        'qp_labels': ['QP_L1','QP_L2','QP_L3','QP_L4','QP_L5','QP_L6','QP_L7','QP_L8','QP_L9','QP_Lss','QP_D1','QP_D2'],
        'qy_labels': ['QY_max (Fv/Fm)','QY_L1','QY_L2','QY_L3','QY_L4','QY_L5','QY_L6','QY_L7','QY_L8','QY_L9','QY_Lss','QY_D1','QY_D2'],
        'etr_labels':['ETR_Fv/Fm','ETR_L1','ETR_L2','ETR_L3','ETR_L4','ETR_L5','ETR_L6','ETR_L7','ETR_L8','ETR_L9','ETR_Lss','ETR_D1','ETR_D2'],
        'last_fm_dark': 'Fm_D2',
        'last_npq_dark': 'NPQ_D2',
        'last_qp_dark': 'Qp_D2',
        'last_qy_dark': 'QY_D2',
        'detect': lambda s: (
            s.astype(str).str.contains('NPQ_L9').any() and
            not s.astype(str).str.contains('NPQ_D7').any()
        ),
    },
}


# ─── helpers ─────────────────────────────────────────────────────────────────

def _safe(v):
    """Convert numpy scalars / NaN to JSON-safe Python types."""
    if v is None:
        return None
    try:
        if np.isnan(v):
            return None
    except Exception:
        pass
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return float(v)
    return v


def _series_to_list(s):
    return [_safe(x) for x in s]


def _cleanup_old_files(folder, seconds=1200):
    now = time.time()
    for fname in os.listdir(folder):
        fpath = os.path.join(folder, fname).replace('\\', '/')
        if os.stat(fpath).st_mtime < now - seconds:
            os.remove(fpath)


# ─── state-transition helpers ────────────────────────────────────────────────

def _fit_single_exp(t_arr, y_arr):
    """
    Fit y = A·exp(-k·(t-t0)) + C where t0 = t_arr[0].
    t_arr and y_arr must be in consistent units (seconds for t, a.u. for y).
    Returns a result dict.
    """
    t = np.asarray([v for v in t_arr if v is not None], dtype=float)
    y = np.asarray([v for v in y_arr if v is not None], dtype=float)
    valid = np.isfinite(t) & np.isfinite(y)
    t, y = t[valid], y[valid]
    n = int(len(t))

    delta_pct = _safe(float((y[-1] - y[0]) / abs(y[0]) * 100)) if n >= 2 and y[0] != 0 else None
    base = dict(n_points=n, delta_fm_pct=delta_pct,
                tau=None, k=None, half_time=None, r_sq=None,
                fit_t=[], fit_y=[], fit_ok=False,
                low_confidence=False, insufficient_data=(n < 4))
    if n < 4 or not _SCIPY_OK:
        return base

    t0  = float(t[0])
    tn  = t - t0
    A0  = float(y[0] - y[-1])
    C0  = float(y[-1])
    k0  = max(1.5 / max(float(tn[-1]), 1e-6), 1e-8)

    def _f(x, A, k, C):
        return A * np.exp(-k * x) + C

    try:
        popt, _ = _scipy_curve_fit(
            _f, tn, y, p0=[A0, k0, C0],
            bounds=([-np.inf, 1e-8, -np.inf], [np.inf, np.inf, np.inf]),
            maxfev=10000,
        )
        A_fit, k_fit, C_fit = popt
        tau       = float(1.0 / k_fit)
        half_time = float(tau * np.log(2))
        y_pred    = _f(tn, *popt)
        ss_res    = float(np.sum((y - y_pred) ** 2))
        ss_tot    = float(np.sum((y - float(np.mean(y))) ** 2))
        r_sq      = float(1.0 - ss_res / ss_tot) if ss_tot > 1e-12 else None
        t_dense   = np.linspace(0.0, float(tn[-1]), 120)
        y_dense   = _f(t_dense, *popt)
        return dict(
            n_points=n, delta_fm_pct=delta_pct,
            tau=_safe(tau), k=_safe(float(k_fit)), half_time=_safe(half_time),
            r_sq=_safe(r_sq),
            fit_t=[_safe(float(v)) for v in (t_dense + t0).tolist()],
            fit_y=[_safe(float(v)) for v in y_dense.tolist()],
            fit_ok=True, low_confidence=(n < 6), insufficient_data=False,
        )
    except Exception:
        return {**base, 'insufficient_data': False}


def _detect_par_phases(par_vals):
    """
    Segment a PAR array into contiguous constant-PAR phases.
    Light phases → AL1, AL2 …; dark phases → D1, D2 …
    Returns list[dict] with keys: label, type, par, start_idx, end_idx.
    """
    phases, n = [], len(par_vals)
    if not n:
        return phases
    cur   = float(round(par_vals[0]))
    start = 0
    li = di = 0
    for i in range(1, n + 1):
        nxt = float(round(par_vals[i])) if i < n else None
        if nxt != cur:
            if cur == 0:
                di += 1; label, ptype = f'D{di}', 'dark'
            else:
                li += 1; label, ptype = f'AL{li}', 'light'
            phases.append(dict(label=label, type=ptype, par=cur,
                               start_idx=start, end_idx=i - 1))
            if i < n:
                cur = nxt; start = i
    return phases


def _calc_st_phases(file_stems, fm_data, t_data_s, phases, include_d1=False):
    """
    Fit a single exponential to Fm' in each phase for every sample.
    fm_data   : {stem: list[float|None]} – Fm' per sample (same length as t_data_s)
    t_data_s  : list[float]              – time in SECONDS
    phases    : list[dict] from _detect_par_phases or built for AquaPen
    include_d1: if False, skip first data point of every dark phase
    Returns   : {stem: [phase_result_dict, …]}
    """
    results = {}
    for stem in file_stems:
        fm = fm_data.get(stem, [])
        file_phases = []
        for ph in phases:
            s = int(ph['start_idx'])
            e = min(int(ph['end_idx']), len(fm) - 1)
            t_seg = list(t_data_s[s: e + 1])
            y_seg = list(fm[s: e + 1])
            if ph['type'] == 'dark' and not include_d1:
                t_seg = t_seg[1:]; y_seg = y_seg[1:]
            pairs = [(tv, yv) for tv, yv in zip(t_seg, y_seg)
                     if tv is not None and yv is not None]
            if not pairs:
                continue
            tv, yv = zip(*pairs)
            fit = _fit_single_exp(list(tv), list(yv))
            file_phases.append(dict(
                label=ph['label'], type=ph['type'], par=ph.get('par'),
                fm_vals=[_safe(v) for v in yv],
                t_vals=[_safe(v) for v in tv],
                **fit,
            ))
        results[stem] = file_phases
    return results


# ─── processing branches ─────────────────────────────────────────────────────

def _process_mcpam_raw(files, reduce_data):
    """
    Read MC-PAM raw data CSV files (2 columns: time;fluorescence).
    Returns unified JSON payload or raises ValueError.
    """
    summary = pd.DataFrame()
    file_stems = []
    for i, file in enumerate(files):
        df = pd.read_csv(file.stream, sep=';', engine='python')
        if len(df.columns) < 2 or 'ETR' in df.columns:
            raise ValueError(
                f'File {secure_filename(file.filename or "")} does not look like a raw data file '
                '(expected 2 columns without ETR column).'
            )
        stem = str.lower(os.path.splitext(file.filename or '')[0])
        file_stems.append(stem)
        if i == 0:
            summary = df.iloc[:, 0:2].copy()
            summary.rename(columns={summary.columns[1]: stem}, inplace=True)
        else:
            col = df.iloc[:, 1:2].copy()
            col.rename(columns={col.columns[0]: stem}, inplace=True)
            summary = pd.concat([summary, col], axis=1)

    if summary.empty:
        raise ValueError('No valid raw data found.')

    if reduce_data and len(summary) > 10000:
        factor = int(len(summary) / 10000)
        summary = summary.iloc[::factor].reset_index(drop=True)

    time_col = summary.iloc[:, 0]
    traces = {}
    for stem in file_stems:
        traces[stem] = _series_to_list(summary[stem].astype(float))

    return {
        'fluorometer': 'MC-PAM',
        'mode': 'raw_data',
        'protocol': None,
        'files': file_stems,
        'file_stem': file_stems[0] if file_stems else '',
        'time_unit': 's',
        'raw_time': _series_to_list(time_col.astype(float)),
        'raw_traces': traces,
        'has_params': False,
        'param_time': [],
        'params': {},
        'has_summary': False,
        'summary': {},
        'has_state_transitions': False,
        'st_include_d1': False,
        'st_phases_meta': [],
        'state_transitions': {},
    }


def _process_mcpam_params(files):
    """
    Read MC-PAM parameter CSV files (columns: time, F, Fm', Y(II), ETR, PAR).
    Returns unified JSON payload or raises ValueError.
    """
    ft_all = fm_all = qy_all = etr_all = par_all = pd.DataFrame()
    file_stems = []

    for i, file in enumerate(files):
        df = pd.read_csv(file.stream, sep=';', engine='python')
        if 'ETR' not in df.columns:
            raise ValueError(
                f'File {secure_filename(file.filename or "")} does not look like a parameter file '
                '(ETR column missing).'
            )
        stem = str.lower(os.path.splitext(file.filename or '')[0])
        file_stems.append(stem)
        if i == 0:
            time_s = df.iloc[:, 0]
            ft_all  = pd.concat([time_s, df['F'].rename(stem)], axis=1)
            fm_all  = pd.concat([time_s, df["Fm'"].rename(stem)], axis=1)
            qy_all  = pd.concat([time_s, df['Y(II)'].rename(stem)], axis=1)
            etr_all = pd.concat([time_s, df['ETR'].rename(stem)], axis=1)
            par_all = pd.concat([time_s, df['PAR'].rename(stem)], axis=1)
            ft_all.rename(columns={ft_all.columns[0]: 'Time (s)'}, inplace=True)
            fm_all.rename(columns={fm_all.columns[0]: 'Time (s)'}, inplace=True)
            qy_all.rename(columns={qy_all.columns[0]: 'Time (s)'}, inplace=True)
            etr_all.rename(columns={etr_all.columns[0]: 'Time (s)'}, inplace=True)
            par_all.rename(columns={par_all.columns[0]: 'Time (s)'}, inplace=True)
        else:
            ft_all  = pd.concat([ft_all,  df['F'].rename(stem)], axis=1)
            fm_all  = pd.concat([fm_all,  df["Fm'"].rename(stem)], axis=1)
            qy_all  = pd.concat([qy_all,  df['Y(II)'].rename(stem)], axis=1)
            etr_all = pd.concat([etr_all, df['ETR'].rename(stem)], axis=1)
            par_all = pd.concat([par_all, df['PAR'].rename(stem)], axis=1)

    # Drop rows with NaN in time or all samples
    for df in [ft_all, fm_all, qy_all, etr_all, par_all]:
        df.dropna(subset=[df.columns[0]], inplace=True)

    time_s = _series_to_list(ft_all.iloc[:, 0].astype(float))

    # Compute derived: Fv, QP, qN, NPQ (using Fm), NPQ (using Fm_max)
    f0 = ft_all.iloc[0, 1:]    # first Ft row per sample
    fm = fm_all.iloc[0, 1:]    # first Fm' row per sample
    fm_max = fm_all.iloc[:, 1:].max()

    fv_all = fm_all.iloc[:, 1:] - ft_all.iloc[:, 1:]
    qp_all = (fm_all.iloc[:, 1:] - ft_all.iloc[:, 1:]) / (fm_all.iloc[:, 1:] - f0.values)
    qn_all = (fm - fm_all.iloc[:, 1:]) / (fm - f0.values)
    npq_fm = (fm - fm_all.iloc[:, 1:]) / fm_all.iloc[:, 1:]
    npq_fm_max = (fm_max - fm_all.iloc[:, 1:]) / fm_all.iloc[:, 1:]

    params = {}
    for stem in file_stems:
        if stem in ft_all.columns:
            params[stem] = {
                'ft':        _series_to_list(ft_all[stem].astype(float)),
                'fm':        _series_to_list(fm_all[stem].astype(float)),
                'fv':        _series_to_list(fv_all[stem].astype(float)),
                'npq':       _series_to_list(npq_fm[stem].astype(float)),
                'npq_fmmax': _series_to_list(npq_fm_max[stem].astype(float)),
                'qn':        _series_to_list(qn_all[stem].astype(float)),
                'qp':        _series_to_list(qp_all[stem].astype(float)),
                'qy':        _series_to_list(qy_all[stem].astype(float)),
                'etr':       _series_to_list(etr_all[stem].astype(float)),
                'par':       _series_to_list(par_all[stem].astype(float)),
            }

    # Summary scalar params: Fv/Fm at t=0, Rfd = (Fp-Fs)/Fs computed from raw trace max/last
    summary_params = {}
    for stem in file_stems:
        fv_fm = _safe(float(fv_all[stem].iloc[0]) / float(fm_all[stem].iloc[0])) if stem in fv_all.columns else None
        summary_params[stem] = {
            'fv_fm': fv_fm,
            'npq_max': _safe(float(npq_fm[stem].max())) if stem in npq_fm.columns else None,
        }

    # ── State transitions ───────────────────────────────────────────────────
    st_result, st_phases_meta, has_st = {}, [], False
    try:
        par_first = par_all.iloc[:, 1].fillna(0).astype(float).tolist() \
                    if par_all.shape[1] > 1 else []
        phases    = _detect_par_phases(par_first)
        fm_dict   = {stem: params[stem]['fm'] for stem in file_stems if stem in params}
        st_phases_meta = [
            dict(label=ph['label'], type=ph['type'], par=ph['par'],
                 t_start=_safe(time_s[int(ph['start_idx'])] if int(ph['start_idx']) < len(time_s) else None),
                 t_end  =_safe(time_s[int(ph['end_idx'])]   if int(ph['end_idx'])   < len(time_s) else None))
            for ph in phases
        ]
        st_result = _calc_st_phases(file_stems, fm_dict, time_s, phases, include_d1=False)
        has_st    = bool(st_result)
    except Exception:
        pass

    return {
        'fluorometer': 'MC-PAM',
        'mode': 'parameters',
        'protocol': None,
        'files': file_stems,
        'file_stem': file_stems[0] if file_stems else '',
        'time_unit': 's',
        'raw_time': time_s,
        'raw_traces': {stem: params[stem]['ft'] for stem in file_stems if stem in params},
        'has_params': True,
        'param_time': time_s,
        'params': params,
        'has_summary': True,
        'summary': summary_params,
        'has_state_transitions': has_st,
        'st_include_d1': False,
        'st_phases_meta': st_phases_meta,
        'state_transitions': st_result,
    }


def _process_aquapen(files, protocol_key, upload_folder):
    """
    Read AquaPen NPQ .txt files and compute all fluorescence parameters.
    Returns unified JSON payload or raises ValueError.
    """
    proto = AQUAPEN_PROTOCOLS[protocol_key]
    summary_combined = pd.DataFrame()
    file_stems = []

    for i, file in enumerate(files):
        fname_full = secure_filename(file.filename or '')
        fpath = os.path.join(upload_folder, fname_full).replace('\\', '/')
        file.save(fpath)
        try:
            with open(fpath, 'r') as fh:
                lines = fh.readlines()
        finally:
            os.remove(fpath)

        df = pd.DataFrame(lines)
        df = df[0].str.split('\t', expand=True)
        # Drop last column (contains '\n')
        df = df.iloc[:, :-1] if df.shape[1] > 1 else df

        stem = str.lower(os.path.splitext(file.filename or '')[0])
        file_stems.append(stem)

        if i == 0:
            summary_combined = df.copy()
            summary_combined.rename(columns={
                summary_combined.columns[0]: 'time_us',
                summary_combined.columns[1]: stem,
            }, inplace=True)
            # Drop extra columns beyond 2
            summary_combined = summary_combined.iloc[:, :2]
        else:
            col = df.iloc[:, 1:2].copy()
            col.rename(columns={col.columns[0]: stem}, inplace=True)
            summary_combined = pd.concat([summary_combined, col.reset_index(drop=True)], axis=1)

    if summary_combined.empty:
        raise ValueError('No valid AquaPen data found.')

    if not summary_combined['time_us'].astype(str).str.contains('NPQ').any():
        raise ValueError('Files do not contain exported parameters (NPQ rows missing). '
                         'Please re-export from AquaPen with parameters enabled.')

    # Detect protocol
    detected_proto = None
    for pkey, pdef in AQUAPEN_PROTOCOLS.items():
        if pdef['detect'](summary_combined['time_us']):
            detected_proto = pkey
            break

    if detected_proto is None:
        raise ValueError('Could not auto-detect NPQ protocol from file content.')

    if detected_proto != protocol_key:
        raise ValueError(
            f'Protocol mismatch: you selected {protocol_key} but file matches {detected_proto}.'
        )

    proto = AQUAPEN_PROTOCOLS[protocol_key]

    # ── Extract scalar parameter rows ────────────────────────────────────────
    def _get_row(label_re, exact=True):
        if exact:
            mask = summary_combined['time_us'].astype(str).str.strip() == label_re
        else:
            mask = summary_combined['time_us'].astype(str).str.contains(label_re)
        return summary_combined[mask].iloc[:, 1:] if mask.any() else pd.DataFrame()

    fo_row      = _get_row('Fo')
    fm_row      = _get_row('Fm', exact=True)
    qy_max_row  = _get_row('QY_max')
    actinic_row = _get_row('ACTINIC-Intensity', exact=False)
    actinic_row = actinic_row.reset_index(drop=True)

    # Indexes for light/dark range slicing
    idx = {}
    for lbl in ['Fm_L1', 'Fm_Lss', 'Fm_D1', proto['last_fm_dark'],
                'NPQ_L1', 'NPQ_Lss', 'NPQ_D1', proto['last_npq_dark'],
                'Qp_L1', 'Qp_Lss', 'Qp_D1', proto['last_qp_dark'],
                'QY_L1', 'QY_Lss', 'QY_D1', proto['last_qy_dark']]:
        rows = summary_combined.index[summary_combined['time_us'].astype(str).str.strip() == lbl].tolist()
        idx[lbl] = rows[0] if rows else None

    def _safe_slice(start_lbl, end_lbl):
        s, e = idx.get(start_lbl), idx.get(end_lbl)
        if s is None or e is None:
            return pd.DataFrame()
        return summary_combined.iloc[s:e+1]

    fm_light = _safe_slice('Fm_L1', 'Fm_Lss')
    fm_dark  = _safe_slice('Fm_D1', proto['last_fm_dark'])
    npq_light = _safe_slice('NPQ_L1', 'NPQ_Lss')
    npq_dark  = _safe_slice('NPQ_D1', proto['last_npq_dark'])
    qp_light  = _safe_slice('Qp_L1', 'Qp_Lss')
    qp_dark   = _safe_slice('Qp_D1', proto['last_qp_dark'])
    qy_light  = _safe_slice('QY_L1', 'QY_Lss')
    qy_dark   = _safe_slice('QY_D1', proto['last_qy_dark'])

    # ── Build Fm_prime timeseries ──────────────────────────────────────────
    timing_fm  = pd.Series(proto['fm_time'], dtype='int64')
    timing_ft  = pd.Series(proto['ft_time'], dtype='int64')
    timing_npq = pd.Series(proto['npq_time'], dtype='int64')

    fm_prime_data = pd.concat([fm_light, fm_dark]).iloc[:, 1:].reset_index(drop=True)
    fm_prime_data = pd.concat([fm_row.reset_index(drop=True), fm_prime_data]).reset_index(drop=True)
    fm_prime_data = fm_prime_data.astype(float)

    fm_max = fm_prime_data.max()
    fm_first = fm_prime_data.iloc[0].to_numpy(dtype=float)

    # ── Compute all parameters per sample ──────────────────────────────────
    # Numeric-only rows for raw trace
    is_numeric = summary_combined['time_us'].str.strip().str.isnumeric()
    raw_num = summary_combined[is_numeric].astype('int64')

    # Ft (nearest time lookup)
    ft_lookup = pd.merge_asof(
        pd.DataFrame({'time_us': proto['ft_time']}, dtype='int64'),
        raw_num,
        on='time_us',
        direction='nearest'
    ).iloc[:, 1:]

    qy_data = pd.concat([qy_light, qy_dark]).iloc[:, 1:]
    qy_data = pd.concat([qy_max_row.reset_index(drop=True), qy_data]).reset_index(drop=True)
    qy_data = qy_data.astype(float)

    qp_data = pd.concat([qp_light, qp_dark]).iloc[:, 1:].reset_index(drop=True).astype(float)

    npq_fm_data = pd.concat([npq_light, npq_dark]).iloc[:, 1:].reset_index(drop=True).astype(float)

    # Fv = Fm' - Ft
    fv_data = fm_prime_data.values - ft_lookup.values
    # qN = (Fm - Fm') / (Fm - Fo)
    fo_first = fo_row.iloc[0].to_numpy(dtype=float)
    qn_data = (fm_first - fm_prime_data.values) / (fm_first - fo_first)
    # NPQ using Fm_max = (Fm_max - Fm') / Fm'
    npq_fmmax_data = (fm_max.values - fm_prime_data.values) / fm_prime_data.values

    params = {}
    for j, stem in enumerate(file_stems):
        col_name = summary_combined.columns[j + 1] if j + 1 < len(summary_combined.columns) else stem
        params[stem] = {
            'ft':  _series_to_list(pd.to_numeric(ft_lookup.iloc[:, j], errors='coerce')),
            'fm':  _series_to_list(pd.to_numeric(fm_prime_data.iloc[:, j], errors='coerce')),
            'fv':  _series_to_list(pd.to_numeric(pd.Series(fv_data[:, j]), errors='coerce')),
            'npq': _series_to_list(pd.to_numeric(npq_fm_data.iloc[:, j], errors='coerce')),
            'npq_fmmax': _series_to_list(pd.to_numeric(pd.Series(npq_fmmax_data[:, j]), errors='coerce')),
            'qn':  _series_to_list(pd.to_numeric(pd.Series(qn_data[:, j]), errors='coerce')),
            'qp':  _series_to_list(pd.to_numeric(qp_data.iloc[:, j], errors='coerce')),
            'qy':  _series_to_list(pd.to_numeric(qy_data.iloc[:, j], errors='coerce')),
            'etr': _series_to_list(pd.to_numeric(qy_data.iloc[:, j], errors='coerce') *
                                   float(actinic_row.iloc[0, j])  # type: ignore[arg-type]
                                   if not actinic_row.empty and j < actinic_row.shape[1] else
                                   pd.Series([0.0] * len(qy_data))),
            'par': None,
        }

    # Scalar summaries (used for Export to Statistics)
    summary_scalars = {}
    for j, stem in enumerate(file_stems):
        fv_fm = None
        rfd_val = None
        actinic_val = None
        try:
            fm0 = float(fm_prime_data.iloc[0, j])  # type: ignore[arg-type]
            ft0 = float(ft_lookup.iloc[0, j])  # type: ignore[arg-type]
            if fm0 > 0:
                fv_fm = _safe((fm0 - ft0) / fm0)
        except Exception:
            pass
        try:
            if not actinic_row.empty and j < actinic_row.shape[1]:
                actinic_val = _safe(float(actinic_row.iloc[0, j]))  # type: ignore[arg-type]
        except Exception:
            pass
        try:
            fp_val = float(raw_num.iloc[:, j + 1].max())  # type: ignore[arg-type]
            fs_val = float(raw_num.iloc[-1, j + 1])  # type: ignore[arg-type]
            if fs_val > 0:
                rfd_val = _safe((fp_val - fs_val) / fs_val)
        except Exception:
            pass
        summary_scalars[stem] = {
            'fv_fm': fv_fm,
            'rfd':   rfd_val,
            'actinic_intensity': actinic_val,
        }

    # Raw fluorescence trace (numeric only)
    raw_time = _series_to_list(raw_num['time_us'].astype(float))
    raw_traces = {}
    for j, stem in enumerate(file_stems):
        raw_traces[stem] = _series_to_list(raw_num.iloc[:, j + 1].astype(float))

    # ── State transitions ───────────────────────────────────────────────────
    st_result, st_phases_meta, has_st = {}, [], False
    try:
        # n_light = count of Fm_L* labels (excludes initial Fm and Fm_D*)
        n_light    = sum(1 for l in proto['fm_labels'] if l.startswith('Fm_L'))
        n_fm_total = len(timing_fm)
        # Times in seconds (for fitting and reporting)
        t_s = [float(t) / 1e6 for t in timing_fm]

        dark_start = n_light + 1          # index of Fm_D1
        ap_phases  = [
            dict(label='AL1', type='light', par=None,
                 start_idx=1, end_idx=n_light),
        ]
        if dark_start < n_fm_total:
            ap_phases.append(dict(label='D1', type='dark', par=0,
                                  start_idx=dark_start, end_idx=n_fm_total - 1))

        fm_dict = {}
        for j2, stem in enumerate(file_stems):
            if j2 < fm_prime_data.shape[1]:
                fm_dict[stem] = list(fm_prime_data.iloc[:, j2].astype(float))

        st_phases_meta = [
            dict(label=ph['label'], type=ph['type'], par=ph.get('par'),
                 t_start=_safe(t_s[int(ph['start_idx'])] if ph['start_idx'] is not None and int(ph['start_idx']) < len(t_s) else None),  # type: ignore[arg-type]
                 t_end  =_safe(t_s[int(ph['end_idx'])]   if ph['end_idx']   is not None and int(ph['end_idx'])   < len(t_s) else None))  # type: ignore[arg-type]
            for ph in ap_phases
        ]
        st_result = _calc_st_phases(file_stems, fm_dict, t_s, ap_phases, include_d1=False)
        has_st    = bool(st_result)
    except Exception:
        pass

    return {
        'fluorometer': 'AquaPen',
        'mode': 'parameters',
        'protocol': protocol_key,
        'files': file_stems,
        'file_stem': file_stems[0] if file_stems else '',
        'time_unit': 'us',
        'raw_time': raw_time,
        'raw_traces': raw_traces,
        'has_params': True,
        'param_time': _series_to_list(timing_fm.astype(float)),
        'param_time_npq': _series_to_list(timing_npq.astype(float)),
        'param_labels': {
            'fm': proto['fm_labels'],
            'ft': proto['ft_labels'],
            'fv': proto['fv_labels'],
            'npq': proto['npq_labels'],
            'qp': proto['qp_labels'],
            'qy': proto['qy_labels'],
            'etr': proto['etr_labels'],
        },
        'params': params,
        'has_summary': True,
        'summary': summary_scalars,
        'has_state_transitions': has_st,
        'st_include_d1': False,
        'st_phases_meta': st_phases_meta,
        'state_transitions': st_result,
    }


# ─── routes ──────────────────────────────────────────────────────────────────

@slow_kin_data_analysis.route('/slow_kin_data_analysis', methods=['GET'])
def analyze_slow_kin_data():
    return render_template('slow_kin_data_analysis.html')


@slow_kin_data_analysis.route('/api/slow_kin_process', methods=['POST'])
def slow_kin_process():
    upload_folder = UPLOAD_FOLDER
    if not os.path.isdir(upload_folder):
        os.mkdir(upload_folder)

    if 'NPQ_files' not in request.files:
        return jsonify({'status': 'error', 'message': 'No files received.'}), 400

    files = request.files.getlist('NPQ_files')
    if not files or secure_filename(files[0].filename or '') == '':
        return jsonify({'status': 'error', 'message': 'Please select one or more files.'}), 400

    if len(files) > 50:
        return jsonify({'status': 'error', 'message': 'Maximum 50 files allowed.'}), 400

    fluorometer = request.form.get('fluorometer', '')
    reduce_data = request.form.get('reduce_data') == 'true'

    try:
        if fluorometer == 'MC-PAM':
            file_type = request.form.get('mc_pam_file_type', 'raw_data')
            if file_type == 'raw_data':
                payload = _process_mcpam_raw(files, reduce_data)
            else:
                payload = _process_mcpam_params(files)

        elif fluorometer == 'AquaPen':
            protocol_key = request.form.get('aquapen_protocol', 'NPQ1')
            if protocol_key not in AQUAPEN_PROTOCOLS:
                return jsonify({'status': 'error', 'message': f'Unknown protocol: {protocol_key}'}), 400
            payload = _process_aquapen(files, protocol_key, upload_folder)

        else:
            return jsonify({'status': 'error', 'message': f'Unknown fluorometer: {fluorometer}'}), 400

    except ValueError as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Processing error: {e}'}), 500

    _cleanup_old_files(upload_folder)

    payload['status'] = 'success'
    return jsonify(payload)


@slow_kin_data_analysis.route('/api/slow_kin_st_refit', methods=['POST'])
def slow_kin_st_refit():
    """
    Refit state transitions with user-supplied phase windows or updated include_d1.
    Body JSON:
      { include_d1: bool,
        phases: [{label, type, par,
                  files_data: {stem: {t: [s…], fm: [a.u.…]}}}] }
    Response JSON:
      { status, state_transitions: {…}, st_phases_meta: […] }
    """
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'status': 'error', 'message': 'No data.'}), 400

        include_d1   = bool(data.get('include_d1', False))
        phases_input = data.get('phases', [])

        all_stems = []
        for ph in phases_input:
            for s in ph.get('files_data', {}).keys():
                if s not in all_stems:
                    all_stems.append(s)

        st_result      = {stem: [] for stem in all_stems}
        st_phases_meta = []

        for ph in phases_input:
            label      = ph.get('label', '')
            ph_type    = ph.get('type', 'light')
            par        = ph.get('par')
            files_data = ph.get('files_data', {})

            all_t = [v for fd in files_data.values() for v in fd.get('t', [])]
            st_phases_meta.append(dict(
                label=label, type=ph_type, par=par,
                t_start=_safe(min(all_t)) if all_t else None,
                t_end  =_safe(max(all_t)) if all_t else None,
            ))

            for stem in all_stems:
                fd    = files_data.get(stem, {})
                t_seg = list(fd.get('t', []))
                y_seg = list(fd.get('fm', []))
                if ph_type == 'dark' and not include_d1:
                    t_seg = t_seg[1:]; y_seg = y_seg[1:]
                fit = _fit_single_exp(t_seg, y_seg)
                st_result[stem].append(dict(
                    label=label, type=ph_type, par=par,
                    fm_vals=[_safe(v) for v in y_seg],
                    t_vals =[_safe(v) for v in t_seg],
                    **fit,
                ))

        return jsonify({'status': 'success',
                        'state_transitions': st_result,
                        'st_phases_meta': st_phases_meta})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@slow_kin_data_analysis.route('/api/slow_kin_export', methods=['POST'])
def slow_kin_export():
    """
    Build and return an .xlsx file from the JSON result payload.
    Receives the full analysis result as JSON body.
    """
    upload_folder = UPLOAD_FOLDER
    if not os.path.isdir(upload_folder):
        os.mkdir(upload_folder)

    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'status': 'error', 'message': 'No data received.'}), 400

        fluorometer = data.get('fluorometer', 'unknown')
        mode        = data.get('mode', '')
        files       = data.get('files', [])
        file_stem   = data.get('file_stem', 'slow_kin') or 'slow_kin'
        raw_time    = data.get('raw_time', [])
        raw_traces  = data.get('raw_traces', {})
        has_params  = data.get('has_params', False)
        param_time  = data.get('param_time', [])
        param_time_npq = data.get('param_time_npq', param_time)
        params      = data.get('params', {})
        has_summary = data.get('has_summary', False)
        summary     = data.get('summary', {})
        param_labels = data.get('param_labels', {})

        wb = Workbook()
        default_sheet = wb.active  # remove after all real sheets are added

        def _write_sheet(wb, title, time_list, series_dict, time_label='Time', labels=None):
            ws = wb.create_sheet(title=title[:31])
            header = [time_label] + list(series_dict.keys())
            ws.append(header)
            n = max(len(v) for v in series_dict.values()) if series_dict else len(time_list)
            for row_i in range(n):
                t_val = time_list[row_i] if row_i < len(time_list) else None
                row = [t_val]
                for vals in series_dict.values():
                    row.append(vals[row_i] if row_i < len(vals) else None)
                ws.append(row)
            # Optionally add label column
            if labels:
                ws.insert_cols(1)
                ws.cell(1, 1, 'Label')
                for ri, lbl in enumerate(labels, start=2):
                    ws.cell(ri, 1, lbl)

        # Raw fluorescence
        _write_sheet(wb, 'Raw fluorescence',
                     raw_time,
                     {f: raw_traces[f] for f in files if f in raw_traces},
                     time_label='Time')

        if has_params and params:
            param_keys = [
                ('ft', 'Ft', param_time, param_labels.get('ft')),
                ('fm', "Fm\u2032", param_time, param_labels.get('fm')),
                ('fv', 'Fv', param_time, param_labels.get('fv')),
                ('npq', 'NPQ (Fm)', param_time_npq, param_labels.get('npq')),
                ('npq_fmmax', 'NPQ (Fm_max)', param_time, param_labels.get('npq')),
                ('qn', 'qN', param_time, param_labels.get('fm')),
                ('qp', 'qP', param_time_npq, param_labels.get('qp')),
                ('qy', 'Y(II)', param_time, param_labels.get('qy')),
                ('etr', 'ETR', param_time, param_labels.get('etr')),
            ]
            for pkey, sheet_name, t_list, lbls in param_keys:
                series = {f: params[f][pkey] for f in files
                          if f in params and pkey in params[f] and params[f][pkey] is not None}
                if series:
                    _write_sheet(wb, sheet_name, t_list or [], series,
                                 time_label='Time', labels=lbls)

        if has_summary and summary:
            ws = wb.create_sheet(title='Summary')
            # Collect all scalar keys
            all_keys = set()
            for v in summary.values():
                all_keys.update(v.keys())
            header = ['Sample'] + sorted(all_keys)
            ws.append(header)
            for fname, vals in summary.items():
                row = [fname] + [vals.get(k) for k in sorted(all_keys)]
                ws.append(row)

        # State-transition sheet
        st_data = data.get('state_transitions', {})
        if st_data:
            ws_st = wb.create_sheet(title='State Transitions')
            ws_st.append(['Sample', 'Phase', 'PAR (µmol m⁻² s⁻¹)', 'n points',
                          'ΔFm\' (%)', 'τ (s)', 't½ (s)', 'R²', 'Notes'])
            for fname, ph_list in st_data.items():
                for ph in (ph_list or []):
                    note = ('low confidence' if ph.get('low_confidence') else
                            'insufficient data' if ph.get('insufficient_data') else
                            'fit failed' if not ph.get('fit_ok') and not ph.get('insufficient_data') else '')
                    ws_st.append([
                        fname,
                        ph.get('label'),
                        ph.get('par'),
                        ph.get('n_points'),
                        ph.get('delta_fm_pct'),
                        ph.get('tau'),
                        ph.get('half_time'),
                        ph.get('r_sq'),
                        note,
                    ])

        # Charts sheet (images captured from Chart.js on the client)
        charts = data.get('charts', [])
        if charts:
            ws_ch = wb.create_sheet(title='Charts')
            row = 1
            for c in charts:
                url = c.get('data_url', '')
                if not url or ',' not in url:
                    continue
                b64 = url.split(',', 1)[1]
                if not b64:
                    continue
                try:
                    img_bytes = base64.b64decode(b64)
                except Exception:
                    continue
                if len(img_bytes) < 8:
                    continue
                img_buf = io.BytesIO(img_bytes)
                try:
                    xl_img = XLImage(img_buf)
                except Exception:
                    continue
                img_buf.seek(0)
                TARGET_W = 700
                orig_w = xl_img.width or 0
                orig_h = xl_img.height or 400
                if orig_w > 0:
                    scale = TARGET_W / orig_w
                    xl_img.width  = TARGET_W
                    xl_img.height = round(orig_h * scale)
                else:
                    xl_img.width, xl_img.height = TARGET_W, 400
                title = c.get('title', '')
                if title:
                    ws_ch.cell(row=row, column=1, value=title)
                    row += 1
                xl_img.anchor = f'A{row}'
                ws_ch.add_image(xl_img)
                row += round(xl_img.height / 20) + 2

        if default_sheet in wb.worksheets:
            wb.remove(default_sheet)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{file_stem}_results.xlsx',
        )

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
