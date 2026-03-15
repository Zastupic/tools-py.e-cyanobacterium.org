from flask import Blueprint, render_template, request, jsonify, send_file
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches # Required for Ellipse
import matplotlib.transforms as transforms # Required for Ellipse rotation
import seaborn as sns
import io
import base64
import traceback
import math
import itertools
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from pyopls import OPLS as PYOPLS
from scipy import stats
from statsmodels.formula.api import ols
import statsmodels.api as sm
from statsmodels.stats.multicomp import pairwise_tukeyhsd, MultiComparison
from statsmodels.stats.multitest import multipletests
from itertools import combinations
from typing import Any

stats_bp = Blueprint('statistics', __name__)

MAX_DATA_ROWS = 100
MAX_DATA_COLUMNS = 100

def _validate_data_limits(data):
    """Validate data does not exceed row/column limits. Returns (True, None) or (False, error_message)."""
    if not data:
        return True, None
    rows = len(data)
    first = data[0]
    cols = len(first) if isinstance(first, (dict, list, tuple)) else 0
    if rows > MAX_DATA_ROWS:
        return False, f"Data exceeds maximum allowed rows ({MAX_DATA_ROWS}). Your data has {rows} rows. Please reduce the dataset."
    if cols > MAX_DATA_COLUMNS:
        return False, f"Data exceeds maximum allowed columns ({MAX_DATA_COLUMNS}). Your data has {cols} columns. Please reduce the dataset."
    return True, None

def _outlier_mask_15iqr(series):
    """Return a boolean array: True where value is outside 1.5×IQR from the box (standard boxplot rule)."""
    q1, q3 = np.quantile(series.dropna(), np.array([0.25, 0.75]))
    iqr = q3 - q1
    if iqr <= 0:
        return pd.Series(False, index=series.index)
    lower = q1 - 1.5 * iqr
    upper = q3 + 1.5 * iqr
    return (series < lower) | (series > upper)


def _box_stats_per_group(clean_df, group_col, value_col):
    """Return list of {group, q1, q3, median, lowerfence, upperfence} using 1.5×IQR rule (same as _outlier_mask_15iqr).
    Fences are exactly Q1 - 1.5*IQR and Q3 + 1.5*IQR so Plotly whiskers match backend outlier logic."""
    out = []
    for grp, sub in clean_df.groupby(group_col, sort=False):
        vals = sub[value_col].dropna()
        if len(vals) == 0:
            continue
        q1, median, q3 = np.quantile(vals, np.array([0.25, 0.5, 0.75]))
        q1, median, q3 = float(q1), float(median), float(q3)
        iqr = q3 - q1
        if iqr <= 0:
            lowerfence = upperfence = median
        else:
            lowerfence = q1 - 1.5 * iqr
            upperfence = q3 + 1.5 * iqr
        out.append({
            "group": str(grp),
            "n": int(len(vals)),
            "q1": q1,
            "q3": q3,
            "median": median,
            "lowerfence": lowerfence,
            "upperfence": upperfence,
        })
    return out

def _build_plot_data(clean_df, group_col, value_col, row_id_col="row_id"):
    """Build list of {group, value, row_id, factor_label, is_outlier} for frontend hover. Outliers = outside 1.5×IQR per group."""
    if row_id_col not in clean_df.columns:
        clean_df = clean_df.copy()
        clean_df[row_id_col] = np.arange(1, len(clean_df) + 1)
    plot_data = []
    for grp, sub in clean_df.groupby(group_col, sort=False):
        vals = sub[value_col]
        outlier = _outlier_mask_15iqr(vals)
        factor_label = str(grp)
        for idx in sub.index:
            row = sub.loc[idx]
            rid = row[row_id_col] if row_id_col in row.index else idx
            if pd.isna(rid):
                rid = int(idx) if isinstance(idx, (int, np.integer)) else idx
            plot_data.append({
                "group": str(grp),
                "value": float(row[value_col]),
                "row_id": int(rid),
                "factor_label": factor_label,
                "is_outlier": bool(outlier.iloc[sub.index.get_loc(idx)]) if idx in sub.index else False,
            })
    return plot_data

def _sanitize(obj):
    """Replace NaN/Inf with None so JSON serialization succeeds."""
    if isinstance(obj, dict):
        return {k: _sanitize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize(v) for v in obj]
    if isinstance(obj, float):
        if obj != obj or obj == float('inf') or obj == float('-inf'):
            return None
        return obj
    if isinstance(obj, (np.floating, np.integer)):
        v = float(obj)
        if v != v or v == float('inf') or v == float('-inf'):
            return None
        return v
    return obj

def get_plot_base64():
    img = io.BytesIO()
    plt.savefig(img, format='png', bbox_inches='tight', dpi=150)
    img.seek(0)
    plt.close()
    return base64.b64encode(img.getvalue()).decode('utf-8')


# ── Shared openpyxl helpers ───────────────────────────────────────────────────
_XL_HDR_FONT   = Font(bold=True, color='FFFFFF', size=10)
_XL_THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)
_XL_CENTER_AL = Alignment(horizontal='center', vertical='center', wrap_text=True)
_XL_FILLS = {
    'blue':   PatternFill('solid', fgColor='2563EB'),
    'green':  PatternFill('solid', fgColor='16A34A'),
    'orange': PatternFill('solid', fgColor='D97706'),
    'purple': PatternFill('solid', fgColor='7C3AED'),
    'teal':   PatternFill('solid', fgColor='0F766E'),
    'grey':   PatternFill('solid', fgColor='374151'),
    'note':   PatternFill('solid', fgColor='FEF9C3'),
    'pass':   PatternFill('solid', fgColor='DCFCE7'),
    'fail':   PatternFill('solid', fgColor='FEE2E2'),
    'light':  PatternFill('solid', fgColor='F3F4F6'),
}

def _xl_auto_width(ws, max_w=45):
    """Auto-fit column widths (openpyxl worksheet)."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_w)

def _xl_style_header(ws, row_num, fill):
    """Apply bold white header style to a row."""
    for cell in ws[row_num]:
        cell.font  = _XL_HDR_FONT
        cell.fill  = fill
        cell.alignment = _XL_CENTER_AL
        cell.border = _XL_THIN_BORDER

def _xl_border_row(ws, row_num, n_cols):
    """Apply thin border to n_cols cells in a row."""
    for c in range(1, n_cols + 1):
        ws.cell(row_num, c).border = _XL_THIN_BORDER

def _xl_add_image(ws, img_b64, cell_addr):
    """Decode base64 PNG and add to worksheet at cell_addr. Returns True on success."""
    try:
        img_bytes = base64.b64decode(img_b64)
        xl_img = XLImage(io.BytesIO(img_bytes))
        ws.add_image(xl_img, cell_addr)
        return True
    except Exception:
        return False
# ─────────────────────────────────────────────────────────────────────────────

@stats_bp.route('/statistics', methods=['GET'])
def statistics_page():
    return render_template("statistics.html") # type: ignore

@stats_bp.route('/run-tests', methods=['POST'])
def run_tests():
    try:
        request_data = request.get_json()
        data = request_data.get('data', [])
        ok, err = _validate_data_limits(data)
        if not ok:
            return jsonify({"error": err}), 400
        df = pd.DataFrame(data)
        factors = request_data.get('factors', [])
        selected_vars = request_data.get('target_columns', [])

        test_results = []

        for var in selected_vars:
            temp_df = df.copy()
            temp_df[var] = pd.to_numeric(temp_df[var], errors='coerce')
            clean_df = temp_df.dropna(subset=[var]).copy()

            if clean_df.empty: continue

            # 1. Prepare Group Labels
            if factors:
                for f in factors:
                    clean_df[f] = clean_df[f].astype(str).replace(['nan', 'None'], "N/A")
                clean_df['Group'] = clean_df[factors].agg(' | '.join, axis=1)
            else:
                clean_df['Group'] = 'All Data'

            # 2. Normality and group data for Shapiro-Wilk / Levene tests
            group_data = []
            shapiro_results = []

            unique_groups = clean_df['Group'].unique().tolist()

            for g_name in unique_groups:
                group_vals = clean_df[clean_df['Group'] == g_name][var].astype(float)
                if len(group_vals) >= 3:
                    stat, p = stats.shapiro(group_vals)
                    is_normal = bool(p > 0.05)
                    group_data.append(group_vals.values)
                    shapiro_results.append({
                        "group": g_name, "stat": float(stat), "p": float(p), "is_normal": is_normal
                    })

            # 3. Calculate Levene's Test
            l_stat, l_p, is_homo = None, None, None
            if len(group_data) > 1:
                l_stat, l_p = stats.levene(*group_data)
                is_homo = bool(l_p > 0.05)

            # 4. Build plot data for Plotly rendering on the frontend
            plot_data = _build_plot_data(clean_df, "Group", var)

            box_stats = _box_stats_per_group(clean_df, "Group", var)
            result_entry = {
                "variable": var,
                "plot_data": plot_data,
                "box_stats": box_stats,
                "shapiro": shapiro_results,
                "levene": {"stat": float(l_stat) if l_stat else None,
                           "p": float(l_p) if l_p else None,
                           "is_homogeneous": is_homo}
            }
            test_results.append(result_entry)

        return jsonify({"results": _sanitize(test_results)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@stats_bp.route('/run-statistics', methods=['POST'])
def run_analysis():
    try:
        request_data = request.get_json()
        data = request_data.get('data', [])
        ok, err = _validate_data_limits(data)
        if not ok:
            return jsonify({"error": err}), 400
        df = pd.DataFrame(data)
        factors = request_data.get('factors', [])
        selected_vars = request_data.get('target_columns', [])

        if not selected_vars:
            cols = [c for c in df.columns.tolist() if c != "row_id"]
            return jsonify({
                "all_columns": cols,
                "variables": cols
            })

        results = []
        for var in selected_vars:
            keep_cols = [c for c in factors + [var] if c in df.columns]
            if "row_id" in df.columns:
                keep_cols = ["row_id"] + [c for c in keep_cols if c != "row_id"]
            temp_df = df[keep_cols].copy()
            temp_df[var] = pd.to_numeric(temp_df[var], errors='coerce')
            clean_df = temp_df.dropna(subset=[var]).copy()

            # Also drop rows where any factor is empty/NaN
            if factors:
                clean_df = clean_df.dropna(subset=factors).copy()
                clean_df = clean_df[~clean_df[factors].isin(['', 'nan', 'None', 'NaN']).any(axis=1)].copy()

            if clean_df.empty:
                continue

            # 1. Numeric Sorting for all factors
            if factors:
                sort_cols = []
                for f in factors:
                    # Create temporary numeric columns for logical sorting
                    clean_df[f + '_sort'] = pd.to_numeric(clean_df[f], errors='coerce')
                    sort_cols.append(f + '_sort')

                clean_df = clean_df.sort_values(by=sort_cols)

                # Convert original factors to string for display
                for f in factors:
                    clean_df[f] = clean_df[f].astype(str).replace(['nan', 'None'], "N/A")

            # Summary Stats (grouped by original factors)
            summary_df = clean_df.groupby(factors)[var].agg(['count', 'mean', 'std']).reset_index()
            summary = summary_df.replace({np.nan: None}).to_dict(orient='records')

            # Build plot_data by combined group (all factors) so frontend splits by both factors
            if factors:
                clean_df = clean_df.copy()
                clean_df['Group'] = clean_df[factors].agg(' | '.join, axis=1)
                group_col = 'Group'
            else:
                group_col = None
            if group_col:
                plot_data = _build_plot_data(clean_df, group_col, var)
                box_stats = _box_stats_per_group(clean_df, group_col, var)
            else:
                plot_data = []
                box_stats = []
            results.append({
                "variable": var,
                "summary": summary,
                "plot_data": plot_data,
                "box_stats": box_stats,
            })

        return jsonify({"mode": "results", "factors": factors, "results": _sanitize(results)})
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@stats_bp.route('/export-excel', methods=['POST'])
def export_excel():
    try:
        request_data = request.get_json()
        results = request_data.get('results', [])
        factors = request_data.get('factors', [])
        # Accept plotly_captures from JS: [{ id: 'plot-viz-0-VarName', image: '<base64>' }]
        plotly_captures = request_data.get('plotly_captures', [])
        # Build a map from plot div id → base64 image for quick lookup
        plotly_map = {pc['id']: pc['image'] for pc in plotly_captures if pc.get('id') and pc.get('image')}

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for res in results:
                sheet_name = "".join([c for c in str(res['variable']) if c.isalnum() or c==' '])[:31]
                df_var = pd.DataFrame(res['summary'])
                df_var.to_excel(writer, sheet_name=sheet_name, index=False)

                ws = writer.sheets[sheet_name]

                # Prefer the Plotly-captured image; fall back to matplotlib backend plot_url
                img_b64 = None
                if plotly_map:
                    # Match by variable name in the plot div id (e.g. 'plot-viz-0-VarName')
                    safe_var = str(res['variable']).replace(r'\W', '_')
                    for k, v in plotly_map.items():
                        if res['variable'].replace(' ', '_') in k or res['variable'] in k:
                            img_b64 = v
                            break
                if img_b64 is None:
                    img_b64 = res.get('plot_url')

                if img_b64:
                    try:
                        img_data = base64.b64decode(img_b64)
                        img = XLImage(io.BytesIO(img_data))
                        ws.add_image(img, f"A{len(df_var) + 4}")
                    except Exception:
                        pass

            for sheetname in writer.sheets:
                _xl_auto_width(writer.sheets[sheetname])

        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='Box_plots.xlsx')
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def confidence_ellipse(x, y, ax, n_std=2.0, facecolor: Any = 'none', **kwargs):
    """
    Utility to create a covariance confidence ellipse.
    """
    if x.size != y.size:
        raise ValueError("x and y must be the same size")

    cov = np.cov(x, y)
    pearson = cov[0, 1] / np.sqrt(cov[0, 0] * cov[1, 1])

    # Using a special case to obtain the eigenvalues of this two-dimensional dataset.
    ell_radius_x = np.sqrt(1 + pearson)
    ell_radius_y = np.sqrt(1 - pearson)

    ellipse = patches.Ellipse((0, 0), width=ell_radius_x * 2, height=ell_radius_y * 2,
                              facecolor=facecolor, **kwargs)

    # Calculating the standard deviation of x from the squareroot of the variance
    scale_x = np.sqrt(cov[0, 0]) * n_std
    mean_x = np.mean(x)

    # calculating the standard deviation of y ...
    scale_y = np.sqrt(cov[1, 1]) * n_std
    mean_y = np.mean(y)

    transf = transforms.Affine2D().rotate_deg(45).scale(scale_x, scale_y).translate(mean_x, mean_y)
    ellipse.set_transform(transf + ax.transData)
    return ax.add_patch(ellipse)

def _nipals_pca(X, n_components, max_iter=500, tol=1e-6):
    """
    NIPALS PCA — handles missing data (NaN) natively without any imputation.
    X must be standardised (nanmean≈0, nanstd≈1 per column) but may contain NaN.
    Returns: T (n×k scores), P (p×k loadings), eigenvalues (k,), evr (k,).
    """
    X = X.copy().astype(float)
    n, p = X.shape
    T = np.zeros((n, n_components))
    P = np.zeros((p, n_components))
    total_ss = float(np.nansum(X ** 2))
    # Save the original observed mask before any deflation — needed for correct
    # explained-variance calculation (we measure SS only at originally-observed positions).
    obs_orig = ~np.isnan(X)

    for k in range(n_components):
        # Initialise score vector with the column that has the most observed values
        best_col = int(np.argmax(np.sum(~np.isnan(X), axis=0)))
        t = np.where(np.isnan(X[:, best_col]), 0.0, X[:, best_col])
        p_vec = np.zeros(p)  # initialised here so it is always bound after the loop

        for _ in range(max_iter):
            # Loading step: p_j = X[:,j]' t / t' t  (only observed rows of col j)
            p_vec = np.zeros(p)
            for j in range(p):
                obs = ~np.isnan(X[:, j])
                if obs.sum() > 0:
                    denom = float(np.dot(t[obs], t[obs]))
                    if denom > 0:
                        p_vec[j] = float(np.dot(X[obs, j], t[obs])) / denom

            p_norm = float(np.linalg.norm(p_vec))
            if p_norm < 1e-12:
                break
            p_vec /= p_norm

            # Score step: t_i = X[i,:] p / p' p  (only observed cols of row i)
            t_new = np.zeros(n)
            for i in range(n):
                obs = ~np.isnan(X[i, :])
                if obs.sum() > 0:
                    denom = float(np.dot(p_vec[obs], p_vec[obs]))
                    if denom > 0:
                        t_new[i] = float(np.dot(X[i, obs], p_vec[obs])) / denom

            t_norm = float(np.linalg.norm(t))
            if t_norm > 0 and float(np.linalg.norm(t_new - t)) / t_norm < tol:
                t = t_new
                break
            t = t_new

        T[:, k] = t
        P[:, k] = p_vec

        # Deflate: subtract the rank-1 contribution from observed positions only
        outer = np.outer(t, p_vec)
        X[obs_orig] -= outer[obs_orig]

    # Explained variance: SS of each rank-1 reconstruction at originally-observed
    # positions divided by total SS of the original standardised matrix.
    # Using ||t||² would be wrong with missing data — rows with more observed
    # variables get higher scores, inflating the estimate above 100%.
    ss_comp = np.array([
        float(np.sum(np.outer(T[:, k], P[:, k])[obs_orig] ** 2))
        for k in range(n_components)
    ])
    eigenvalues = ss_comp / max(n - 1, 1)
    evr = ss_comp / total_ss if total_ss > 0 else np.zeros(n_components)
    return T, P, eigenvalues, evr


def _em_pca(X, n_components, max_iter=200, tol=1e-5):
    """
    EM-PCA (iterative SVD / PPCA-lite) — handles missing data natively.
    Iteratively reconstructs missing values from the current PCA model until convergence.
    No imputation bias: missing values are constrained only by the low-rank structure.
    X must be standardised (nanmean≈0, nanstd≈1 per column) but may contain NaN.
    Returns: T (n×k scores), P (p×k loadings), eigenvalues (k,), evr (k,).
    """
    X = X.copy().astype(float)
    n, p = X.shape
    missing_mask = np.isnan(X)
    # Save the original observed mask and observed SS before any imputation.
    # We measure explained variance relative to the original observed data only,
    # consistent with the NIPALS fix — using the full imputed matrix as denominator
    # would underestimate evr by roughly the fraction of missing cells.
    obs_orig    = ~missing_mask
    total_ss    = float(np.nansum(X ** 2))

    # Initialise missing positions with column means (observed only)
    col_means = np.nanmean(X, axis=0)
    col_means = np.where(np.isnan(col_means), 0.0, col_means)
    X_imp = X.copy()
    for j in range(p):
        X_imp[missing_mask[:, j], j] = col_means[j]

    prev_vals = X_imp[missing_mask].copy() if missing_mask.any() else np.array([])

    for _ in range(max_iter):
        try:
            U, S, Vt = np.linalg.svd(X_imp, full_matrices=False)
        except np.linalg.LinAlgError:
            break

        k = n_components
        X_recon = (U[:, :k] * S[:k]) @ Vt[:k, :]

        if missing_mask.any():
            X_imp[missing_mask] = X_recon[missing_mask]

            curr_vals = X_imp[missing_mask]
            if prev_vals.size > 0:
                scale = float(np.linalg.norm(prev_vals)) + 1e-10
                if float(np.linalg.norm(curr_vals - prev_vals)) / scale < tol:
                    break
            prev_vals = curr_vals.copy()
        else:
            break  # No missing data — one pass is sufficient

    # Final decomposition on converged matrix
    U, S, Vt = np.linalg.svd(X_imp, full_matrices=False)
    k = n_components
    T = U[:, :k] * S[:k]           # n × k scores
    P = Vt[:k, :].T                 # p × k loadings (unit-length columns)

    # Explained variance: SS of each rank-1 reconstruction at originally-observed
    # positions only, divided by the original observed SS — same basis as NIPALS.
    # X_imp[obs_orig] == X_original[obs_orig] exactly (EM only fills missing cells),
    # so the sum of all components' ss at obs_orig equals total_ss, bounding evr ≤ 1.
    ss_comp = np.array([
        float(np.sum(np.outer(T[:, k_], P[:, k_])[obs_orig] ** 2))
        for k_ in range(n_components)
    ])
    eigenvalues = ss_comp / max(n - 1, 1)
    evr = ss_comp / total_ss if total_ss > 0 else np.zeros(k)
    return T, P, eigenvalues, evr


@stats_bp.route('/run-pca', methods=['POST'])
def run_pca():
    try:
        request_data = request.get_json()
        data = request_data.get('data', [])
        ok, err = _validate_data_limits(data)
        if not ok:
            return jsonify({"error": err}), 400
        df = pd.DataFrame(data)
        factors        = request_data.get('factors', [])
        selected_vars  = list(request_data.get('variables', []))
        pca_method     = request_data.get('pca_method', 'svd')          # 'svd' | 'nipals' | 'em_pca'
        missing_strategy  = request_data.get('missing_strategy', 'hybrid')
        missing_threshold = float(request_data.get('missing_threshold', 30)) / 100.0
        min_row_coverage  = float(request_data.get('min_row_coverage', 0)) / 100.0
        average_by_factors = request_data.get('average_by_factors', False)
        label_col = request_data.get('label_col', None) or None

        # ── 1. Numeric coercion ──────────────────────────────────────────────
        clean_df = df.copy()
        for var in selected_vars:
            clean_df[var] = pd.to_numeric(clean_df[var], errors='coerce')

        n_input_rows    = len(clean_df)
        vars_excluded   = []
        n_imputed_cells = 0
        rows_dropped    = 0
        rows_sparse_dropped = 0

        # ── 1c. Per-variable transforms (applied before imputation) ──────────
        per_var_transforms: dict = request_data.get('per_var_transforms') or {}
        if per_var_transforms:
            clean_df = _apply_per_var_transforms(clean_df, per_var_transforms, selected_vars)

        # ── 1b. Row sparsity filter (shared with correlation) ────────────────
        if min_row_coverage > 0 and selected_vars:
            n_sel = len(selected_vars)
            coverage = clean_df[selected_vars].notna().sum(axis=1) / n_sel
            before_sparse = len(clean_df)
            clean_df = clean_df[coverage >= min_row_coverage].copy()
            rows_sparse_dropped = before_sparse - len(clean_df)

        # ── 2. Missing-data handling ─────────────────────────────────────────
        native_missing = pca_method in ('nipals', 'em_pca')

        if native_missing:
            # NIPALS / EM-PCA handle NaN natively — only apply variable exclusion
            # to drop columns that are almost entirely missing (>threshold).
            n_rows = len(clean_df)
            for var in list(selected_vars):
                frac = clean_df[var].isna().sum() / n_rows if n_rows > 0 else 0
                if frac > missing_threshold:
                    vars_excluded.append(var)
            selected_vars = [v for v in selected_vars if v not in vars_excluded]
            if len(selected_vars) < 2:
                return jsonify({"error": (
                    f"Too many variables excluded by the missing-value threshold "
                    f"({missing_threshold*100:.0f}%). Only {len(selected_vars)} variable(s) "
                    "remain. Lower the threshold or deselect sparse variables."
                )}), 400
            # Drop rows where ALL selected variables are missing (uninformative rows)
            before = len(clean_df)
            clean_df = clean_df[clean_df[selected_vars].notna().any(axis=1)].copy()
            rows_dropped = before - len(clean_df)

        elif missing_strategy == 'drop_rows':
            before = len(clean_df)
            clean_df = clean_df.dropna(subset=selected_vars).copy()
            rows_dropped = before - len(clean_df)

        elif missing_strategy == 'impute_group_mean':
            if factors:
                for var in selected_vars:
                    mask = clean_df[var].isna()
                    if mask.any():
                        group_means = clean_df.groupby(factors)[var].transform('mean')
                        n_imputed_cells += int(mask.sum())
                        clean_df.loc[mask, var] = group_means[mask]
            for var in selected_vars:
                mask = clean_df[var].isna()
                if mask.any():
                    n_imputed_cells += int(mask.sum())
                    clean_df.loc[mask, var] = clean_df[var].mean()

        elif missing_strategy == 'exclude_vars':
            n_rows = len(clean_df)
            for var in list(selected_vars):
                frac = clean_df[var].isna().sum() / n_rows if n_rows > 0 else 0
                if frac > missing_threshold:
                    vars_excluded.append(var)
            selected_vars = [v for v in selected_vars if v not in vars_excluded]
            if len(selected_vars) < 2:
                return jsonify({"error": (
                    f"Too many variables excluded by the missing-value threshold "
                    f"({missing_threshold*100:.0f}%). Only {len(selected_vars)} variable(s) "
                    "remain. Lower the threshold or choose a different strategy."
                )}), 400
            before = len(clean_df)
            clean_df = clean_df.dropna(subset=selected_vars).copy()
            rows_dropped = before - len(clean_df)

        elif missing_strategy == 'knn':
            from sklearn.impute import KNNImputer
            n_neighbors = min(5, max(1, len(clean_df) - 1))
            imputer = KNNImputer(n_neighbors=n_neighbors)
            arr = clean_df[selected_vars].values
            n_imputed_cells = int(np.isnan(arr).sum())
            clean_df[selected_vars] = imputer.fit_transform(arr)

        else:  # hybrid (default)
            n_rows = len(clean_df)
            for var in list(selected_vars):
                frac = clean_df[var].isna().sum() / n_rows if n_rows > 0 else 0
                if frac > missing_threshold:
                    vars_excluded.append(var)
            selected_vars = [v for v in selected_vars if v not in vars_excluded]
            if len(selected_vars) < 2:
                return jsonify({"error": (
                    f"Too many variables excluded by the missing-value threshold "
                    f"({missing_threshold*100:.0f}%). Only {len(selected_vars)} variable(s) "
                    "remain. Lower the threshold or choose a different strategy."
                )}), 400
            if factors:
                for var in selected_vars:
                    mask = clean_df[var].isna()
                    if mask.any():
                        group_means = clean_df.groupby(factors)[var].transform('mean')
                        n_imputed_cells += int(mask.sum())
                        clean_df.loc[mask, var] = group_means[mask]
            for var in selected_vars:
                mask = clean_df[var].isna()
                if mask.any():
                    n_imputed_cells += int(mask.sum())
                    clean_df.loc[mask, var] = clean_df[var].mean()

        # Average replicates before PCA if requested
        if average_by_factors and factors:
            clean_df = clean_df.groupby(factors)[selected_vars].mean().reset_index()

        # For SVD path: drop rows still containing NaN after strategy
        if not native_missing:
            clean_df = clean_df.dropna(subset=selected_vars).copy()

        if len(clean_df) < 2:
            return jsonify({"error": "Insufficient data: PCA requires at least 2 valid samples after applying the missing-data strategy."}), 400
        if len(selected_vars) < 2:
            return jsonify({"error": "PCA requires at least 2 variables."}), 400

        # SVD-path diagnostics (not needed for native-missing methods)
        if not native_missing:
            still_nan = [v for v in selected_vars if clean_df[v].isna().any()]
            if still_nan:
                return jsonify({"error": (
                    f"Variable(s) still contain missing values after imputation — "
                    f"likely all values in a factor group are missing: {', '.join(still_nan)}. "
                    "Try a different strategy or increase the exclusion threshold."
                )}), 400
            zero_var = [v for v in selected_vars if clean_df[v].std() == 0]
            if zero_var:
                return jsonify({"error": (
                    f"Variable(s) have zero variance (all identical values): "
                    f"{', '.join(zero_var)}. Please exclude them from PCA."
                )}), 400

        # ── 3. Factor labels ─────────────────────────────────────────────────
        if factors:
            for f in factors:
                clean_df[f] = clean_df[f].astype(str).replace(['nan', 'None'], "N/A")
            hue_col = " & ".join(factors)
            clean_df[hue_col] = clean_df[factors].agg(' | '.join, axis=1)
        else:
            hue_col = None

        # ── 4. PCA decomposition ─────────────────────────────────────────────
        n_comp = min(4, len(selected_vars), len(clean_df) - 1)

        def _safe_f(x):
            f = float(x)
            return None if (f != f or f == float('inf') or f == float('-inf')) else f

        if pca_method == 'nipals':
            # Standardise using observed values only (nanmean / nanstd)
            col_means = np.nanmean(clean_df[selected_vars].values, axis=0)
            col_stds  = np.nanstd(clean_df[selected_vars].values, axis=0, ddof=1)
            col_stds  = np.where(col_stds < 1e-10, 1.0, col_stds)
            X_sc = (clean_df[selected_vars].values.astype(float) - col_means) / col_stds
            T_mat, P_mat, eigenvalues, evr = _nipals_pca(X_sc, n_comp)
            loadings = P_mat * np.sqrt(np.maximum(eigenvalues, 0))
            pca_features = T_mat

        elif pca_method == 'em_pca':
            col_means = np.nanmean(clean_df[selected_vars].values, axis=0)
            col_stds  = np.nanstd(clean_df[selected_vars].values, axis=0, ddof=1)
            col_stds  = np.where(col_stds < 1e-10, 1.0, col_stds)
            X_sc = (clean_df[selected_vars].values.astype(float) - col_means) / col_stds
            T_mat, P_mat, eigenvalues, evr = _em_pca(X_sc, n_comp)
            loadings = P_mat * np.sqrt(np.maximum(eigenvalues, 0))
            pca_features = T_mat

        else:  # svd (sklearn)
            scaler = StandardScaler()
            scaled_data = scaler.fit_transform(clean_df[selected_vars])
            pca = PCA(n_components=n_comp)
            pca_features = pca.fit_transform(scaled_data)
            loadings = pca.components_.T * np.sqrt(pca.explained_variance_)
            evr = pca.explained_variance_ratio_

        for k in range(n_comp):
            clean_df[f'PC{k + 1}'] = pca_features[:, k]

        # ── 5. NaN guard on loadings ─────────────────────────────────────────
        nan_loading_vars = [
            var for i, var in enumerate(selected_vars)
            if any(np.isnan(loadings[i, k]) for k in range(n_comp))
        ]
        if nan_loading_vars:
            return jsonify({"error": (
                f"PCA produced NaN loadings for: {', '.join(nan_loading_vars)}. "
                "These variables likely have near-zero variance or are fully collinear. "
                "Try excluding them or use a different algorithm."
            ), "nan_variables": nan_loading_vars}), 400

        loadings_list = [
            {**{"Variable": var},
             **{f'PC{k + 1}_Loading': _safe_f(loadings[i, k]) for k in range(n_comp)}}
            for i, var in enumerate(selected_vars)
        ]

        # ── 6. Scores & response ─────────────────────────────────────────────
        pc_cols = [f'PC{k + 1}' for k in range(n_comp)]
        scores = clean_df[pc_cols].copy()
        scores['group'] = clean_df[hue_col].values if hue_col else 'All samples'
        if label_col and label_col in clean_df.columns:
            scores['label'] = clean_df[label_col].astype(str).values
        else:
            scores['label'] = [str(i + 1) for i in range(len(clean_df))]

        # Cast to object dtype first so that pandas does not silently convert
        # Python None back to float NaN (which would produce bare NaN in JSON).
        scores_records    = scores.astype(object).mask(pd.isnull(scores)).to_dict(orient='records')
        pca_table_records = clean_df.astype(object).mask(pd.isnull(clean_df)).to_dict(orient='records')

        return jsonify(_make_json_safe({
            "n_samples":        len(clean_df),
            "n_input_rows":     n_input_rows,
            "n_components":     n_comp,
            "pca_method":       pca_method,
            "explained_variance": [_safe_f(v) for v in evr.tolist()],
            "loadings":         loadings_list,
            "scores":           scores_records,
            "pca_table":        pca_table_records,
            "vars_excluded":    vars_excluded,
            "n_imputed_cells":  n_imputed_cells,
            "rows_dropped":        rows_dropped,
            "rows_sparse_dropped": rows_sparse_dropped,
            "vars_used":           selected_vars,
            "hue_col":             hue_col,
        }))
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@stats_bp.route('/export-pca-excel', methods=['POST'])
def export_pca_excel():
    try:
        request_data = request.get_json()
        pca_details = request_data.get('pca_details', {})

        # 1. Create the Main Data Sheet (Original Data + PCA Scores)
        # Coordinates usually contains original data + PC1 + PC2
        df_main = pd.DataFrame(pca_details['coordinates'])

        # 2. Create the Loadings Sheet (The "Arrows" data)
        # We expect the frontend to send 'loadings' which is a list of {var: name, pc1: val, pc2: val}
        loadings_data = pca_details.get('loadings', [])
        df_loadings = pd.DataFrame(loadings_data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Results and Plot
            df_main.to_excel(writer, sheet_name='PCA Scores', index=False, startrow=5)
            ws1 = writer.sheets['PCA Scores']
            ws1['A1'] = "PCA Analysis Report"
            ws1['A2'] = f"Number of Samples: {pca_details['n_samples']}"
            ws1['A3'] = f"Explained Variance: PC1 ({pca_details['variance'][0]:.1%}), PC2 ({pca_details['variance'][1]:.1%})"

            # Embed Plot Image
            img_data = base64.b64decode(pca_details['plot_url'])
            img = XLImage(io.BytesIO(img_data))
            ws1.add_image(img, "M1")

            # Sheet 2: Variable Loadings (Arrow Coordinates)
            if not df_loadings.empty:
                df_loadings.to_excel(writer, sheet_name='Variable Loadings', index=False)
                ws2 = writer.sheets['Variable Loadings']
                # Add a note explaining what this is
                ws2.append([])
                ws2.append(["Note: These values represent the 'Arrows' on the Biplot."])
                ws2.append(["They show the correlation of each variable with the Principal Components."])

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='PCA_Full_Analysis.xlsx'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@stats_bp.route('/run-opls', methods=['POST'])
def run_opls():
    try:
        from sklearn.cross_decomposition import PLSRegression
        from sklearn.preprocessing import LabelBinarizer
        from sklearn.model_selection import KFold, LeaveOneOut
        from sklearn.metrics import r2_score as sk_r2

        request_data = request.get_json()
        data = request_data.get('data', [])
        ok, err = _validate_data_limits(data)
        if not ok:
            return jsonify({"error": err}), 400
        df = pd.DataFrame(data)
        factors = request_data.get('factors', [])
        selected_vars = request_data.get('variables', [])
        method = request_data.get('method', 'opls-da')  # 'opls-da' or 'opls'
        y_col = request_data.get('y_col')
        n_ortho = int(request_data.get('n_ortho', 1))
        compute_cv = bool(request_data.get('compute_cv', True))

        if not y_col:
            return jsonify({"error": "Response variable (Y) is required."}), 400
        if len(selected_vars) < 2:
            return jsonify({"error": "OPLS requires at least 2 predictor variables."}), 400

        # Prepare X
        clean_df = df.copy()
        for var in selected_vars:
            clean_df[var] = pd.to_numeric(clean_df[var], errors='coerce')
        clean_df = clean_df.dropna(subset=selected_vars).copy()

        if len(clean_df) < 4:
            return jsonify({"error": "Insufficient data: OPLS requires at least 4 valid samples."}), 400

        # Group labels for colouring (use factors if available, else Y column)
        if factors:
            for f in factors:
                clean_df[f] = clean_df[f].astype(str).replace(['nan', 'None'], 'N/A')
            hue_col = ' & '.join(factors)
            clean_df[hue_col] = clean_df[factors].agg(' | '.join, axis=1)
            group_labels = clean_df[hue_col].tolist()
        else:
            group_labels = clean_df[y_col].astype(str).tolist()

        # Standardise X
        scaler = StandardScaler()
        X = scaler.fit_transform(clean_df[selected_vars].values)

        # Prepare Y
        class_labels = None
        if method == 'opls-da':
            lb = LabelBinarizer()
            Y_bin: np.ndarray = np.asarray(lb.fit_transform(clean_df[y_col].astype(str)))
            y = (Y_bin.ravel() if Y_bin.shape[1] == 1 else Y_bin[:, 0]).astype(float)
            class_labels = lb.classes_.tolist()
        else:
            y = pd.to_numeric(clean_df[y_col], errors='coerce').values
            if np.isnan(y).any():
                return jsonify({"error": f"Column '{y_col}' contains non-numeric or missing values."}), 400

        n = len(clean_df)
        n_ortho = max(1, min(n_ortho, min(n - 2, len(selected_vars) - 1)))

        # Fit OPLS — removes orthogonal variance from X
        opls = PYOPLS(n_ortho)
        opls.fit(X, y)
        X_filtered = opls.transform(X)

        # Fit PLS on deflated X to get predictive scores
        pls = PLSRegression(n_components=1)
        pls.fit(X_filtered, y)
        T_pred = pls.x_scores_.ravel()
        T_ortho_arr = opls.T_ortho_[:, 0]  # type: ignore[index]

        # S-plot: covariance and correlation of each X column with T_pred
        cov_xT = np.array([np.cov(X[:, j], T_pred)[0, 1] for j in range(X.shape[1])])
        corr_xT = np.array([np.corrcoef(X[:, j], T_pred)[0, 1] for j in range(X.shape[1])])

        # VIP scores
        W = pls.x_weights_          # (p, 1) predictive weights
        T_mat = pls.x_scores_       # (n, 1)
        Q = pls.y_loadings_         # (1, 1)
        SS = float(np.sum(T_mat**2)) * float(np.sum(Q**2))
        p_vars = X.shape[1]
        if SS > 0:
            W_norm = W / np.sqrt(np.sum(W**2, axis=0, keepdims=True))
            VIP = np.sqrt(p_vars * np.sum((W_norm**2) * SS, axis=1) / SS)
        else:
            VIP = np.ones(p_vars)

        # R²Y
        y_pred_full = pls.predict(X_filtered).ravel()
        r2y = float(sk_r2(y, y_pred_full))

        # R²X (predictive + orthogonal reconstructed)
        X_recon = T_pred[:, None] @ pls.x_loadings_.T + opls.T_ortho_ @ opls.P_ortho_.T  # type: ignore[union-attr]
        ss_x_total = float(np.sum(X**2))
        r2x = (float(np.sum(X_recon**2)) / ss_x_total) if ss_x_total > 0 else 0.0

        # Q² via cross-validation
        q2 = None
        if compute_cv and n >= 4:
            cv_iter = LeaveOneOut() if n <= 20 else KFold(n_splits=min(7, n), shuffle=True, random_state=42)
            ss_total = float(np.sum((y - np.mean(y))**2))
            press = 0.0
            for train_idx, test_idx in cv_iter.split(X):
                X_tr, X_te = X[train_idx], X[test_idx]
                y_tr, y_te = y[train_idx], y[test_idx]
                try:
                    n_o = max(1, min(n_ortho, len(X_tr) - 2))
                    opls_cv = PYOPLS(n_o)
                    opls_cv.fit(X_tr, y_tr)
                    X_filt_tr = opls_cv.transform(X_tr)
                    X_filt_te = opls_cv.transform(X_te)
                    pls_cv = PLSRegression(n_components=1)
                    pls_cv.fit(X_filt_tr, y_tr)
                    y_pred_cv = pls_cv.predict(X_filt_te).ravel()
                    press += float(np.sum((y_te - y_pred_cv)**2))
                except Exception:
                    pass
            q2 = float(1.0 - press / ss_total) if ss_total > 0 else None

        def _jf(v):
            """Return float safe for JSON serialisation; NaN / Inf → None."""
            if v is None:
                return None
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else f

        scores = [
            {'T': _jf(T_pred[i]), 'T_ortho': _jf(T_ortho_arr[i]),
             'label': str(i + 1), 'group': str(group_labels[i])}
            for i in range(n)
        ]
        splot = [
            {'var': selected_vars[j], 'cov': _jf(cov_xT[j]), 'corr': _jf(corr_xT[j]),
             'vip': _jf(VIP[j])}
            for j in range(len(selected_vars))
        ]
        vip_sorted = sorted(
            [{'var': selected_vars[j], 'vip': _jf(VIP[j])} for j in range(len(selected_vars))],
            key=lambda x: (x['vip'] or 0), reverse=True
        )

        return jsonify({
            'method': method,
            'scores': scores,
            'splot': splot,
            'vip': vip_sorted,
            'r2x': _jf(r2x),
            'r2y': _jf(r2y),
            'q2': _jf(q2),
            'n_samples': int(n),
            'n_ortho': int(n_ortho),
            'y_col': y_col,
            'class_labels': class_labels,
            'vars_used': selected_vars,
        })
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@stats_bp.route('/export-opls-excel', methods=['POST'])
def export_opls_excel():
    try:
        request_data = request.get_json()
        result = request_data.get('result', {})
        images = request_data.get('images', {})  # {score: base64, splot: base64, vip: base64}

        output = io.BytesIO()
        wb = Workbook()

        # Sheet 1: Summary + scores
        ws1: Worksheet = wb.active  # type: ignore[assignment]
        ws1.title = 'OPLS Results'
        method_str = 'OPLS-DA' if result.get('method') == 'opls-da' else 'OPLS'
        ws1['A1'] = f'{method_str} Analysis Report'
        ws1['A1'].font = Font(bold=True, size=13)
        ws1['A2'] = f"Y variable: {result.get('y_col', '')}"
        ws1['A3'] = f"Samples: {result.get('n_samples', '')}"
        ws1['A4'] = f"Orthogonal components: {result.get('n_ortho', '')}"
        r2x = result.get('r2x')
        r2y = result.get('r2y')
        q2 = result.get('q2')
        ws1['A5'] = f"R\u00b2X = {r2x:.4f}  |  R\u00b2Y = {r2y:.4f}  |  Q\u00b2 = {f'{q2:.4f}' if q2 is not None else 'N/A'}"

        # Score table
        header = ['Sample', 'Group', 'T (predictive)', 'T_orth (orthogonal)']
        ws1.append([])
        ws1.append(header)
        hrow = ws1.max_row
        for cell in ws1[hrow]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='D9E1F2')
        for s in result.get('scores', []):
            ws1.append([s.get('label'), s.get('group'), s.get('T'), s.get('T_ortho')])

        # Sheet 2: S-plot data
        ws2 = wb.create_sheet('S-plot Data')
        ws2.append(['Variable', 'Covariance (p)', 'Correlation (p_corr)', 'VIP'])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='D9E1F2')
        for s in result.get('splot', []):
            ws2.append([s.get('var'), s.get('cov'), s.get('corr'), s.get('vip')])

        # Sheet 3: VIP data
        ws3 = wb.create_sheet('VIP Scores')
        ws3.append(['Variable', 'VIP Score'])
        for cell in ws3[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='D9E1F2')
        vip_ref_row = None
        for i, v in enumerate(result.get('vip', [])):
            ws3.append([v.get('var'), v.get('vip')])
            if v.get('vip', 0) >= 1.0 and vip_ref_row is None:
                vip_ref_row = ws3.max_row

        # Embed plot images
        for sheet_name, img_key in [('Score Plot', 'score'), ('S-plot', 'splot'), ('VIP Plot', 'vip')]:
            img_b64 = images.get(img_key)
            if img_b64:
                ws_img = wb.create_sheet(sheet_name)
                img_data = base64.b64decode(img_b64)
                img = XLImage(io.BytesIO(img_data))
                ws_img.add_image(img, 'A1')

        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='OPLS_Analysis.xlsx'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@stats_bp.route('/run-pls', methods=['POST'])
def run_pls():
    try:
        from sklearn.cross_decomposition import PLSRegression
        from sklearn.preprocessing import LabelBinarizer
        from sklearn.model_selection import KFold, LeaveOneOut
        from sklearn.metrics import r2_score as sk_r2

        request_data = request.get_json()
        data = request_data.get('data', [])
        ok, err = _validate_data_limits(data)
        if not ok:
            return jsonify({"error": err}), 400
        df = pd.DataFrame(data)
        factors = request_data.get('factors', [])
        selected_vars = request_data.get('variables', [])
        method = request_data.get('method', 'pls-da')   # 'pls-da' or 'pls'
        y_col = request_data.get('y_col')
        n_components = int(request_data.get('n_components', 2))
        compute_cv = bool(request_data.get('compute_cv', True))

        if not y_col:
            return jsonify({"error": "Response variable (Y) is required."}), 400
        if len(selected_vars) < 2:
            return jsonify({"error": "PLS requires at least 2 predictor variables."}), 400

        clean_df = df.copy()
        for var in selected_vars:
            clean_df[var] = pd.to_numeric(clean_df[var], errors='coerce')
        clean_df = clean_df.dropna(subset=selected_vars).copy()

        if len(clean_df) < 4:
            return jsonify({"error": "Insufficient data: PLS requires at least 4 valid samples."}), 400

        # Group labels for colouring (factors or Y column)
        if factors:
            for f in factors:
                clean_df[f] = clean_df[f].astype(str).replace(['nan', 'None'], 'N/A')
            hue_col = ' & '.join(factors)
            clean_df[hue_col] = clean_df[factors].agg(' | '.join, axis=1)
            group_labels = clean_df[hue_col].tolist()
        else:
            group_labels = clean_df[y_col].astype(str).tolist()

        # Standardise X
        scaler = StandardScaler()
        X = scaler.fit_transform(clean_df[selected_vars].values)

        # Prepare Y
        class_labels = None
        if method == 'pls-da':
            lb = LabelBinarizer()
            Y_bin: np.ndarray = np.asarray(lb.fit_transform(clean_df[y_col].astype(str)))
            y = (Y_bin.ravel() if Y_bin.shape[1] == 1 else Y_bin[:, 0]).astype(float)
            class_labels = lb.classes_.tolist()
        else:
            y = pd.to_numeric(clean_df[y_col], errors='coerce').values
            if np.isnan(y).any():
                return jsonify({"error": f"Column '{y_col}' contains non-numeric or missing values."}), 400

        n, p = len(clean_df), X.shape[1]
        n_components = max(1, min(n_components, min(n - 1, p)))

        pls = PLSRegression(n_components=n_components)
        pls.fit(X, y)

        T = pls.x_scores_    # (n, n_comp) — LV scores
        W = pls.x_weights_   # (p, n_comp) — weights for VIP + biplot
        P = pls.x_loadings_  # (p, n_comp) — X loadings
        Q = pls.y_loadings_  # (1, n_comp) — Y loadings

        # VIP scores (all components)
        T_sq = np.sum(T ** 2, axis=0)
        Q_sq = np.sum(Q ** 2, axis=0)
        SS = T_sq * Q_sq
        SS_total = float(np.sum(SS))
        if SS_total > 0:
            W_norm = W / np.sqrt(np.sum(W ** 2, axis=0, keepdims=True))
            VIP = np.sqrt(p * np.sum((W_norm ** 2) * SS, axis=1) / SS_total)
        else:
            VIP = np.ones(p)

        # Cumulative R²X per component
        x_var_total = float(np.sum(X ** 2))
        X_rec = np.zeros_like(X)
        r2x_per_comp = []
        for k in range(n_components):
            X_rec += T[:, k:k+1] @ P[:, k:k+1].T
            r2x_per_comp.append(float(np.sum(X_rec ** 2)) / float(x_var_total) if x_var_total > 0 else 0.0)

        # R²Y
        r2y = float(sk_r2(y, pls.predict(X).ravel()))

        # Q²
        q2 = None
        if compute_cv and n >= 4:
            cv_iter = LeaveOneOut() if n <= 20 else KFold(n_splits=min(7, n), shuffle=True, random_state=42)
            ss_total = float(np.sum((y - np.mean(y)) ** 2))
            press = 0.0
            for train_idx, test_idx in cv_iter.split(X):
                X_tr, X_te = X[train_idx], X[test_idx]
                y_tr, y_te = y[train_idx], y[test_idx]
                try:
                    n_c = max(1, min(n_components, len(X_tr) - 1))
                    pls_cv = PLSRegression(n_components=n_c)
                    pls_cv.fit(X_tr, y_tr)
                    press += float(np.sum((y_te - pls_cv.predict(X_te).ravel()) ** 2))
                except Exception:
                    pass
            q2 = float(1.0 - press / ss_total) if ss_total > 0 else None

        def _jf(v):
            if v is None:
                return None
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else f

        scores = [
            {'T1': _jf(T[i, 0]),
             'T2': _jf(T[i, 1]) if n_components >= 2 else 0.0,
             'label': str(i + 1), 'group': str(group_labels[i])}
            for i in range(n)
        ]
        weights = [
            {'var': selected_vars[j], 'vip': _jf(VIP[j]),
             **{f'W{k+1}': _jf(W[j, k]) for k in range(n_components)}}
            for j in range(p)
        ]
        vip_sorted = sorted(
            [{'var': selected_vars[j], 'vip': _jf(VIP[j])} for j in range(p)],
            key=lambda x: (x['vip'] or 0), reverse=True
        )
        return jsonify({
            'method': method,
            'n_components': n_components,
            'scores': scores,
            'weights': weights,
            'vip': vip_sorted,
            'r2x': _jf(r2x_per_comp[-1]),
            'r2x_per_comp': [_jf(v) for v in r2x_per_comp],
            'r2y': _jf(r2y),
            'q2': _jf(q2),
            'n_samples': int(n),
            'y_col': y_col,
            'class_labels': class_labels,
            'vars_used': selected_vars,
        })
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@stats_bp.route('/export-pls-excel', methods=['POST'])
def export_pls_excel():
    try:
        request_data = request.get_json()
        result = request_data.get('result', {})
        images = request_data.get('images', {})

        output = io.BytesIO()
        wb = Workbook()

        ws1: Worksheet = wb.active  # type: ignore[assignment]
        ws1.title = 'PLS Results'
        method_str = 'PLS-DA' if result.get('method') == 'pls-da' else 'PLS'
        ws1['A1'] = f'{method_str} Analysis Report'
        ws1['A1'].font = Font(bold=True, size=13)
        ws1['A2'] = f"Y variable: {result.get('y_col', '')}"
        ws1['A3'] = f"Samples: {result.get('n_samples', '')} | Components: {result.get('n_components', '')}"
        r2x = result.get('r2x'); r2y = result.get('r2y'); q2 = result.get('q2')
        ws1['A4'] = f"R\u00b2X = {r2x:.4f}  |  R\u00b2Y = {r2y:.4f}  |  Q\u00b2 = {f'{q2:.4f}' if q2 is not None else 'N/A'}"

        ws1.append([])
        header = ['Sample', 'Group', 'T1 (LV1 score)', 'T2 (LV2 score)']
        ws1.append(header)
        hrow = ws1.max_row
        for cell in ws1[hrow]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='D9E1F2')
        for s in result.get('scores', []):
            ws1.append([s.get('label'), s.get('group'), s.get('T1'), s.get('T2')])

        n_comp = result.get('n_components', 2)
        ws2 = wb.create_sheet('Weights')
        ws2.append(['Variable'] + [f'W{k+1}' for k in range(n_comp)] + ['VIP'])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='D9E1F2')
        for row in result.get('weights', []):
            ws2.append([row.get('var')] + [row.get(f'W{k+1}') for k in range(n_comp)] + [row.get('vip')])

        ws3 = wb.create_sheet('VIP Scores')
        ws3.append(['Variable', 'VIP Score'])
        for cell in ws3[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='D9E1F2')
        for v in result.get('vip', []):
            ws3.append([v.get('var'), v.get('vip')])

        for sheet_name, img_key in [('Score Plot', 'score'), ('Weights Plot', 'weights'), ('VIP Plot', 'vip')]:
            img_b64 = images.get(img_key)
            if img_b64:
                ws_img = wb.create_sheet(sheet_name)
                img_data = base64.b64decode(img_b64)
                img = XLImage(io.BytesIO(img_data))
                ws_img.add_image(img, 'A1')

        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='PLS_Analysis.xlsx'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@stats_bp.route('/run-anova', methods=['POST'])
def run_anova():
    """
    Main ANOVA handler (v2).
    Supports: One-way ANOVA, Welch's ANOVA, Kruskal-Wallis,
              Two-way ANOVA, Scheirer-Ray-Hare, MANOVA (Pillai's Trace), ART ANOVA.
    Auto-selects test based on factor count + assumption results, with manual override.
    """
    try:
        request_data = request.get_json()
        data = request_data.get('data', [])
        ok, err = _validate_data_limits(data)
        if not ok:
            return jsonify({"error": err}), 400
        df = pd.DataFrame(data)
        factors = request_data.get('factors', [])
        selected_vars = request_data.get('target_columns', [])
        grouping_mode = request_data.get('grouping_mode', 'all_combined')
        manual_override = request_data.get('manual_override', 'auto')

        if not selected_vars or not factors:
            return jsonify({"error": "Missing target_columns or factors"}), 400

        # ── Quick assumption check to drive auto-selection ────────────────────
        all_normal_list = []
        homogeneous_list = []
        for var in selected_vars:
            if var not in df.columns:
                continue
            tmp = df.copy()
            tmp[var] = pd.to_numeric(tmp[var], errors='coerce')
            tmp = tmp.dropna(subset=[var])
            for f in factors:
                tmp[f] = tmp[f].astype(str).replace(['nan', 'None'], 'N/A')
            tmp['_group'] = tmp[factors[0]]
            group_vals = {g: grp[var].astype(float).values
                          for g, grp in tmp.groupby('_group')
                          if len(grp) >= 3}
            if len(group_vals) < 2:
                continue
            norms = [bool(stats.shapiro(v)[1] > 0.05) for v in group_vals.values()]
            all_normal_list.append(all(norms))
            _, lp = stats.levene(*group_vals.values())
            homogeneous_list.append(bool(lp > 0.05))

        all_normal = all(all_normal_list) if all_normal_list else False
        homogeneous = all(homogeneous_list) if homogeneous_list else False
        n_factors = len(factors)

        # ── Grouping-mode override ────────────────────────────────────────────
        # When the user chose 'across:F' or 'per:F|G', the intent is to run
        # a single-factor comparison per slice, regardless of how many factors
        # were selected in the panel.  Force 1-factor test selection so that
        # _run_one_factor_tests (and _anova_build_slices) is used.
        uses_slicing = (n_factors > 1 and
                        (grouping_mode.startswith('across:') or
                         grouping_mode.startswith('per:')))

        # ── Select test ───────────────────────────────────────────────────────
        if manual_override and manual_override != 'auto':
            test_key = manual_override
            test_rationale = f"Manual override: {test_key.replace('_', ' ').title()} selected by user"
            # If the override is a multi-factor test but slicing was requested,
            # fall back to auto 1-factor selection instead.
            if uses_slicing and test_key not in ('one_way_anova', 'welch_anova', 'kruskal_wallis'):
                test_key, test_rationale = _select_test_auto(1, all_normal, homogeneous)
                test_rationale += f' (sliced by grouping mode: {grouping_mode})'
        else:
            eff_n = 1 if uses_slicing else n_factors
            test_key, test_rationale = _select_test_auto(eff_n, all_normal, homogeneous)
            if uses_slicing:
                test_rationale += f' (sliced by grouping mode: {grouping_mode})'

        # ── Dispatch ──────────────────────────────────────────────────────────
        if test_key in ('one_way_anova', 'welch_anova', 'kruskal_wallis'):
            all_results = _run_one_factor_tests(
                df, selected_vars, factors, grouping_mode,
                test_key, all_normal, homogeneous
            )
        elif test_key == 'two_way_anova':
            all_results = _run_two_way_anova(df, selected_vars, factors, all_normal, homogeneous)
        elif test_key == 'scheirer_ray_hare':
            all_results = _run_scheirer_ray_hare(df, selected_vars, factors, all_normal, homogeneous)
        elif test_key == 'manova':
            all_results = _run_manova(df, selected_vars, factors, all_normal, homogeneous)
        elif test_key == 'art_anova':
            all_results = _run_art_anova(df, selected_vars, factors, all_normal, homogeneous)
        else:
            return jsonify({"error": f"Unknown test key: {test_key}"}), 400

        return jsonify(_make_json_safe({
            "results": all_results,
            "test_key": test_key,
            "test_rationale": test_rationale
        }))

    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# ════════════════════════════════════════════════════════════════════════════════
# ANOVA v2 — HELPER FUNCTIONS
# ════════════════════════════════════════════════════════════════════════════════

def _select_test_auto(n_factors, all_normal, homogeneous):
    """Return (test_key, rationale_str) based on factor count and assumption results."""
    if n_factors == 1:
        if all_normal and homogeneous:
            return 'one_way_anova', 'One-way ANOVA: normality ✓ and homogeneity ✓ both passed'
        elif all_normal and not homogeneous:
            return 'welch_anova', "Welch's ANOVA: normality ✓ passed but homogeneity ✗ failed (heteroscedastic data)"
        else:
            return 'kruskal_wallis', 'Kruskal–Wallis: normality ✗ failed — non-parametric test selected'
    elif n_factors == 2:
        if all_normal and homogeneous:
            return 'two_way_anova', 'Two-way ANOVA: normality ✓ and homogeneity ✓ both passed'
        else:
            return 'scheirer_ray_hare', 'Scheirer–Ray–Hare: assumptions not fully met — non-parametric two-factor test selected'
    else:  # 3+
        if all_normal and homogeneous:
            return 'manova', "MANOVA (Pillai's Trace): normality ✓ and homogeneity ✓ — multivariate test selected"
        else:
            return 'art_anova', 'ART ANOVA (Aligned Rank Transform): assumptions not met — non-parametric multi-factor test selected'


def _bh_correction(p_values):
    """Benjamini-Hochberg FDR correction. Returns adjusted p-values list."""
    n = len(p_values)
    if n == 0:
        return []
    indexed = sorted(enumerate(p_values), key=lambda x: x[1])
    adjusted = [1.0] * n
    prev = 1.0
    for rank, (orig_idx, p) in enumerate(reversed(indexed), 1):
        adj = p * n / (n - rank + 1)
        adj = min(adj, prev)
        prev = adj
        adjusted[orig_idx] = min(adj, 1.0)
    return adjusted


def _dunns_test_bh(groups_dict):
    """
    Dunn (1964) pairwise post-hoc for Kruskal-Wallis, with Benjamini-Hochberg correction.
    groups_dict: { group_name: array_of_values }
    Returns list of { group1, group2, p_raw, p_adj, significant }.
    """
    group_names = list(groups_dict.keys())
    all_values, group_labels = [], []
    for gname in group_names:
        for v in groups_dict[gname]:
            all_values.append(v)
            group_labels.append(gname)

    n_total = len(all_values)
    ranks = stats.rankdata(all_values)

    mean_ranks, group_sizes = {}, {}
    for gname in group_names:
        idxs = [i for i, g in enumerate(group_labels) if g == gname]
        group_ranks = [ranks[i] for i in idxs]
        mean_ranks[gname] = np.mean(group_ranks)
        group_sizes[gname] = len(idxs)

    unique_vals, tie_counts = np.unique(all_values, return_counts=True)
    tie_factor = sum(t**3 - t for t in tie_counts if t > 1)
    tie_correction = 1 - tie_factor / (n_total**3 - n_total) if n_total > 1 else 1.0

    pairs = list(itertools.combinations(group_names, 2))
    z_stats, p_raws = [], []
    for g1, g2 in pairs:
        n1, n2 = group_sizes[g1], group_sizes[g2]
        mr1, mr2 = mean_ranks[g1], mean_ranks[g2]
        se = math.sqrt(tie_correction * n_total * (n_total + 1) / 12.0 * (1.0 / n1 + 1.0 / n2))
        z = abs(mr1 - mr2) / se if se > 0 else 0.0
        z_stats.append(z)
        p_raws.append(float(2 * (1 - stats.norm.cdf(z))))

    p_adjs = _bh_correction(p_raws)
    return [
        {'group1': g1, 'group2': g2,
         'p_adj': float(p_adjs[i]), 'significant': p_adjs[i] < 0.05}
        for i, (g1, g2) in enumerate(pairs)
    ]


def _games_howell(groups_dict):
    """
    Games-Howell pairwise post-hoc for Welch's ANOVA (unequal variances).
    Uses Welch t-statistics with BH correction.
    groups_dict: { group_name: array_of_values }
    Returns list of { group1, group2, p_adj, significant }.
    """
    group_names = list(groups_dict.keys())
    pairs = list(itertools.combinations(group_names, 2))
    results = []
    for g1, g2 in pairs:
        v1 = np.array(groups_dict[g1], dtype=float)
        v2 = np.array(groups_dict[g2], dtype=float)
        n1, n2 = len(v1), len(v2)
        if n1 < 2 or n2 < 2:
            results.append({'group1': g1, 'group2': g2, 'p_raw': 1.0})
            continue
        m1, m2 = float(np.mean(v1)), float(np.mean(v2))
        s1, s2 = float(np.var(v1, ddof=1)), float(np.var(v2, ddof=1))
        se = math.sqrt(s1 / float(n1) + s2 / float(n2))
        if se == 0:
            results.append({'group1': g1, 'group2': g2, 'p_raw': 1.0})
            continue
        t = abs(m1 - m2) / se
        num = (s1 / n1 + s2 / n2) ** 2
        den = (s1 / n1) ** 2 / (n1 - 1) + (s2 / n2) ** 2 / (n2 - 1)
        df = num / den if den > 0 else min(n1, n2) - 1
        results.append({'group1': g1, 'group2': g2, 'p_raw': float(2 * stats.t.sf(t, df))})

    p_adjs = _bh_correction([r['p_raw'] for r in results])
    for i, r in enumerate(results):
        r['p_adj'] = float(p_adjs[i])
        r['significant'] = p_adjs[i] < 0.05
        del r['p_raw']
    return results


def _group_stats(groups_dict):
    """Return list of { group, mean, std, n, values } preserving key order."""
    return [
        {'group': g,
         'mean': float(np.mean(v)),
         'std': float(np.std(v, ddof=1)) if len(v) > 1 else 0.0,
         'n': int(len(v)),
         'values': [float(x) for x in v]}
        for g, v in groups_dict.items()
    ]


def _letter_groups_from_posthoc(groups_dict, posthoc):
    """Assign CLD letters from post-hoc pairs and return enriched stats list."""
    group_names = list(groups_dict.keys())
    stats_list = _group_stats(groups_dict)

    sig_pairs = set()
    for ph in posthoc:
        if ph.get('significant'):
            sig_pairs.add((ph['group1'], ph['group2']))
            sig_pairs.add((ph['group2'], ph['group1']))

    if not sig_pairs:
        for s in stats_list:
            s['letter'] = 'a'
        return stats_list

    # Adjacency of "not significantly different" pairs
    not_sig_adj = {g: set() for g in group_names}
    for i in range(len(group_names)):
        for j in range(i + 1, len(group_names)):
            g1, g2 = group_names[i], group_names[j]
            if (g1, g2) not in sig_pairs:
                not_sig_adj[g1].add(g2)
                not_sig_adj[g2].add(g1)

    import string
    letter_pool = list(string.ascii_lowercase)
    group_letters = {g: set() for g in group_names}
    letter_idx = 0
    sorted_groups = sorted(group_names, key=lambda g: -np.mean(groups_dict[g]))
    assigned = set()

    for g in sorted_groups:
        if g in assigned and group_letters[g]:
            continue
        clique = {g}
        for candidate in sorted_groups:
            if candidate in clique:
                continue
            if all(candidate in not_sig_adj[m] for m in clique):
                clique.add(candidate)
        letter = letter_pool[letter_idx % len(letter_pool)]
        letter_idx += 1
        for member in clique:
            group_letters[member].add(letter)
            assigned.add(member)

    for g in group_names:
        if not group_letters[g]:
            group_letters[g].add(letter_pool[letter_idx % len(letter_pool)])
            letter_idx += 1

    letter_map = {g: ''.join(sorted(group_letters[g])) for g in group_names}
    for s in stats_list:
        s['letter'] = letter_map.get(s['group'], 'a')
    return stats_list



def _anova_build_slices(df, factors, grouping_mode):
    """Parse grouping_mode and return list of (label, slice_df, [factor(s)])."""
    if grouping_mode == 'all_combined' or len(factors) == 1:
        return [('All', df, factors)]
    if grouping_mode.startswith('across:'):
        compare_factor = grouping_mode.split(':', 1)[1]
        return [('All', df, [compare_factor])]
    if grouping_mode.startswith('per:'):
        parts = grouping_mode.split(':', 1)[1]
        compare_factor, stratify_str = parts.split('|', 1)
        stratify_factors = [s.strip() for s in stratify_str.split(',')]
        df2 = df.copy()
        df2['_stratify_'] = df2[stratify_factors].agg(' | '.join, axis=1)
        slices = []
        for strat_val in sorted(df2['_stratify_'].unique()):
            subset = df2[df2['_stratify_'] == strat_val].copy()
            label = f"{' & '.join(stratify_factors)} = {strat_val}"
            slices.append((label, subset, [compare_factor]))
        return slices
    return [('All', df, factors)]


# ── One-factor tests ──────────────────────────────────────────────────────────

def _run_one_factor_tests(df, selected_vars, factors, grouping_mode,
                          test_key, all_normal, homogeneous):
    """Handles one_way_anova / welch_anova / kruskal_wallis with grouping_mode slices."""
    test_labels = {
        'one_way_anova': ("One-way ANOVA + Tukey HSD", "Tukey HSD"),
        'welch_anova':   ("Welch's ANOVA + Games-Howell", "Games-Howell"),
        'kruskal_wallis':("Kruskal–Wallis + Dunn's (BH)", "Dunn's test (BH)"),
    }
    all_results = []

    for var in selected_vars:
        tmp = df.copy()
        tmp[var] = pd.to_numeric(tmp[var], errors='coerce')
        tmp = tmp.dropna(subset=[var]).copy()
        if tmp.empty:
            continue
        for f in factors:
            tmp[f] = tmp[f].astype(str).replace(['nan', 'None'], 'N/A')

        slices = _anova_build_slices(tmp, factors, grouping_mode)

        for slice_label, slice_df, group_factors in slices:
            # Build Group column
            if len(group_factors) > 1:
                slice_df = slice_df.copy()
                slice_df['Group'] = slice_df[group_factors].agg(' | '.join, axis=1)
            else:
                slice_df = slice_df.copy()
                slice_df['Group'] = slice_df[group_factors[0]].astype(str)

            all_ordered = slice_df['Group'].unique().tolist()
            valid_groups = [g for g in all_ordered
                            if len(slice_df[slice_df['Group'] == g]) >= 3]
            if len(valid_groups) < 2:
                continue

            groups_dict = {g: slice_df[slice_df['Group'] == g][var].astype(float).values.tolist()
                           for g in valid_groups}

            # Per-slice assumption re-check (for accuracy)
            sp = [bool(stats.shapiro(v)[1] > 0.05) for v in groups_dict.values() if len(v) >= 3]
            slice_normal = all(sp) if sp else False
            _, lp = stats.levene(*[np.array(v) for v in groups_dict.values()])
            slice_homo = bool(lp > 0.05)

            test_name, posthoc_method = test_labels[test_key]

            if test_key == 'one_way_anova':
                f_stat, overall_p = stats.f_oneway(*[np.array(v) for v in groups_dict.values()])
                overall_p = float(overall_p)
                posthoc = []
                if overall_p < 0.05:
                    all_vals = np.concatenate([np.array(v) for v in groups_dict.values()])
                    all_lbls = np.concatenate([[g]*len(groups_dict[g]) for g in valid_groups])
                    try:
                        mc = MultiComparison(all_vals, all_lbls)
                        tukey = mc.tukeyhsd()
                        for row in tukey.summary().data[1:]:
                            g1, g2, _, p_adj, _, _, reject = row
                            posthoc.append({'group1': str(g1), 'group2': str(g2),
                                            'p_adj': float(p_adj), 'significant': bool(reject)})
                    except Exception:
                        pass
                # Effect size η²
                grand_mean = np.concatenate([np.array(v) for v in groups_dict.values()]).mean()
                ss_b = sum(len(v)*(np.mean(v)-grand_mean)**2 for v in groups_dict.values())
                ss_t = sum(((np.array(v)-grand_mean)**2).sum() for v in groups_dict.values())
                effect_size = float(ss_b/ss_t) if ss_t > 0 else 0.0

            elif test_key == 'welch_anova':
                _, overall_p = _welch_anova([np.array(v) for v in groups_dict.values()])
                overall_p = float(overall_p) if overall_p is not None else 1.0
                posthoc = _games_howell(groups_dict) if overall_p < 0.05 else []
                effect_size = None

            else:  # kruskal_wallis
                _, overall_p = stats.kruskal(*[np.array(v) for v in groups_dict.values()])
                overall_p = float(overall_p)
                posthoc = _dunns_test_bh(groups_dict) if overall_p < 0.05 else []
                n_total = sum(len(v) for v in groups_dict.values())
                k = len(groups_dict)
                kw_stat, _ = stats.kruskal(*[np.array(v) for v in groups_dict.values()])
                eta2_kw = max(0.0, float((kw_stat - k + 1) / (n_total - k))) if n_total > k else 0.0
                effect_size = eta2_kw

            letter_groups = _letter_groups_from_posthoc(groups_dict, posthoc)

            result = {
                "variable": var,
                "slice_label": slice_label,
                "test_used": test_name,
                "posthoc_method": posthoc_method,
                "overall_p": overall_p,
                "effect_size": effect_size,
                "effect_size_label": "η²",
                "assumptions": {"all_normal": slice_normal, "homogeneous": slice_homo},
                "letter_groups": letter_groups,
                "posthoc": posthoc,
            }
            all_results.append(result)

    return all_results


# ── Two-way ANOVA ─────────────────────────────────────────────────────────────

def _run_two_way_anova(df, selected_vars, factors, all_normal, homogeneous):
    f1, f2 = factors[0], factors[1]
    all_results = []

    for var in selected_vars:
        if var not in df.columns:
            continue
        sub = df[[var, f1, f2]].copy()
        sub[var] = pd.to_numeric(sub[var], errors='coerce')
        for f in [f1, f2]:
            sub[f] = sub[f].astype(str).replace(['nan', 'None'], 'N/A')
        sub = sub.dropna()
        if len(sub) < 6:
            continue

        try:
            formula = f'Q("{var}") ~ C(Q("{f1}")) + C(Q("{f2}")) + C(Q("{f1}")):C(Q("{f2}"))'
            model = ols(formula, data=sub).fit()
            anova_table = sm.stats.anova_lm(model, typ=2)

            def get_p(keyword, exclude=None):
                keys = [k for k in anova_table.index
                        if keyword in str(k) and (exclude is None or exclude not in str(k))]
                return float(anova_table.loc[keys[0], 'PR(>F)']) if keys else 1.0 # type: ignore

            p_f1 = get_p(f1, ':')
            p_f2 = get_p(f2, ':')
            p_int_keys = [k for k in anova_table.index if ':' in str(k)]
            p_int = float(anova_table.loc[p_int_keys[0], 'PR(>F)']) if p_int_keys else None # type: ignore

            ss_res = anova_table.loc['Residual', 'sum_sq']
            ss_others = sum(anova_table.loc[k, 'sum_sq'] # type: ignore
                            for k in anova_table.index if 'Residual' not in str(k)) # type: ignore
            ss_total = ss_res + ss_others
            eta2_f1 = get_p(f1, ':')  # placeholder, compute below
            eta2_f1_keys = [k for k in anova_table.index if f1 in str(k) and ':' not in str(k)]
            eta2_f1 = float(anova_table.loc[eta2_f1_keys[0], 'sum_sq'] / ss_total) if eta2_f1_keys and ss_total > 0 else 0.0

            overall_p = min(p_f1, p_f2)

            # Post-hoc: Tukey on combined group labels
            sub['Group'] = sub[f1].astype(str) + ' / ' + sub[f2].astype(str)
            groups_dict = {g: grp[var].astype(float).values.tolist()
                           for g, grp in sub.groupby('Group') if len(grp) >= 1}
            posthoc = []
            if overall_p < 0.05 and len(groups_dict) >= 2:
                all_vals = np.concatenate([np.array(v) for v in groups_dict.values()])
                all_lbls = np.concatenate([[g]*len(groups_dict[g]) for g in groups_dict])
                try:
                    mc = MultiComparison(all_vals, all_lbls)
                    tukey = mc.tukeyhsd()
                    for row in tukey.summary().data[1:]:
                        g1, g2, _, p_adj, _, _, reject = row
                        posthoc.append({'group1': str(g1), 'group2': str(g2),
                                        'p_adj': float(p_adj), 'significant': bool(reject)})
                except Exception:
                    pass

            letter_groups = _letter_groups_from_posthoc(groups_dict, posthoc)
            result = {
                "variable": var, "slice_label": "All",
                "test_used": "Two-way ANOVA",
                "posthoc_method": "Tukey HSD (combined groups)",
                "overall_p": overall_p,
                "p_factor1": float(p_f1), "p_factor2": float(p_f2),
                "p_interaction": float(p_int) if p_int is not None else None,
                "effect_size": eta2_f1, "effect_size_label": f"η² ({f1})",
                "assumptions": {"all_normal": all_normal, "homogeneous": homogeneous},
                "letter_groups": letter_groups, "posthoc": posthoc,
            }
            all_results.append(result)
        except Exception as e:
            all_results.append({"variable": var, "slice_label": "All", "error": str(e)})

    return all_results


# ── Scheirer-Ray-Hare ─────────────────────────────────────────────────────────

def _scheirer_ray_hare(df, value_col, f1, f2):
    """Returns {f1: {H, df, p}, f2: {…}, 'interaction': {…}} or None."""
    df = df[[value_col, f1, f2]].dropna().copy()
    n = len(df)
    if n < 4:
        return None
    df['_rank'] = stats.rankdata(df[value_col])
    ss_total = np.sum((df['_rank'] - df['_rank'].mean()) ** 2)

    def ss_effect(col):
        gm = df.groupby(col)['_rank'].mean()
        gn = df.groupby(col)['_rank'].count()
        return sum(gn[g] * (gm[g] - df['_rank'].mean()) ** 2 for g in gm.index)

    ss1 = ss_effect(f1)
    ss2 = ss_effect(f2)
    df['_cell'] = df[f1].astype(str) + '|||' + df[f2].astype(str)
    ss_cells = ss_effect('_cell')
    ss_int = ss_cells - ss1 - ss2
    df1 = df[f1].nunique() - 1
    df2 = df[f2].nunique() - 1
    df_int = df1 * df2
    ms_err = ss_total / (n - 1) if n > 1 else 1.0

    def h_p(ss, dfe):
        h = (ss / ms_err) if ms_err > 0 else 0.0
        p = float(stats.chi2.sf(h, dfe)) if dfe > 0 else 1.0
        return float(h), p

    h1, p1 = h_p(ss1, df1)
    h2, p2 = h_p(ss2, df2)
    hi, pi = h_p(ss_int, df_int)
    return {
        f1: {'H': h1, 'df': int(df1), 'p': p1},
        f2: {'H': h2, 'df': int(df2), 'p': p2},
        'interaction': {'H': hi, 'df': int(df_int), 'p': pi}
    }


def _run_scheirer_ray_hare(df, selected_vars, factors, all_normal, homogeneous):
    f1, f2 = factors[0], factors[1]
    all_results = []

    for var in selected_vars:
        if var not in df.columns:
            continue
        sub = df[[var, f1, f2]].copy()
        sub[var] = pd.to_numeric(sub[var], errors='coerce')
        for f in [f1, f2]:
            sub[f] = sub[f].astype(str).replace(['nan', 'None'], 'N/A')
        sub = sub.dropna()
        if len(sub) < 6:
            continue

        srh = _scheirer_ray_hare(sub, var, f1, f2)
        if srh is None:
            continue

        overall_p = min(srh[f1]['p'], srh[f2]['p'])
        sub['Group'] = sub[f1].astype(str) + ' / ' + sub[f2].astype(str)
        groups_dict = {g: grp[var].astype(float).values.tolist()
                       for g, grp in sub.groupby('Group') if len(grp) >= 2}
        posthoc = _dunns_test_bh(groups_dict) if overall_p < 0.05 and len(groups_dict) >= 2 else []

        letter_groups = _letter_groups_from_posthoc(groups_dict, posthoc)
        result = {
            "variable": var, "slice_label": "All",
            "test_used": "Scheirer–Ray–Hare",
            "posthoc_method": "Dunn's test (BH)",
            "overall_p": overall_p,
            "p_factor1": srh[f1]['p'], "p_factor2": srh[f2]['p'],
            "p_interaction": srh['interaction']['p'],
            "effect_size": None,
            "assumptions": {"all_normal": all_normal, "homogeneous": homogeneous},
            "letter_groups": letter_groups, "posthoc": posthoc,
            "srh_table": {f1: srh[f1], f2: srh[f2], 'interaction': srh['interaction']}
        }
        all_results.append(result)

    return all_results


# ── MANOVA (Pillai's Trace) ───────────────────────────────────────────────────

def _manova_pillai(df, variable_cols, group_col):
    """Compute Pillai's Trace MANOVA. Returns dict with pillais_trace, F, p, or None."""
    df_clean = df[[group_col] + variable_cols].dropna()
    if len(df_clean) < len(variable_cols) + 2:
        return None
    groups = df_clean[group_col].unique()
    k, p, n = len(groups), len(variable_cols), len(df_clean)
    grand_mean = df_clean[variable_cols].mean().values

    H = np.zeros((p, p))
    for g in groups:
        g_data = df_clean[df_clean[group_col] == g][variable_cols].values
        g_mean = g_data.mean(axis=0)
        diff = (g_mean - grand_mean).reshape(-1, 1)
        H += len(g_data) * (diff @ diff.T)

    E = np.zeros((p, p))
    for g in groups:
        g_data = df_clean[df_clean[group_col] == g][variable_cols].values
        diffs = g_data - g_data.mean(axis=0)
        E += diffs.T @ diffs

    try:
        eigvals = np.linalg.eigvals(np.linalg.pinv(E) @ H).real
        eigvals = eigvals[eigvals > 0]
    except np.linalg.LinAlgError:
        return None

    pillai = float(sum(l / (1 + l) for l in eigvals))
    s = min(k - 1, p)
    m = (abs(k - 1 - p) - 1) / 2
    nn = (n - k - p - 1) / 2
    if s == 0 or nn <= 0:
        return {'pillais_trace': pillai, 'F': None, 'p': None}

    df_num = s * (2 * m + s + 1)
    df_den = s * (2 * nn + s + 1)
    F_stat = (pillai / s) * df_den / ((1 - pillai / s) * df_num) if pillai < s else None
    p_val = float(stats.f.sf(F_stat, df_num, df_den)) if F_stat and F_stat > 0 else None

    return {'pillais_trace': pillai, 'F': F_stat, 'p': p_val}


def _run_manova(df, selected_vars, factors, all_normal, homogeneous):
    """MANOVA (Pillai's Trace) + per-variable ANOVA follow-ups with Bonferroni."""
    df = df.copy()
    df['_group'] = df[factors[0]].astype(str)
    for f in factors[1:]:
        df['_group'] += ' / ' + df[f].astype(str)

    clean = df[['_group'] + selected_vars].copy()
    for v in selected_vars:
        clean[v] = pd.to_numeric(clean[v], errors='coerce')
    clean = clean.dropna()

    manova_res = _manova_pillai(clean, selected_vars, '_group')
    manova_p = manova_res['p'] if manova_res else None
    pillai = manova_res['pillais_trace'] if manova_res else None

    bonferroni_alpha = 0.05 / len(selected_vars) if selected_vars else 0.05
    all_results = []

    for var in selected_vars:
        if var not in df.columns:
            continue
        groups_dict = {g: grp[var].dropna().astype(float).values.tolist()
                       for g, grp in clean.groupby('_group') if len(grp) >= 3}
        if len(groups_dict) < 2:
            continue

        f_stat, overall_p = stats.f_oneway(*[np.array(v) for v in groups_dict.values()])
        overall_p = float(overall_p)
        posthoc = []
        if overall_p < 0.05:
            all_vals = np.concatenate([np.array(v) for v in groups_dict.values()])
            all_lbls = np.concatenate([[g]*len(groups_dict[g]) for g in groups_dict])
            try:
                mc = MultiComparison(all_vals, all_lbls)
                tukey = mc.tukeyhsd()
                for row in tukey.summary().data[1:]:
                    g1, g2, _, p_adj, _, _, reject = row
                    posthoc.append({'group1': str(g1), 'group2': str(g2),
                                    'p_adj': float(p_adj),
                                    'significant': float(p_adj) < bonferroni_alpha})
            except Exception:
                pass

        letter_groups = _letter_groups_from_posthoc(groups_dict, posthoc)
        result = {
            "variable": var, "slice_label": "All",
            "test_used": "MANOVA → per-variable ANOVA",
            "posthoc_method": f"Tukey HSD (Bonferroni α={bonferroni_alpha:.4f})",
            "overall_p": overall_p,
            "effect_size": None,
            "assumptions": {"all_normal": all_normal, "homogeneous": homogeneous},
            "letter_groups": letter_groups, "posthoc": posthoc,
            "manova_pillais_trace": float(pillai) if pillai is not None else None,
            "manova_p": float(manova_p) if manova_p is not None else None,
        }
        all_results.append(result)

    return all_results


# ── ART ANOVA (Aligned Rank Transform) ───────────────────────────────────────

def _art_anova_per_factor(df, value_col, factors):
    """
    ART ANOVA: for each factor, align values by removing all other effects,
    rank the aligned values, then run one-way ANOVA on them.
    Returns {factor: {F, p}}.
    """
    df = df[factors + [value_col]].dropna().copy()
    if len(df) < len(factors) * 3:
        return None
    results = {}
    for target in factors:
        others = [f for f in factors if f != target]
        if others:
            df['_art_mean'] = df.groupby(others)[value_col].transform('mean')
        else:
            df['_art_mean'] = df[value_col].mean()
        df['_aligned'] = df[value_col] - df['_art_mean']
        df['_ranked'] = stats.rankdata(df['_aligned'])
        group_arrs = [grp['_ranked'].values for _, grp in df.groupby(target)]
        if len(group_arrs) < 2:
            continue
        f_stat, p_val = stats.f_oneway(*group_arrs)
        results[target] = {'F': float(f_stat), 'p': float(p_val)}
    return results


def _run_art_anova(df, selected_vars, factors, all_normal, homogeneous):
    """ART ANOVA + Dunn's post-hoc with Bonferroni correction across variables."""
    df = df.copy()
    df['_group'] = df[factors[0]].astype(str)
    for f in factors[1:]:
        df['_group'] += ' / ' + df[f].astype(str)

    bonferroni_alpha = 0.05 / len(selected_vars) if selected_vars else 0.05
    all_results = []

    for var in selected_vars:
        if var not in df.columns:
            continue
        sub = df[[var] + factors + ['_group']].copy()
        sub[var] = pd.to_numeric(sub[var], errors='coerce')
        for f in factors:
            sub[f] = sub[f].astype(str).replace(['nan', 'None'], 'N/A')
        sub = sub.dropna(subset=[var])
        if len(sub) < 6:
            continue

        art_res = _art_anova_per_factor(sub, var, factors)
        if not art_res:
            continue
        overall_p = min(r['p'] for r in art_res.values())

        groups_dict = {g: grp[var].astype(float).values.tolist()
                       for g, grp in sub.groupby('_group') if len(grp) >= 2}
        posthoc = []
        if overall_p < 0.05 and len(groups_dict) >= 2:
            posthoc = _dunns_test_bh(groups_dict)
            for ph in posthoc:
                ph['significant'] = ph['p_adj'] < bonferroni_alpha

        letter_groups = _letter_groups_from_posthoc(groups_dict, posthoc)
        result = {
            "variable": var, "slice_label": "All",
            "test_used": "ART ANOVA",
            "posthoc_method": f"Dunn's test (Bonferroni α={bonferroni_alpha:.4f})",
            "overall_p": overall_p,
            "effect_size": None,
            "assumptions": {"all_normal": all_normal, "homogeneous": homogeneous},
            "letter_groups": letter_groups, "posthoc": posthoc,
            "art_factor_table": art_res,
        }
        all_results.append(result)

    return all_results


# ════════════════════════════════════════════════════════════════════════════════
# FULL REPORT EXCEL EXPORT  (/export-full-report)
# ════════════════════════════════════════════════════════════════════════════════

@stats_bp.route('/export-full-report', methods=['POST'])
def export_full_report():
    """
    Generates a 6-sheet .xlsx Full Statistical Report.

    Expected JSON payload:
    {
        original_data: [...],          # raw rows before any processing
        analysed_data: [...],          # rows used for ANOVA (after transforms/exclusions)
        transform_notes: "...",        # human-readable transform summary
        anova_results: { results: [...], test_rationale: "..." },
        assumption_results: [...],     # from /run-tests
        original_assumption_results: [],
        plot_captures: [{ label, type: 'box'|'residuals'|'qq', image: '<base64 png>' }],
        factors: ['F1', 'F2'],
        target_columns: ['var1', 'var2']
    }
    """
    try:
        payload = request.get_json(force=True)
        original_data = payload.get('original_data', [])
        analysed_data = payload.get('analysed_data', [])
        transform_notes = payload.get('transform_notes', 'None')
        anova_results = payload.get('anova_results', {})
        assumption_results_raw = payload.get('assumption_results') or {}
        orig_assumption_results_raw = payload.get('original_assumption_results') or {}
        # lastTestResults in JS has shape { results: [...] }
        assumption_results = assumption_results_raw.get('results', []) if isinstance(assumption_results_raw, dict) else (assumption_results_raw or [])
        orig_assumption_results = orig_assumption_results_raw.get('results', []) if isinstance(orig_assumption_results_raw, dict) else (orig_assumption_results_raw or [])
        plot_captures = payload.get('plot_captures', [])
        factors = payload.get('factors', [])
        target_cols = payload.get('target_columns', [])

        wb = Workbook()

        # Use module-level style helpers and fill palette
        fills      = _XL_FILLS
        hdr_font   = _XL_HDR_FONT
        thin_border = _XL_THIN_BORDER
        center_al  = _XL_CENTER_AL
        left_al    = Alignment(horizontal='left', vertical='center', wrap_text=True)
        style_header_row = _xl_style_header
        auto_width       = _xl_auto_width
        border_row       = _xl_border_row

        # ══ Sheet 1: Original Data ════════════════════════════════════════════
        ws1 = wb.create_sheet('1 - Original Data', 0)
        if original_data:
            orig_df = pd.DataFrame(original_data).drop(columns=['row_id'], errors='ignore')
            headers = list(orig_df.columns)
            ws1.append(headers)
            style_header_row(ws1, 1, fills['blue'])
            for _, row in orig_df.iterrows():
                ws1.append([row.get(h, '') for h in headers])
            auto_width(ws1)

        # ══ Sheet 2: Analysed Data ════════════════════════════════════════════
        ws2 = wb.create_sheet('2 - Analysed Data')
        ws2.cell(1, 1, 'Data used for ANOVA analysis').font = Font(bold=True, size=12)
        ws2.cell(2, 1, f'Transformations applied: {transform_notes}').fill = fills['note']

        orig_ids = {r.get('row_id') for r in original_data}
        anal_ids = {r.get('row_id') for r in analysed_data}
        excluded = orig_ids - anal_ids
        if excluded:
            ws2.cell(3, 1, 'Excluded row IDs: ' + ', '.join(str(e) for e in sorted(excluded))).fill = fills['fail']

        ws2.append([])
        if analysed_data:
            an_df = pd.DataFrame(analysed_data).drop(columns=['row_id'], errors='ignore')
            headers = list(an_df.columns)
            ws2.append(headers)
            style_header_row(ws2, ws2.max_row, fills['green'])
            for _, row in an_df.iterrows():
                ws2.append([row.get(h, '') for h in headers])
            auto_width(ws2)

        # ══ Sheet 3: Assumption Tests ══════════════════════════════════════════
        ws3 = wb.create_sheet('3 - Assumption Tests')
        has_transforms = bool(transform_notes and transform_notes.strip() not in ('', 'None'))

        if has_transforms:
            hdrs3 = ['Variable', 'Group',
                     'SW p (original)', 'Normal? (original)',
                     'SW p (transformed)', 'Normal? (transformed)',
                     'Levene p (original)', 'Homo? (original)',
                     'Levene p (transformed)', 'Homo? (transformed)']
        else:
            hdrs3 = ['Variable', 'Group', 'Shapiro-Wilk p', 'Normal?',
                     'Levene p', 'Homogeneous?']

        ws3.append(hdrs3)
        style_header_row(ws3, 1, fills['orange'])

        def find_var_result(result_list, var_name):
            for r in (result_list or []):
                if r.get('variable') == var_name:
                    return r
            return None

        for var in target_cols:
            res_t = find_var_result(assumption_results, var)
            res_o = find_var_result(orig_assumption_results, var) if orig_assumption_results else None
            if res_t is None:
                continue
            lev_t = res_t.get('levene', {})
            lp_t = lev_t.get('p')
            lh_t = lev_t.get('is_homogeneous')
            lev_o = res_o.get('levene', {}) if res_o else {}
            lp_o = lev_o.get('p') if lev_o else None
            lh_o = lev_o.get('is_homogeneous') if lev_o else None

            for sg in res_t.get('shapiro', []):
                grp = sg.get('group', '')
                sw_p_t = sg.get('p')
                is_n_t = sg.get('is_normal')
                sw_p_o, is_n_o = None, None
                if res_o:
                    for osg in res_o.get('shapiro', []):
                        if osg.get('group') == grp:
                            sw_p_o, is_n_o = osg.get('p'), osg.get('is_normal')
                            break

                if has_transforms:
                    row_data = [
                        var, grp,
                        round(sw_p_o, 4) if sw_p_o is not None else '',
                        'Pass' if is_n_o else 'Fail',
                        round(sw_p_t, 4) if sw_p_t is not None else '',
                        'Pass' if is_n_t else 'Fail',
                        round(lp_o, 4) if lp_o is not None else '',
                        'Pass' if lh_o else 'Fail',
                        round(lp_t, 4) if lp_t is not None else '',
                        'Pass' if lh_t else 'Fail',
                    ]
                else:
                    row_data = [
                        var, grp,
                        round(sw_p_t, 4) if sw_p_t is not None else '',
                        'Pass' if is_n_t else 'Fail',
                        round(lp_t, 4) if lp_t is not None else '',
                        'Pass' if lh_t else 'Fail',
                    ]
                ws3.append(row_data)
                r_idx = ws3.max_row
                for c_idx, val in enumerate(row_data, 1):
                    cell = ws3.cell(r_idx, c_idx)
                    cell.border = thin_border
                    if val == 'Pass':
                        cell.fill = fills['pass']
                    elif val == 'Fail':
                        cell.fill = fills['fail']
        auto_width(ws3)

        # ══ Sheet 4: ANOVA Results ═════════════════════════════════════════════
        ws4 = wb.create_sheet('4 - ANOVA Results')
        anova_r = anova_results.get('results', [])
        rationale = anova_results.get('test_rationale', '')
        if rationale:
            ws4.cell(1, 1, 'Test selection:').font = Font(bold=True)
            ws4.cell(1, 2, rationale).fill = fills['note']
            ws4.append([])

        by_var = defaultdict(list)
        for r in anova_r:
            by_var[r.get('variable', 'Unknown')].append(r)

        cur_row = ws4.max_row + 1
        for var in target_cols:
            if var not in by_var:
                continue
            # Variable heading
            ws4.cell(cur_row, 1, var)
            ws4.cell(cur_row, 1).font = hdr_font
            ws4.cell(cur_row, 1).fill = fills['purple']
            cur_row += 1

            for res in by_var[var]:
                ws4.cell(cur_row, 1, f"Grouping: {res.get('slice_label', 'All')}")
                ws4.cell(cur_row, 1).font = Font(bold=True, italic=True, size=10)
                cur_row += 1

                for lbl, key in [
                    ('Test used', 'test_used'),
                    ('Post-hoc method', 'posthoc_method'),
                    ('Overall p-value', 'overall_p'),
                    ('Effect size', 'effect_size'),
                    (f'p ({factors[0] if factors else "Factor1"})', 'p_factor1'),
                    (f'p ({factors[1] if len(factors) > 1 else "Factor2"})', 'p_factor2'),
                    ('p (Interaction)', 'p_interaction'),
                    ("MANOVA Pillai's Trace", 'manova_pillais_trace'),
                    ('MANOVA overall p', 'manova_p'),
                ]:
                    val = res.get(key)
                    if val is not None:
                        ws4.cell(cur_row, 1, lbl + ':')
                        ws4.cell(cur_row, 2, round(val, 6) if isinstance(val, float) else val)
                        cur_row += 1

                # Letter-group table
                letter_groups = res.get('letter_groups', [])
                if letter_groups:
                    cur_row += 1
                    for c_idx, h in enumerate(['Group', 'N', 'Mean', 'SD', 'Letter Group'], 1):
                        cell = ws4.cell(cur_row, c_idx, h)
                        cell.font = hdr_font
                        cell.fill = fills['teal']
                        cell.alignment = center_al
                        cell.border = thin_border
                    cur_row += 1
                    for lg in letter_groups:
                        row_vals = [lg.get('group', ''), lg.get('n', 0),
                                    round(lg.get('mean', 0), 4), round(lg.get('std', 0), 4),
                                    lg.get('letter', '')]
                        for c_idx, v in enumerate(row_vals, 1):
                            cell = ws4.cell(cur_row, c_idx, v)
                            cell.alignment = center_al
                            cell.border = thin_border
                        cur_row += 1
                cur_row += 1
            cur_row += 1
        auto_width(ws4)

        # ══ Sheet 5: Pairwise Comparisons ═════════════════════════════════════
        ws5 = wb.create_sheet('5 - Pairwise Comparisons')
        hdrs5 = ['Variable', 'Grouping', 'Group 1', 'Group 2', 'p (adjusted)', 'Significant?']
        ws5.append(hdrs5)
        style_header_row(ws5, 1, fills['grey'])

        for var in target_cols:
            for res in by_var.get(var, []):
                for ph in res.get('posthoc', []):
                    ws5.append([
                        var, res.get('slice_label', 'All'),
                        ph.get('group1', ''), ph.get('group2', ''),
                        round(ph.get('p_adj', 1.0), 6),
                        'Yes' if ph.get('significant') else 'No'
                    ])
                    r = ws5.max_row
                    ws5.cell(r, 6).fill = fills['pass'] if ph.get('significant') else fills['light']
                    border_row(ws5, r, 6)
        auto_width(ws5)

        # ══ Sheet 6: Plots ═════════════════════════════════════════════════════
        ws6 = wb.create_sheet('6 - Plots')
        ws6.cell(1, 1, 'Statistical Diagnostic Plots').font = Font(bold=True, size=14)
        ws6.append([])

        type_labels = {'box': 'Box Plot', 'residuals': 'Residuals vs Fitted', 'qq': 'Normal Q-Q'}
        col_positions = [1, 12]   # Two plots per row (columns A and L)
        cur_plot_row = 3
        cur_col_idx = 0

        if not plot_captures:
            ws6.cell(3, 1, 'No plots captured. Run assumption tests before exporting.').fill = fills['note']
        else:
            for pc in plot_captures:
                img_b64 = pc.get('image', '')
                label = pc.get('label', '')
                ptype = pc.get('type', '')
                type_label = type_labels.get(ptype, ptype)

                label_cell = ws6.cell(cur_plot_row, col_positions[cur_col_idx],
                                       f'{type_label}: {label}')
                label_cell.font = Font(bold=True, size=9)
                label_cell.fill = fills['note']

                if img_b64:
                    try:
                        img_bytes = base64.b64decode(img_b64)
                        xl_img = XLImage(io.BytesIO(img_bytes))
                        xl_img.width = 500 # type: ignore
                        xl_img.height = 300 # type: ignore
                        cell_addr = f'{get_column_letter(col_positions[cur_col_idx])}{cur_plot_row + 1}'
                        ws6.add_image(xl_img, cell_addr)
                    except Exception as img_err:
                        ws6.cell(cur_plot_row + 1, col_positions[cur_col_idx],
                                  f'[Image error: {img_err}]')

                cur_col_idx += 1
                if cur_col_idx >= len(col_positions):
                    cur_col_idx = 0
                    cur_plot_row += 22   # rows per image block

        # ── Serialise ─────────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            download_name='Statistical_Analysis_Full_Report.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# ================== CORRELATION ANALYSIS ==================

@stats_bp.route('/run-correlation', methods=['POST'])
def run_correlation():
    try:
        request_data = request.get_json()
        data = request_data.get('data', [])
        ok, err = _validate_data_limits(data)
        if not ok:
            return jsonify({"error": err}), 400

        df = pd.DataFrame(data)
        variables = request_data.get('variables', [])
        method = request_data.get('method', 'pearson')  # pearson | spearman | kendall
        min_row_coverage = float(request_data.get('min_row_coverage', 0)) / 100.0

        if len(variables) < 2:
            return jsonify({"error": "Correlation requires at least 2 variables."}), 400

        # Convert to numeric, coerce errors to NaN
        for v in variables:
            df[v] = pd.to_numeric(df[v], errors='coerce')

        # Row sparsity filter (shared with PCA)
        rows_sparse_dropped = 0
        if min_row_coverage > 0 and variables:
            n_sel = len(variables)
            coverage = df[variables].notna().sum(axis=1) / n_sel
            before_sparse = len(df)
            df = df[coverage >= min_row_coverage].copy()
            rows_sparse_dropped = before_sparse - len(df)

        subset = df[variables]

        # Compute correlation matrix and p-value matrix pairwise (handles NaN)
        n_vars = len(variables)
        corr_matrix: list = [[None] * n_vars for _ in range(n_vars)]
        pval_matrix: list = [[None] * n_vars for _ in range(n_vars)]
        n_pairs     = [[0]    * n_vars for _ in range(n_vars)]

        method_fn = {
            'pearson':  stats.pearsonr,
            'spearman': stats.spearmanr,
            'kendall':  stats.kendalltau,
        }.get(method, stats.pearsonr)

        for i in range(n_vars):
            for j in range(n_vars):
                if i == j:
                    corr_matrix[i][j] = 1.0
                    pval_matrix[i][j] = 0.0
                    valid = subset[variables[i]].dropna()
                    n_pairs[i][j] = int(valid.count())
                else:
                    xy = subset[[variables[i], variables[j]]].dropna()
                    n = len(xy)
                    n_pairs[i][j] = n
                    if n < 3:
                        corr_matrix[i][j] = None
                        pval_matrix[i][j] = None
                    else:
                        r, p = method_fn(xy[variables[i]], xy[variables[j]])
                        corr_matrix[i][j] = float(r)
                        pval_matrix[i][j] = float(p)

        # Hierarchical clustering order (Ward linkage on correlation distance)
        hclust_dendrogram = None
        hclust_error      = None
        try:
            from scipy.cluster.hierarchy import linkage, leaves_list, dendrogram as sp_dendrogram
            from scipy.spatial.distance import squareform

            # Check for variables with no valid pairs (all-None correlation row)
            null_vars = [variables[i] for i in range(n_vars)
                         if all(corr_matrix[i][j] is None for j in range(n_vars) if j != i)]
            if null_vars:
                raise ValueError(
                    f"Cannot cluster: variable(s) {null_vars} have no valid correlations "
                    f"(possibly non-numeric or constant data)."
                )

            # Build a complete correlation matrix with NaN → 0 for clustering
            corr_np = np.array([[corr_matrix[i][j] if corr_matrix[i][j] is not None else 0.0
                                  for j in range(n_vars)] for i in range(n_vars)])
            np.fill_diagonal(corr_np, 1.0)
            dist = 1.0 - np.abs(corr_np)
            np.fill_diagonal(dist, 0.0)
            dist = np.clip(dist, 0, None)
            condensed = squareform(dist, checks=False)
            Z = linkage(condensed, method='ward')
            hclust_order = leaves_list(Z).tolist()
            dend = sp_dendrogram(Z, labels=variables, no_plot=True, color_threshold=-1)
            hclust_dendrogram = {
                'icoord':     dend['icoord'],
                'dcoord':     dend['dcoord'],
                'color_list': dend['color_list'],
                'ivl':        dend['ivl'],
            }
        except Exception as e:
            hclust_order = list(range(n_vars))
            hclust_error = str(e)

        # Compute redundant pairs for Step 4 (variable reduction)
        corr_threshold = float(request_data.get('corr_threshold', 0.80))
        ordered_vars = variables  # already ordered from input
        n = len(ordered_vars)
        redundant_pairs = []
        for i in range(n):
            for j in range(i + 1, n):
                r = corr_matrix[i][j]
                if r is not None and abs(r) >= corr_threshold:
                    redundant_pairs.append({
                        'var_a': ordered_vars[i],
                        'var_b': ordered_vars[j],
                        'r': round(r, 4),
                        'p': round(pval_matrix[i][j], 4) if pval_matrix[i][j] is not None else None,
                        'abs_r': round(abs(r), 4)
                    })

        return jsonify(_make_json_safe({
            "variables":         variables,
            "method":            method,
            "corr_matrix":       corr_matrix,
            "pval_matrix":       pval_matrix,
            "n_pairs":           n_pairs,
            "hclust_order":        hclust_order,
            "hclust_dendrogram":   hclust_dendrogram,
            "hclust_error":        hclust_error,
            "rows_sparse_dropped": rows_sparse_dropped,
            "redundant_pairs":     redundant_pairs,
            "corr_threshold":      corr_threshold,
        }))

    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@stats_bp.route('/run-normalization', methods=['POST'])
def run_normalization():
    """Server-side normalization endpoint (for logging/export purposes).
    Most normalization is done client-side in JS; this route provides a
    server-side equivalent for validation and audit.
    Accepts: { data, variables, method }
    Returns: { normalized_data, method, variables }
    """
    try:
        request_data = request.get_json()
        data = request_data.get('data', [])
        variables = request_data.get('variables', [])
        method = request_data.get('method', 'none')

        if not data or not variables:
            return jsonify({"error": "data and variables are required"}), 400

        df = pd.DataFrame(data)
        for v in variables:
            df[v] = pd.to_numeric(df[v], errors='coerce')

        if method == 'none':
            pass  # no transformation
        elif method == 'log10':
            for v in variables:
                df[v] = df[v].apply(lambda x: float(np.log10(x)) if x is not None and not np.isnan(x) and x > 0 else None)
        elif method == 'ln':
            for v in variables:
                df[v] = df[v].apply(lambda x: float(np.log(x)) if x is not None and not np.isnan(x) and x > 0 else None)
        elif method == 'sqrt':
            for v in variables:
                df[v] = df[v].apply(lambda x: float(np.sqrt(x)) if x is not None and not np.isnan(x) and x >= 0 else None)
        elif method == 'zscore':
            for v in variables:
                col = df[v].dropna()
                if len(col) > 1:
                    m, s = col.mean(), col.std(ddof=1)
                    if s > 0:
                        df[v] = (df[v] - m) / s
                    else:
                        df[v] = 0.0
        elif method == 'minmax':
            for v in variables:
                col = df[v].dropna()
                if len(col) > 1:
                    lo, hi = col.min(), col.max()
                    if hi > lo:
                        df[v] = (df[v] - lo) / (hi - lo)
                    else:
                        df[v] = 0.0

        normalized = df.astype(object).mask(pd.isnull(df)).to_dict(orient='records')
        return jsonify(_make_json_safe({
            "normalized_data": normalized,
            "method": method,
            "variables": variables,
        }))
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@stats_bp.route('/export-correlation-excel', methods=['POST'])
def export_correlation_excel():
    try:
        request_data = request.get_json()
        variables   = request_data.get('variables', [])
        method      = request_data.get('method', 'pearson')
        corr_matrix = request_data.get('corr_matrix', [])
        pval_matrix = request_data.get('pval_matrix', [])
        n_pairs     = request_data.get('n_pairs', [])
        heatmap_img = request_data.get('heatmap_img', None)   # base64 PNG (optional)
        scatter_img = request_data.get('scatter_img', None)   # base64 PNG (optional)

        wb = Workbook()

        # ── Sheet 1: Correlation Matrix ──────────────────────────────────────
        ws_corr: Worksheet = wb.active  # type: ignore[assignment]
        ws_corr.title = 'Correlation Matrix'

        header_fill   = PatternFill("solid", fgColor="2F5496")
        header_font   = Font(color="FFFFFF", bold=True, name='Calibri', size=11)
        subhdr_fill   = PatternFill("solid", fgColor="D6E4F0")
        subhdr_font   = Font(bold=True, name='Calibri', size=10)
        center_align  = Alignment(horizontal='center', vertical='center')
        thin_border   = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

        def _corr_fill(r):
            """Blue-white-red fill for r in [-1, 1]."""
            if r is None:
                return PatternFill("solid", fgColor="F2F2F2")
            r = max(-1.0, min(1.0, r))
            if r >= 0:
                intensity = int(r * 200)
                return PatternFill("solid", fgColor=f"FF{255-intensity:02X}{255-intensity:02X}")
            else:
                intensity = int(-r * 200)
                return PatternFill("solid", fgColor=f"{255-intensity:02X}{255-intensity:02X}FF")

        # Title row
        ws_corr.merge_cells(start_row=1, start_column=1,
                            end_row=1, end_column=len(variables) + 1)
        title_cell = ws_corr.cell(row=1, column=1,
                                  value=f"Correlation Matrix ({method.capitalize()})")
        title_cell.font   = Font(bold=True, size=13, name='Calibri')
        title_cell.fill   = header_fill
        title_cell.font   = Font(color="FFFFFF", bold=True, size=13, name='Calibri')
        title_cell.alignment = center_align
        ws_corr.row_dimensions[1].height = 20  # type: ignore[index]

        # Header row (column labels)
        ws_corr.cell(row=2, column=1, value='Variable').fill = subhdr_fill
        ws_corr.cell(row=2, column=1).font = subhdr_font
        ws_corr.cell(row=2, column=1).alignment = center_align
        for j, var in enumerate(variables):
            c = ws_corr.cell(row=2, column=j + 2, value=var)
            c.fill = subhdr_fill; c.font = subhdr_font; c.alignment = center_align
            c.border = thin_border

        # Data rows
        for i, var_i in enumerate(variables):
            row_lbl = ws_corr.cell(row=i + 3, column=1, value=var_i)
            row_lbl.fill = subhdr_fill; row_lbl.font = subhdr_font
            row_lbl.alignment = center_align; row_lbl.border = thin_border
            for j in range(len(variables)):
                r_val = corr_matrix[i][j] if corr_matrix else None
                p_val = pval_matrix[i][j] if pval_matrix else None
                disp  = f"{r_val:.3f}" if r_val is not None else "—"
                if p_val is not None and p_val < 0.001:
                    disp += " ***"
                elif p_val is not None and p_val < 0.01:
                    disp += " **"
                elif p_val is not None and p_val < 0.05:
                    disp += " *"
                c = ws_corr.cell(row=i + 3, column=j + 2, value=disp)
                c.fill      = _corr_fill(r_val)
                c.alignment = center_align
                c.border    = thin_border
                c.font      = Font(name='Calibri', size=10,
                                   bold=(i == j))

        # Column widths
        ws_corr.column_dimensions[get_column_letter(1)].width = max(
            len(v) for v in variables) + 2
        for j in range(len(variables)):
            ws_corr.column_dimensions[get_column_letter(j + 2)].width = max(
                len(variables[j]) + 2, 12)

        # Embed heatmap image if provided
        if heatmap_img:
            img_data = base64.b64decode(heatmap_img)
            img_stream = io.BytesIO(img_data)
            img_xl = XLImage(img_stream)
            img_row = len(variables) + 5
            ws_corr.add_image(img_xl, f'A{img_row}')

        # ── Sheet 2: P-value Matrix ───────────────────────────────────────────
        ws_pval = wb.create_sheet('P-value Matrix')
        ws_pval.merge_cells(start_row=1, start_column=1,
                            end_row=1, end_column=len(variables) + 1)
        t2 = ws_pval.cell(row=1, column=1,
                          value=f"P-value Matrix ({method.capitalize()})")
        t2.font = Font(color="FFFFFF", bold=True, size=13, name='Calibri')
        t2.fill = header_fill; t2.alignment = center_align

        ws_pval.cell(row=2, column=1, value='Variable').fill = subhdr_fill
        ws_pval.cell(row=2, column=1).font = subhdr_font
        for j, var in enumerate(variables):
            c = ws_pval.cell(row=2, column=j + 2, value=var)
            c.fill = subhdr_fill; c.font = subhdr_font; c.alignment = center_align

        sig_fill  = PatternFill("solid", fgColor="E2EFDA")   # green tint for sig
        insig_fill = PatternFill("solid", fgColor="FCE4D6")  # red tint for non-sig

        for i, var_i in enumerate(variables):
            c0 = ws_pval.cell(row=i + 3, column=1, value=var_i)
            c0.fill = subhdr_fill; c0.font = subhdr_font; c0.alignment = center_align
            for j in range(len(variables)):
                p_val = pval_matrix[i][j] if pval_matrix else None
                n_val = n_pairs[i][j]   if n_pairs   else None
                if i == j:
                    disp = "—"; fill = PatternFill("solid", fgColor="F2F2F2")
                elif p_val is None:
                    disp = "n/a"; fill = PatternFill("solid", fgColor="F2F2F2")
                else:
                    disp = f"{p_val:.4f}"
                    if n_val: disp += f" (n={n_val})"
                    fill = sig_fill if p_val < 0.05 else insig_fill
                c = ws_pval.cell(row=i + 3, column=j + 2, value=disp)
                c.fill = fill; c.alignment = center_align
                c.border = thin_border; c.font = Font(name='Calibri', size=10)

        ws_pval.column_dimensions[get_column_letter(1)].width = max(
            len(v) for v in variables) + 2
        for j in range(len(variables)):
            ws_pval.column_dimensions[get_column_letter(j + 2)].width = max(
                len(variables[j]) + 2, 14)

        # ── Sheet 3: Scatter Matrix image ─────────────────────────────────────
        if scatter_img:
            ws_scatter = wb.create_sheet('Scatter Matrix')
            ws_scatter.cell(row=1, column=1,
                            value='Scatter Matrix').font = Font(bold=True, size=13, name='Calibri')
            img_data2   = base64.b64decode(scatter_img)
            img_stream2 = io.BytesIO(img_data2)
            img_xl2     = XLImage(img_stream2)
            ws_scatter.add_image(img_xl2, 'A3')

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            download_name='Correlation_Analysis.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# ================== HELPER FUNCTIONS ==================

def _apply_per_var_transforms(df: pd.DataFrame, transform_map: dict, variables: list) -> pd.DataFrame:
    """Apply per-variable transformations in-place on a copy of df.

    transform_map: { varName: 'none'|'log10'|'ln'|'sqrt'|'zscore'|'minmax' }
    Variables not in transform_map are left untouched.
    Invalid transform results (e.g. log of non-positive) become NaN so the
    downstream imputation step can fill them.
    """
    df = df.copy()
    for v in variables:
        method = transform_map.get(v, 'none')
        if not method or method == 'none':
            continue
        col = pd.to_numeric(df[v], errors='coerce')
        if method == 'log10':
            df[v] = col.apply(lambda x: float(np.log10(x)) if pd.notna(x) and x > 0 else np.nan)
        elif method == 'ln':
            df[v] = col.apply(lambda x: float(np.log(x)) if pd.notna(x) and x > 0 else np.nan)
        elif method == 'sqrt':
            df[v] = col.apply(lambda x: float(np.sqrt(x)) if pd.notna(x) and x >= 0 else np.nan)
        elif method == 'zscore':
            valid = col.dropna()
            if len(valid) > 1:
                m, s = valid.mean(), valid.std(ddof=1)
                df[v] = (col - m) / s if s > 0 else 0.0
        elif method == 'minmax':
            valid = col.dropna()
            if len(valid) > 1:
                lo, hi = valid.min(), valid.max()
                df[v] = (col - lo) / (hi - lo) if hi > lo else 0.0
    return df


def _make_json_safe(obj):
    """Convert numpy types (bool_, float64, int64, etc.) to native Python types and sanitize NaN/Inf"""
    if isinstance(obj, np.generic):
        val = obj.item()
        if isinstance(val, float) and (val != val or val == float('inf') or val == float('-inf')):
            return None
        return val
    elif isinstance(obj, float):
        if obj != obj or obj == float('inf') or obj == float('-inf'):
            return None
        return obj
    elif isinstance(obj, dict):
        return {k: _make_json_safe(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_make_json_safe(item) for item in obj]
    elif isinstance(obj, tuple):
        return tuple(_make_json_safe(item) for item in obj)
    return obj


def _welch_anova(data_list):
    """
    Welch's one-way ANOVA (Welch 1951) — handles unequal variances.
    Returns (F, p). Use instead of stats.f_oneway when homogeneity fails but normality holds.
    Formula: F_W = num / denom, with adjusted df2 = (k²-1) / (3·λ)
      where λ = Σ ((1 - wi/W)² / (ni-1))  and  wi = ni/si²
    """
    from scipy.stats import f as f_dist
    k = len(data_list)
    if k < 2:
        return None, None
    n  = np.array([len(d)             for d in data_list], dtype=float)
    m  = np.array([np.mean(d)         for d in data_list])
    v  = np.array([np.var(d, ddof=1)  for d in data_list])
    if np.any(v <= 0):
        return None, None
    w  = n / v
    W  = w.sum()
    grand_mean = (w * m).sum() / W
    lambda_    = ((1 - w / W) ** 2 / (n - 1)).sum()
    num   = (w * (m - grand_mean) ** 2).sum() / (k - 1)
    denom = 1 + (2 * (k - 2) / (k ** 2 - 1)) * lambda_
    F     = num / denom
    df1   = float(k - 1)
    df2   = (k ** 2 - 1) / (3 * lambda_) if lambda_ > 0 else 1e6
    p     = float(1 - f_dist.cdf(F, df1, df2))
    return float(F), p


def _pairwise_posthoc(group_names, group_data, method='mannwhitneyu', correction='fdr_bh'):

    pairs = list(combinations(range(len(group_names)), 2))
    p_values = []
    comparisons = []

    for i, j in pairs:
        g1, g2 = group_names[i], group_names[j]
        d1, d2 = group_data[g1], group_data[g2]

        if method == 'mannwhitneyu':
            _, p = stats.mannwhitneyu(d1, d2, alternative='two-sided')
        else:  # welch
            _, p = stats.ttest_ind(d1, d2, equal_var=False, alternative='two-sided')

        p_values.append(p)
        comparisons.append((g1, g2))

    reject, p_adj, _, _ = multipletests(p_values, alpha=0.05, method=correction)

    return [
        {
            "group1": str(g1),
            "group2": str(g2),
            "p_adj": float(p),
            "significant": bool(r)
        }
        for (g1, g2), p, r in zip(comparisons, p_adj, reject)
    ]


def _assign_letter_groups(group_names, group_data, posthoc_results, var, slice_df):
    """
    Assign compact letter display (CLD) from post-hoc pairwise results.
    Groups that are NOT significantly different share the same letter.
    """
    import string

    n = len(group_names)
    if n == 0:
        return []

    # Build a set of pairs that ARE significantly different
    sig_pairs = set()
    if posthoc_results:
        for ph in posthoc_results:
            if ph["significant"]:
                sig_pairs.add((ph["group1"], ph["group2"]))
                sig_pairs.add((ph["group2"], ph["group1"]))

    # If no posthoc or no significant differences, all get 'a'
    if not sig_pairs:
        result = []
        for g in group_names:
            d = group_data[g]
            result.append({
                "group": str(g),
                "mean": float(np.mean(d)),
                "std": float(np.std(d, ddof=1)) if len(d) > 1 else 0.0,
                "n": int(len(d)),
                "letter": "a"
            })
        return result

    # Greedy CLD algorithm:
    # Start by assigning letter 'a' to all groups.
    # For each pair that is significantly different, ensure they don't share ALL letters.
    # If they do, add a new letter to one of them.

    # Simple approach: absorption algorithm
    # 1. Create initial grouping: each group starts with its own potential set
    # 2. Merge groups that are NOT significantly different

    # Use Union-Find like approach but for CLD
    # Alternative: connected-components of "not-significantly-different" groups
    # Groups in the same connected component share a letter

    # Build adjacency for NOT significant (i.e., similar groups)
    not_sig_adj = {g: set() for g in group_names}
    for i in range(n):
        for j in range(i + 1, n):
            g1, g2 = group_names[i], group_names[j]
            if (g1, g2) not in sig_pairs:
                not_sig_adj[g1].add(g2)
                not_sig_adj[g2].add(g1)

    # Find all maximal cliques of "not significantly different" groups
    # Each clique gets one letter
    letters = list(string.ascii_lowercase)
    group_letters = {g: set() for g in group_names}
    letter_idx = 0

    # Sort groups by mean (descending) for consistent letter assignment
    sorted_groups = sorted(group_names, key=lambda g: -np.mean(group_data[g]))

    assigned = set()
    for g in sorted_groups:
        if g in assigned and group_letters[g]:
            continue
        # Find all groups connected to g (not sig different) that can form a clique
        clique = {g}
        for candidate in sorted_groups:
            if candidate in clique:
                continue
            # candidate must be not-sig-different from ALL current clique members
            if all(candidate in not_sig_adj[member] for member in clique):
                clique.add(candidate)

        # Assign letter to this clique
        if letter_idx < len(letters):
            letter = letters[letter_idx]
        else:
            letter = letters[letter_idx % 26] + str(letter_idx // 26)
        letter_idx += 1

        for member in clique:
            group_letters[member].add(letter)
            assigned.add(member)

    # Check: any group with no letter gets the next available
    for g in group_names:
        if not group_letters[g]:
            if letter_idx < len(letters):
                group_letters[g].add(letters[letter_idx])
            letter_idx += 1

    # Build a lookup dictionary for the calculated statistics
    stats_map = {}
    for g in sorted_groups:
        d = group_data[g]
        stats_map[g] = {
            "group": str(g),
            "mean": float(np.mean(d)),
            "std": float(np.std(d, ddof=1)) if len(d) > 1 else 0.0,
            "n": int(len(d)),
            "letter": "".join(sorted(group_letters[g]))
        }

    # Return the results in the original order provided in group_names
    return [stats_map[g] for g in group_names]

@stats_bp.route('/export-anova-excel', methods=['POST'])
def export_anova_excel():
    try:
        data = request.get_json()
        if not data or 'results' not in data:
            return jsonify({"error": "No data provided"}), 400

        summary_rows = []
        pairwise_rows = []

        for res in data['results']:
            var_name = res.get('variable', 'N/A')
            slice_info = res.get('slice_label', 'All Data')
            test_type = res.get('test_used', 'N/A')
            overall_p = res.get('overall_p', 'N/A')

            # --- NEW: Extract Assumption Results ---
            assumptions = res.get('assumptions', {})
            all_normal = "Yes" if assumptions.get('all_normal') else "No"
            homogeneous = "Yes" if assumptions.get('homogeneous') else "No"

            # 1. Process Summary (Adding Assumption Columns)
            if 'letter_groups' in res:
                for lg in res['letter_groups']:
                    summary_rows.append({
                        "Variable": var_name,
                        "Group": lg.get('group'),
                        "Slice/Subset": slice_info,
                        "Significance Letter": lg.get('letter'),
                        "Test Selection": test_type,
                        "Normality (All Groups)": all_normal,
                        "Homogeneity (Levene)": homogeneous,
                        "Overall p-value": overall_p,
                        "Mean": lg.get('mean'),
                        "Std Dev": lg.get('std'),
                        "N": lg.get('n')
                    })

            # 2. Process Detailed Pairwise
            if 'posthoc' in res:
                for ph in res['posthoc']:
                    pairwise_rows.append({
                        "Variable": var_name,
                        "Slice/Subset": slice_info,
                        "Comparison": f"{ph.get('group1')} vs {ph.get('group2')}",
                        "p-adj": ph.get('p_adj'),
                        "Significant": "Yes" if ph.get('significant') else "No"
                    })

        # Create DataFrames
        df_summary = pd.DataFrame(summary_rows)
        df_pairwise = pd.DataFrame(pairwise_rows)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if not df_summary.empty:
                df_summary.to_excel(writer, index=False, sheet_name='Summary Letters')
            if not df_pairwise.empty:
                df_pairwise.to_excel(writer, index=False, sheet_name='Detailed Pairwise')

            for sheetname in writer.sheets:
                _xl_auto_width(writer.sheets[sheetname])

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name="Significance_Report_With_Assumptions.xlsx"
        )

    except Exception as e:
        print(f"Excel Export Error: {str(e)}")
        return jsonify({"error": str(e)}), 500