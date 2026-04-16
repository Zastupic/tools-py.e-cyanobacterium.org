from flask import Blueprint, render_template, request, jsonify
import re
import io
import os
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from . import UPLOAD_FOLDER
from werkzeug.utils import secure_filename

sigma_bp = Blueprint('sigma_analysis', __name__)

WAVELENGTHS = [440, 480, 540, 590, 625]


# Maps canonical output key → possible header names (case-insensitive, stripped)
_COLUMN_ALIASES = {
    'fo':       ['fo'],
    'i1':       ['i1'],
    'par':      ['par'],
    'p':        ['p'],
    'j':        ['j'],
    'tau':      ['tau'],
    'tau_reox': ['1.r.tau', '1r.tau', 'tau_reox'],
    'sigma':    ['sigma'],
    'error':    ['error'],
}


def _build_col_map(header_parts):
    """
    Given the split header row, return a dict {output_key: column_index}.
    Matching is case-insensitive and ignores surrounding whitespace.
    """
    normalised = [p.strip().lower() for p in header_parts]
    col_map = {}
    for key, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalised:
                col_map[key] = normalised.index(alias)
                break
    return col_map


def _parse_sigma_csv(file_bytes):
    """
    Parse a semicolon-delimited Sigma(II) CSV file.
    Returns a dict with keys: wavelengths, sigma, tau, tau_reox, p, j, fo, i1, par, error.
    Each value is a list ordered by ascending wavelength.
    """
    text = file_bytes.decode('utf-8', errors='replace')
    lines = [l for l in text.splitlines() if l.strip()]

    col_map = {}   # populated from the header row
    rows_by_wl = {}

    for line in lines:
        parts = [p.strip() for p in line.split(';')]

        # Detect header row: first field is 'Nr.' (or empty) and 'Sigma' appears somewhere
        if not parts[0].isdigit():
            candidate = _build_col_map(parts)
            if candidate:   # at least one known column found → treat as header
                col_map = candidate
            continue

        if not col_map:
            continue  # data row before any header was recognised

        # Extract wavelength from Comment (always 3rd field, index 2)
        comment = parts[2] if len(parts) > 2 else ''
        m = re.search(r'(\d{3})nm', comment)
        if not m:
            continue
        wl = int(m.group(1))
        if wl not in WAVELENGTHS:
            continue

        def _f(key):
            idx = col_map.get(key)
            if idx is None:
                return None
            try:
                return float(parts[idx])
            except (IndexError, ValueError):
                return None

        rows_by_wl[wl] = {k: _f(k) for k in _COLUMN_ALIASES}

    # Build ordered lists following WAVELENGTHS order
    result = {k: [] for k in ['wavelengths', 'sigma', 'tau', 'tau_reox', 'p', 'j',
                               'fo', 'i1', 'par', 'error']}
    for wl in WAVELENGTHS:
        if wl in rows_by_wl:
            r = rows_by_wl[wl]
            result['wavelengths'].append(wl)
            for key in ['sigma', 'tau', 'tau_reox', 'p', 'j', 'fo', 'i1', 'par', 'error']:
                result[key].append(r[key])

    return result


def _build_xlsx(samples, charts=None):
    """Build an Excel workbook from a list of sample dicts. Returns bytes.
    Optional charts: list of {title, data_url} dicts with JPEG data URLs."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Sigma(II) Summary'

    header_fill = PatternFill('solid', fgColor='0A3C58')
    header_font = Font(bold=True, color='FFFFFF')

    COLS = [
        ('Sample', 18), ('Wavelength (nm)', 16), ('Sigma(II) (nm²)', 16),
        ('Tau (ms)', 12), ('1.r.Tau (ms)', 13), ('p', 8), ('J', 8),
        ('Fo (V)', 10), ('I1 (V)', 10), ('PAR (µE m⁻² s⁻¹)', 18), ('Error (rel.)', 12),
    ]
    for col_idx, (title, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = width

    row = 2
    for s in samples:
        name = s['name']
        for i, wl in enumerate(s['wavelengths']):
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=wl)
            ws.cell(row=row, column=3, value=s['sigma'][i])
            ws.cell(row=row, column=4, value=s['tau'][i])
            ws.cell(row=row, column=5, value=s['tau_reox'][i])
            ws.cell(row=row, column=6, value=s['p'][i])
            ws.cell(row=row, column=7, value=s['j'][i])
            ws.cell(row=row, column=8, value=s['fo'][i])
            ws.cell(row=row, column=9, value=s['i1'][i])
            ws.cell(row=row, column=10, value=s['par'][i])
            ws.cell(row=row, column=11, value=s['error'][i])
            row += 1

    # ── Charts sheet ──────────────────────────────────────────────────────
    if charts:
        ws_charts = wb.create_sheet('Charts')
        chart_row = 1
        TARGET_W = 700
        for c in charts:
            url = c.get('data_url', '')
            if not url or ',' not in url:
                continue
            b64 = url.split(',', 1)[1]
            if not b64:
                continue
            try:
                img_bytes = base64.b64decode(b64)
            except Exception:
                continue
            if len(img_bytes) < 8:
                continue
            img_buf = io.BytesIO(img_bytes)
            try:
                xl_img = XLImage(img_buf)
            except Exception:
                continue
            orig_w = xl_img.width   # type: ignore[attr-defined]
            orig_h = xl_img.height  # type: ignore[attr-defined]
            if orig_w and orig_w > 0:
                scale = TARGET_W / orig_w
                xl_img.width  = TARGET_W                    # type: ignore[attr-defined]
                xl_img.height = round(orig_h * scale)       # type: ignore[attr-defined]
            else:
                xl_img.width, xl_img.height = TARGET_W, 400  # type: ignore[attr-defined]
            title = c.get('title', '')
            if title:
                ws_charts.cell(row=chart_row, column=1, value=title)
                chart_row += 1
            xl_img.anchor = f'A{chart_row}'
            ws_charts.add_image(xl_img)
            chart_row += round(xl_img.height / 20) + 2     # type: ignore[attr-defined]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@sigma_bp.route('/sigma_analysis')
def sigma_analysis():
    return render_template('sigma_analysis.html')


@sigma_bp.route('/api/sigma_process', methods=['POST'])
def sigma_process():
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'No files uploaded'}), 400

    samples = []
    errors = []

    for f in files:
        if not f.filename:
            continue
        raw_name = os.path.splitext(os.path.basename(f.filename))[0]
        try:
            data = _parse_sigma_csv(f.read())
        except Exception as exc:
            errors.append(f'{raw_name}: {exc}')
            continue
        if not data['wavelengths']:
            errors.append(f'{raw_name}: no recognisable wavelength rows found')
            continue
        samples.append({'name': raw_name, **data})

    if not samples:
        msg = 'Could not parse any files.'
        if errors:
            msg += ' ' + '; '.join(errors)
        return jsonify({'error': msg}), 400

    # Build xlsx and save to uploads folder
    xlsx_bytes = _build_xlsx(samples)
    xlsx_name = 'sigma_summary.xlsx'
    xlsx_path = os.path.join(UPLOAD_FOLDER, xlsx_name)
    with open(xlsx_path, 'wb') as fh:
        fh.write(xlsx_bytes)

    return jsonify({
        'samples': samples,
        'wavelengths': WAVELENGTHS,
        'xlsx_url': f'/static/uploads/{xlsx_name}',
        'errors': errors,
    })


@sigma_bp.route('/api/sigma_export', methods=['POST'])
def sigma_export():
    """Rebuild xlsx with client-captured chart images embedded."""
    data    = request.get_json(force=True)
    samples = data.get('samples', [])
    charts  = data.get('charts', [])

    if not samples:
        return jsonify({'error': 'No sample data provided'}), 400

    try:
        xlsx_bytes = _build_xlsx(samples, charts)
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500

    xlsx_name = 'sigma_summary.xlsx'
    xlsx_path = os.path.join(UPLOAD_FOLDER, xlsx_name)
    with open(xlsx_path, 'wb') as fh:
        fh.write(xlsx_bytes)

    return jsonify({'xlsx_url': f'/static/uploads/{xlsx_name}'})
