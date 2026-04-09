from flask import Blueprint, render_template, request, jsonify
import os, base64, io, time
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from scipy.optimize import curve_fit
from scipy import stats
from . import UPLOAD_FOLDER
from werkzeug.utils import secure_filename

light_curves_analysis = Blueprint('light_curves_analysis', __name__)

# ── protocol definitions ─────────────────────────────────────────────────────
# To add new protocols: append to LC_PROTOCOLS and add a radio button in the template.
LC_PROTOCOLS = {
    'LC1': {'name': 'LC1', 'n_steps': 6,  'par': [10, 20, 50, 100, 300, 500]},
    'LC2': {'name': 'LC2', 'n_steps': 5,  'par': [100, 200, 300, 500, 1000]},
    'LC3': {'name': 'LC3', 'n_steps': 7,  'par': [10, 20, 50, 100, 300, 500, 1000]},
}


# ── helpers ──────────────────────────────────────────────────────────────────

def _safe(v):
    """Convert to float, returning None for NaN/inf. Rounds to 8 decimal places."""
    try:
        f = float(v)
        if np.isnan(f) or np.isinf(f):
            return None
        return round(f, 8)
    except Exception:
        return None


def _cleanup_old_files(upload_folder, max_age_s=1200):
    """Delete files older than max_age_s seconds from upload_folder."""
    try:
        current_time = time.time()
        for fname in os.listdir(upload_folder):
            fpath = os.path.join(upload_folder, fname).replace('\\', '/')
            try:
                if os.stat(fpath).st_mtime < current_time - max_age_s:
                    os.remove(fpath)
            except Exception:
                pass
    except Exception:
        pass


# ── routes ───────────────────────────────────────────────────────────────────

@light_curves_analysis.route('/light_curves_analysis', methods=['GET'])
def analyze_light_curves():
    return render_template('light_curves_analysis.html')


@light_curves_analysis.route('/api/lc_process', methods=['POST'])
def lc_process():
    upload_folder = UPLOAD_FOLDER
    if not os.path.isdir(upload_folder):
        os.mkdir(upload_folder)

    if 'light_curve_files' not in request.files:
        return jsonify({'status': 'error', 'message': 'No files received.'}), 400

    files = request.files.getlist('light_curve_files')
    if not files or secure_filename(files[0].filename or '') == '':
        return jsonify({'status': 'error', 'message': 'Please select one or more files.'}), 400

    fluorometer    = request.form.get('fluorometer', 'AquaPen')
    protocol_key   = request.form.get('protocol', 'LC3')
    etr_max_factor = int(request.form.get('etr_max_factor', 10))

    if len(files) > 50:
        return jsonify({'status': 'error', 'message': 'Maximum 50 files allowed.'}), 400

    if protocol_key not in LC_PROTOCOLS:
        return jsonify({'status': 'error', 'message': f'Unknown protocol: {protocol_key}'}), 400

    protocol     = LC_PROTOCOLS[protocol_key]
    n_steps      = protocol['n_steps']
    PAR          = protocol['par']

    # ── parse files ───────────────────────────────────────────────────────────
    # Merged dataframe structure:
    #   rows: all rows from first file (time_us column + fluorescence column)
    #   additional columns: fluorescence from subsequent files

    import pandas as pd

    Summary_file  = pd.DataFrame()
    file_names    = []
    last_stem     = 'lc'

    for file_number, file in enumerate(files):
        fname_no_ext  = str.lower(os.path.splitext(file.filename or '')[0])
        ext           = str.lower(os.path.splitext(file.filename or '')[1])
        fname_full    = secure_filename(file.filename or '')
        last_stem     = fname_no_ext

        if ext != '.txt':
            return jsonify({'status': 'error',
                            'message': f'Wrong file type for {fname_full}. Expected .txt'}), 400

        raw = file.read().decode('utf-8', errors='replace').splitlines(keepends=True)
        df  = pd.DataFrame(raw)
        df  = df[0].str.split('\t', expand=True).iloc[:, :2]

        # Validate AquaPen/FluorPen file
        if not df[0].astype(str).str.strip().str.contains('FluorPen|AquaPen', case=False).any():
            return jsonify({'status': 'error',
                            'message': f'{fname_full}: not a valid AquaPen/FluorPen file.'}), 400

        if file_number == 0:
            Summary_file = df.rename(columns={df.columns[0]: 'time_us', df.columns[1]: fname_no_ext})
        else:
            Summary_file[fname_no_ext] = df.iloc[:, 1].values

        file_names.append(fname_no_ext)

    if Summary_file.empty:
        return jsonify({'status': 'error', 'message': 'No valid data found in uploaded files.'}), 400

    # ── validate protocol ─────────────────────────────────────────────────────
    # Count Fm_L* rows to determine actual n_steps in the data
    fm_rows = Summary_file['time_us'].astype(str).str.strip().str.match(r'^Fm_L\d+$').sum()
    if fm_rows != n_steps:
        return jsonify({
            'status': 'error',
            'message': (f'Protocol mismatch: found {fm_rows} Fm_L* rows in the data, '
                        f'but protocol {protocol_key} expects {n_steps}. '
                        f'Please select the correct protocol.')
        }), 400

    # Check Fo row present (parameter export)
    if not Summary_file['time_us'].astype(str).str.strip().str.contains('^Fo$', regex=True).any():
        return jsonify({'status': 'error',
                        'message': 'File does not contain exported parameters (Fo row missing). '
                                   'Please export with parameters from AquaPen/FluorPen software.'}), 400

    # ── extract F values ──────────────────────────────────────────────────────
    def _row(label_regex):
        """Return numeric row matching label_regex as DataFrame with file columns."""
        mask = Summary_file['time_us'].astype(str).str.strip().str.match(label_regex)
        return Summary_file[mask].iloc[:, 1:].apply(pd.to_numeric, errors='coerce')

    F0_df  = _row(r'^Fo$').reset_index(drop=True)   # shape (1, n_files)
    FM_df  = _row(r'^Fm$').reset_index(drop=True)   # shape (1, n_files)  — max FM overall

    # Ft_L1..Ft_Ln and Fm_L1..Fm_Ln
    ft_rows, fm_rows_list = [], []
    for k in range(1, n_steps + 1):
        ft_rows.append(_row(rf'^Ft_L{k}$').reset_index(drop=True))
        fm_rows_list.append(_row(rf'^Fm_L{k}$').reset_index(drop=True))

    FTALL = pd.concat(ft_rows, ignore_index=True)   # shape (n_steps, n_files)
    FMALL = pd.concat(fm_rows_list, ignore_index=True)

    # ── raw fluorescence signal (numeric rows only) ───────────────────────────
    good = Summary_file['time_us'].astype(str).str.isnumeric()
    raw_df = Summary_file[good].astype(int).reset_index(drop=True)
    raw_time_us  = raw_df['time_us'].tolist()
    raw_curves   = {fname: raw_df[fname].tolist() for fname in file_names}

    # ── calculate derived parameters ─────────────────────────────────────────
    F0_row   = F0_df.iloc[0]   # Series indexed by file names
    FM_row   = FM_df.iloc[0]   # overall Fm (first Fm row = pre-LC max)

    FMMAX    = FMALL.max()     # max Fm' across steps per file
    FMMAX_df = pd.DataFrame(
        np.repeat(np.asarray(FMMAX.values, dtype=float).reshape(1, -1), n_steps, axis=0), columns=FMALL.columns)
    F0_df2   = pd.DataFrame(
        np.repeat(np.asarray(F0_row.values, dtype=float).reshape(1, -1), n_steps, axis=0), columns=FTALL.columns)

    # Ensure column names are consistent
    FTALL.columns = file_names
    FMALL.columns = file_names
    FMMAX_df.columns = file_names
    F0_df2.columns   = file_names

    QYALL   = (FMALL - FTALL) / FMALL                        # QY = (FM' - Ft) / FM'
    NPQALL  = (FMMAX_df - FMALL) / FMALL                     # NPQ = (FMmax - FM') / FM'
    QP      = (FMALL - FTALL) / (FMALL - F0_df2)             # qP = (FM' - Ft) / (FM' - F0)
    QN      = (FMMAX_df - FMALL) / (FMMAX_df - F0_df2)       # qN = (FMmax - FM') / (FMmax - F0)
    ETRALL  = QYALL.mul(pd.Series(PAR), axis=0)              # rETR = QY * PAR

    # ── fit Platt curves ──────────────────────────────────────────────────────
    def model_platt(x, ETRmPot, alpha, beta):
        return ETRmPot * (1 - np.exp(-(alpha * x / ETRmPot))) * np.exp(-(beta * x / ETRmPot))

    ETRMAX_measured = ETRALL.max()   # max measured ETR per file

    # Check for zero or negative ETR (unfittable data)
    if (ETRMAX_measured <= 0).any():
        bad = ETRMAX_measured[ETRMAX_measured <= 0].index.tolist()
        return jsonify({
            'status': 'error',
            'message': ('Some light curves contain zero or negative ETR values, '
                        f'which cannot be fitted: {bad}. '
                        'Please check the data and select only valid rapid light curves.')
        }), 400

    step_data = {fname: {
        'ft': [], 'fm': [], 'qy': [], 'etr_measured': [],
        'etr_fitted': [], 'npq': [], 'qp': [], 'qn': []
    } for fname in file_names}

    params_out = {}
    par_arr    = np.array(PAR, dtype=float)

    for fname in file_names:
        etr_measured = ETRALL[fname].values.astype(float)
        etr_max_obs  = float(ETRMAX_measured.at[fname])
        etrmPot_init = etr_max_factor * etr_max_obs

        try:
            popt, _ = curve_fit(
                model_platt, par_arr, etr_measured,
                p0=np.array([etrmPot_init, 0.05, 0.05]),
                bounds=((0, 0, 0), (etrmPot_init, 25, 25)),
                maxfev=2000
            )
        except Exception as e:
            return jsonify({'status': 'error',
                            'message': f'Curve fitting failed for {fname}: {e}'}), 400

        ETRmPot_fit, _, _ = popt
        fit_etr = model_platt(par_arr, *popt)

        # Re-derive alpha and beta from slopes of fitted curve
        slope1, _, _, _, _ = stats.linregress(par_arr[:3],   fit_etr[:3])
        slope2, _, _, _, _ = stats.linregress(par_arr[-2:],  fit_etr[-2:])
        alpha     = slope1
        beta      = slope2
        beta_abs  = abs(slope2)

        # ETRmax from alpha/beta formula (Platt 1980; β must be positive)
        if (alpha + beta_abs) > 0 and alpha > 0 and beta_abs > 0:
            etr_max_from_ab = ETRmPot_fit * (alpha / (alpha + beta_abs)) * (beta_abs / (alpha + beta_abs)) ** (beta_abs / alpha)
        else:
            etr_max_from_ab = float('nan')

        Ik = etr_max_from_ab / alpha    if (alpha    != 0 and not np.isnan(etr_max_from_ab)) else float('nan')
        Ib = etr_max_from_ab / beta_abs if (beta_abs != 0 and not np.isnan(etr_max_from_ab)) else float('nan')

        params_out[fname] = {
            'alpha':              _safe(alpha),
            'beta':               _safe(beta_abs),
            'etr_max_measured':   _safe(etr_max_obs),
            'etr_max_from_ab':    _safe(etr_max_from_ab),
            'etr_mpot':           _safe(ETRmPot_fit),
            'ik':                 _safe(Ik),
            'ib':                 _safe(Ib),
        }

        step_data[fname] = {
            'ft':           [_safe(v) for v in FTALL[fname].values],
            'fm':           [_safe(v) for v in FMALL[fname].values],
            'qy':           [_safe(v) for v in QYALL[fname].values],
            'etr_measured': [_safe(v) for v in etr_measured],
            'etr_fitted':   [_safe(v) for v in fit_etr],
            'npq':          [_safe(v) for v in NPQALL[fname].values],
            'qp':           [_safe(v) for v in QP[fname].values],
            'qn':           [_safe(v) for v in QN[fname].values],
        }

    _cleanup_old_files(upload_folder)

    return jsonify({
        'status':           'success',
        'fluorometer':      fluorometer,
        'protocol':         protocol_key,
        'light_intensities': PAR,
        'files':            file_names,
        'file_stem':        last_stem,
        'raw_time_us':      raw_time_us,
        'raw_curves':       raw_curves,
        'step_data':        step_data,
        'params':           params_out,
    })


@light_curves_analysis.route('/api/lc_export', methods=['POST'])
def lc_export():
    """
    Build summary xlsx from client-supplied JSON.
    Receives: {files, step_data, params, raw_curves, light_intensities,
               raw_time_us, file_stem, group_export, charts}
    Returns:  {status, xlsx_path}
    """
    data             = request.get_json(force=True)
    files            = data.get('files', [])
    step_data        = data.get('step_data', {})
    params           = data.get('params', {})
    raw_curves       = data.get('raw_curves', {})
    light_intensities = data.get('light_intensities', [])
    raw_time_us      = data.get('raw_time_us', [])
    file_stem        = secure_filename(data.get('file_stem', 'lc') or 'lc')
    group_export     = data.get('group_export')
    charts           = data.get('charts', [])

    if not file_stem:
        file_stem = 'lc'

    out_fname  = f'{file_stem}_lc_results.xlsx'
    out_path   = os.path.join(UPLOAD_FOLDER, out_fname).replace('\\', '/')
    out_static = f'uploads/{out_fname}'

    try:
        wb = Workbook()

        # ── Parameters sheet ──────────────────────────────────────────────────
        ws_params = wb.worksheets[0]
        ws_params.title = 'Parameters'
        param_keys    = ['alpha', 'beta', 'etr_max_measured', 'etr_max_from_ab',
                         'etr_mpot', 'ik', 'ib']
        param_labels  = {
            'alpha': 'Alpha', 'beta': 'Beta',
            'etr_max_measured': 'ETRmax (measured)',
            'etr_max_from_ab':  'ETRmax (alpha/beta)',
            'etr_mpot':         'ETRmPot',
            'ik':               'Ik', 'ib': 'Ib',
        }
        ws_params.append(['Sample'] + [param_labels.get(k, k) for k in param_keys])
        for fname in files:
            p = params.get(fname, {})
            ws_params.append([fname] + [p.get(k) for k in param_keys])

        # ── Step-data sheets ──────────────────────────────────────────────────
        metric_sheets = [
            ('ETR_measured', 'etr_measured'),
            ('ETR_fitted',   'etr_fitted'),
            ('Ft',           'ft'),
            ('Fm',           'fm'),
            ('QY',           'qy'),
            ('NPQ',          'npq'),
            ('qP',           'qp'),
            ('qN',           'qn'),
        ]
        for sheet_name, key in metric_sheets:
            ws = wb.create_sheet(sheet_name)
            ws.append(['Light_intensity'] + files)
            n_steps = len(light_intensities)
            for i, par in enumerate(light_intensities):
                row = [par]
                for fname in files:
                    vals = (step_data.get(fname) or {}).get(key, [])
                    row.append(vals[i] if i < len(vals) else None)
                ws.append(row)

        # ── Raw fluorescence sheet ────────────────────────────────────────────
        ws_raw = wb.create_sheet('Raw_fluorescence')
        ws_raw.append(['time_us'] + files)
        for i, t in enumerate(raw_time_us):
            row = [t]
            for fname in files:
                vals = raw_curves.get(fname, [])
                row.append(vals[i] if i < len(vals) else None)
            ws_raw.append(row)

        # ── Charts sheet ──────────────────────────────────────────────────────
        ws_charts = wb.create_sheet('Charts')
        row = 1
        for c in charts:
            url = c.get('data_url', '')
            if not url or ',' not in url:
                continue
            b64 = url.split(',', 1)[1]
            if not b64:
                continue
            try:
                img_bytes = base64.b64decode(b64)
            except Exception:
                continue
            if len(img_bytes) < 8:
                continue
            img_buf = io.BytesIO(img_bytes)
            try:
                xl_img = Image(img_buf)
            except Exception:
                continue
            img_buf.seek(0)
            TARGET_W = 700
            orig_w, orig_h = xl_img.width, xl_img.height  # type: ignore[attr-defined]
            if orig_w > 0:
                scale = TARGET_W / orig_w
                xl_img.width  = TARGET_W                  # type: ignore[attr-defined]
                xl_img.height = round(orig_h * scale)     # type: ignore[attr-defined]
            else:
                xl_img.width, xl_img.height = TARGET_W, 400  # type: ignore[attr-defined]
            title = c.get('title', '')
            if title:
                ws_charts.cell(row=row, column=1, value=title)
                row += 1
            xl_img.anchor = f'A{row}'
            ws_charts.add_image(xl_img)
            row += round(xl_img.height / 20) + 2  # type: ignore[attr-defined]

        # ── Group statistics sheets ───────────────────────────────────────────
        if group_export:
            grp_stats    = group_export.get('stats', {})
            samples      = group_export.get('samples', [])
            param_order  = group_export.get('param_order', param_keys)
            grp_labels   = group_export.get('param_labels', param_labels)
            grp_names    = list(grp_stats.keys())

            if grp_stats and param_order:
                ws_st = wb.create_sheet('Group_Statistics')
                hdr   = ['Parameter']
                for g in grp_names:
                    hdr += [f'{g} mean', f'{g} SD', f'{g} N']
                ws_st.append(hdr)
                for p in param_order:
                    stat_row = [grp_labels.get(p, p)]
                    for g in grp_names:
                        s = (grp_stats.get(g) or {}).get('params', {}).get(p)
                        if s:
                            stat_row += [round(s['mean'], 6), round(s['sd'], 6), s.get('n')]
                        else:
                            stat_row += [None, None, None]
                    ws_st.append(stat_row)

            if samples and param_order:
                ws_sp = wb.create_sheet('Group_Samples')
                ws_sp.append(['Sample', 'Group'] + [grp_labels.get(p, p) for p in param_order])
                for sr in samples:
                    sample_row = [sr.get('sample'), sr.get('group')]
                    for p in param_order:
                        v = sr.get(p)
                        sample_row.append(round(v, 6) if v is not None else None)
                    ws_sp.append(sample_row)

        wb.save(out_path)
        _cleanup_old_files(UPLOAD_FOLDER)
        return jsonify({'status': 'success', 'xlsx_path': out_static})

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
