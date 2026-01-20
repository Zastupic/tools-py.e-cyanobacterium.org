from flask import Blueprint, render_template, request, flash
import os, base64, io, time, openpyxl
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from openpyxl.drawing.image import Image
from . import UPLOAD_FOLDER
from time import strftime, localtime
# from flask_login import current_user
from werkzeug.utils import secure_filename

ex_em_spectra_analysis = Blueprint('ex_em_spectra_analysis', __name__)

@ex_em_spectra_analysis.route('/ex_em_spectra_analysis', methods=['GET', 'POST'])
def analyze_ex_em_spectra():
#    if current_user.is_authenticated:
    if request.method == "POST":
        # Define global variables
        max_number_of_files = 50
        upload_folder = UPLOAD_FOLDER
        files_extensions = set()
        excitation_wavelengths = []
        emission_wavelengths = []
        wavelengths_for_norm = []
        files_names = []
        spectra_2d_maps = [] 
        spectra_plot_from_memory = normalized_spectra_plot_from_memory = bar_plot_from_memory = ()
        Ex_Em_spectra_file = param_all = pd.DataFrame()
        Excitation_1 = Excitation_2 = Excitation_3 = Excitation_4 = Excitation_5 = Excitation_6 = Emission_1 = Emission_2 = Emission_3 = Emission_4 = Emission_5 = Emission_6 = pd.DataFrame()
        Excitation_1_norm = Excitation_2_norm = Excitation_3_norm = Excitation_4_norm = Excitation_5_norm = Excitation_6_norm = Emission_1_norm = Emission_2_norm = Emission_3_norm = Emission_4_norm = Emission_5_norm = Emission_6_norm = pd.DataFrame()
        PBS_free_fluo = PBS_PSII_fluo = PBS_PSI_fluo = PBS_tot_fluo = Chl_PSII_fluo = Chl_PSI_fluo = Chl_tot_fluo = pd.Series()
        PBS_free_norm = PBS_PSII_norm = PBS_PSI_norm = Chl_PSII_norm = Chl_PSI_norm = PSII_to_PSI = PBS_PSII_to_PBS_PSI = PC_to_PE = pd.Series()
        fluo_ex_360_all = fluo_ex_440_all = fluo_ex_560_all = fluo_ex_620_all = Chl_tot_fluo_DF = pd.DataFrame()
        fluo_ex_360_em_580 = fluo_ex_360_em_662 = fluo_ex_360_em_689 = fluo_ex_360_em_724 = pd.Series()
        fluo_ex_440_em_580 = fluo_ex_440_em_662 = fluo_ex_440_em_689 = fluo_ex_440_em_724 = pd.Series()
        fluo_ex_560_em_580 = fluo_ex_560_em_662 = fluo_ex_560_em_689 = fluo_ex_560_em_724 = pd.Series()
        fluo_ex_620_em_580 = fluo_ex_620_em_662 = fluo_ex_620_em_689 = fluo_ex_620_em_724 = pd.Series()
        dictionary_ex_em = {
            'Excitation_1': Excitation_1,
            'Excitation_2': Excitation_2,
            'Excitation_3': Excitation_3,
            'Excitation_4': Excitation_4,
            'Excitation_5': Excitation_5,
            'Excitation_6': Excitation_6,
            'Emission_1': Emission_1,
            'Emission_2': Emission_2,
            'Emission_3': Emission_3,
            'Emission_4': Emission_4,
            'Emission_5': Emission_5,
            'Emission_6': Emission_6
            }
        ALLOWED_EXTENSIONS = set(['.csv, .CSV'])
        xlsx_file_path = file_name_without_extension = str('')
        # create upload directory, if there is not any
        if os.path.isdir(upload_folder) == False:
            os.mkdir(upload_folder)
        ################################################
        ### Read Excitations and Emissions from HTML ###
        ################################################
        # Collect selected EXCITATIONS and EMISSIONS
        if str(request.form.get('ex_1')) != "":
            excitation_wavelengths.append(int(str(request.form.get('ex_1'))))
        if str(request.form.get('ex_2')) != "":
            excitation_wavelengths.append(int(str(request.form.get('ex_2'))))
        if str(request.form.get('ex_3')) != "":
            excitation_wavelengths.append(int(str(request.form.get('ex_3'))))
        if str(request.form.get('ex_4')) != "":
            excitation_wavelengths.append(int(str(request.form.get('ex_4'))))
        if str(request.form.get('ex_5')) != "":
            excitation_wavelengths.append(int(str(request.form.get('ex_5'))))
        if str(request.form.get('ex_6')) != "":
            excitation_wavelengths.append(int(str(request.form.get('ex_6'))))
        if str(request.form.get('em_1')) != "":
            emission_wavelengths.append(int(str(request.form.get('em_1'))))
        if str(request.form.get('em_2')) != "":
            emission_wavelengths.append(int(str(request.form.get('em_2'))))
        if str(request.form.get('em_3')) != "":
            emission_wavelengths.append(int(str(request.form.get('em_3'))))
        if str(request.form.get('em_4')) != "":
            emission_wavelengths.append(int(str(request.form.get('em_4'))))
        if str(request.form.get('em_5')) != "":
            emission_wavelengths.append(int(str(request.form.get('em_5'))))
        if str(request.form.get('em_6')) != "":
            emission_wavelengths.append(int(str(request.form.get('em_6'))))
        ##################
        ### Load files ###
        ##################
        ### Check if at least one excitation and emission wavelength was selected
        if len(excitation_wavelengths) > 0 and len(emission_wavelengths) > 0:
            # get the current time
            current_time = time.time()
            current_time_formated = strftime('%Y_%m_%d_%H_%M_%S', localtime(current_time))
            if str(request.form.get('ex_for_norm')) != "" and str(request.form.get('em_for_norm')) != "":
                wavelengths_for_norm.append(int(str(request.form.get('ex_for_norm'))))
                wavelengths_for_norm.append(int(str(request.form.get('em_for_norm'))))
                # check if some file is selected
                if '77K_files' in request.files:
                    # get list of files
                    files = request.files.getlist("77K_files")
                    spectrofluorometer = (request.form.get('spectrofluorometer'))
                    # check if at least one file is selected
                    if secure_filename(files[0].filename) == '': # type: ignore
                        flash('Please select one or more files to analyze.', category='error')
                    else:
                        # limit number of uploaded files
                        if len(files) <= max_number_of_files:
                            file_number = 0
                            #########################
                            #### DO FOR EACH FILE ###
                            #########################
                            for file in files:
                                # get image names and extension
                                file_name_without_extension = str.lower(os.path.splitext(file.filename)[0]) # type: ignore
                                file_extension = str.lower(os.path.splitext(file.filename)[1]) # type: ignore
                                file_name_full = secure_filename(file.filename) # type: ignore
                                files_names.append(file_name_without_extension)
                                # append all extensions to a set
                                files_extensions.add(file_extension)
                                if spectrofluorometer == 'FP-8050 Series Spectrofluorometers (Jasco Inc.)' and '.csv' in files_extensions:
                                    ##########################################
                                    ### Read ALL excitations and Emissions ###
                                    ##########################################
                                    # read csv file directly, without uploading to server: read only rows = results in full number of rows with a single column with values for all columns
                                    Ex_Em_spectra_file = pd.read_csv(files[(file_number)], sep="\0", skip_blank_lines=True, header=None) # type: ignore
                                    # split the single column into multiple columns, according to the length of the file
                                    Ex_Em_spectra_file = pd.DataFrame(np.concatenate([Ex_Em_spectra_file[col].str.split(pat=',', expand=True).add_prefix(col) for col in Ex_Em_spectra_file.columns]))
                                    # get index of a row that contains 'XYDATA'
                                    spectra_start_index = Ex_Em_spectra_file[Ex_Em_spectra_file[0].str.contains('XYDATA')].index.values.astype(int)[0]
                                    # drop header and keep only the fluorescence values
                                    Ex_Em_spectra_file =Ex_Em_spectra_file.iloc[(spectra_start_index+1):,:]
                                    # Convert dataframe strings to numeric
                                    Ex_Em_spectra_file = Ex_Em_spectra_file.apply(pd.to_numeric)
                                    # Replace NA
                                    Ex_Em_spectra_file = Ex_Em_spectra_file.fillna(0)
                                    # reset index and drop first column
                                    Ex_Em_spectra_file.reset_index(inplace = True)
                                    Ex_Em_spectra_file = Ex_Em_spectra_file.drop(Ex_Em_spectra_file.columns[0], axis=1) # drop('index', axis=1)
                                    # Append values to 2D maps
                                    spectra_2d_maps.append((file_name_without_extension, 
                                        Ex_Em_spectra_file.iloc[0, 1:].values,  # excitation wavelengths (first row, skip first col)
                                        Ex_Em_spectra_file.iloc[1:, 0].values,  # emission wavelengths (first col, skip first row)
                                        Ex_Em_spectra_file.iloc[1:, 1:].values))  # intensity matrix
                                    #################################################################
                                    ### Append individual excitations and emissions to dataframes ###
                                    #################################################################
                                    ### Validate selection of Excitations and Emissions: check if ALL excitations and ALL emissions are within the measured range
                                    if set(excitation_wavelengths).issubset(Ex_Em_spectra_file.iloc[0]) and set(emission_wavelengths).issubset(Ex_Em_spectra_file[0]):
                                        ### Get indexes of Excitations and Emissions + Append excitations and emissions to DF + rename columns (Ex) or rows (Em) based on file names
                                        if str(request.form.get('ex_1')) != "":
                                            Excitation_1 = pd.concat([Excitation_1, Ex_Em_spectra_file[np.where(Ex_Em_spectra_file.iloc[0] == int(str(request.form.get('ex_1'))))[0][0]]], axis=1)
                                            Excitation_1.rename(columns={Excitation_1.columns[file_number]: file_name_without_extension}, inplace = True)
                                        if str(request.form.get('ex_2')) != "":
                                            Excitation_2 = pd.concat([Excitation_2, Ex_Em_spectra_file[np.where(Ex_Em_spectra_file.iloc[0] == int(str(request.form.get('ex_2'))))[0][0]]], axis=1)
                                            Excitation_2.rename(columns={Excitation_2.columns[file_number]: file_name_without_extension}, inplace = True)
                                        if str(request.form.get('ex_3')) != "":
                                            Excitation_3 = pd.concat([Excitation_3, Ex_Em_spectra_file[np.where(Ex_Em_spectra_file.iloc[0] == int(str(request.form.get('ex_3'))))[0][0]]], axis=1)
                                            Excitation_3.rename(columns={Excitation_3.columns[file_number]: file_name_without_extension}, inplace = True)
                                        if str(request.form.get('ex_4')) != "":
                                            Excitation_4 = pd.concat([Excitation_4, Ex_Em_spectra_file[np.where(Ex_Em_spectra_file.iloc[0] == int(str(request.form.get('ex_4'))))[0][0]]], axis=1)
                                            Excitation_4.rename(columns={Excitation_4.columns[file_number]: file_name_without_extension}, inplace = True)
                                        if str(request.form.get('ex_5')) != "":
                                            Excitation_5 = pd.concat([Excitation_5, Ex_Em_spectra_file[np.where(Ex_Em_spectra_file.iloc[0] == int(str(request.form.get('ex_5'))))[0][0]]], axis=1)
                                            Excitation_5.rename(columns={Excitation_5.columns[file_number]: file_name_without_extension}, inplace = True)
                                        if str(request.form.get('ex_6')) != "":
                                            Excitation_6 = pd.concat([Excitation_6, Ex_Em_spectra_file[np.where(Ex_Em_spectra_file.iloc[0] == int(str(request.form.get('ex_6'))))[0][0]]], axis=1)
                                            Excitation_6.rename(columns={Excitation_6.columns[file_number]: file_name_without_extension}, inplace = True)
                                        if str(request.form.get('em_1')) != "":
                                            Emission_1 = pd.concat([Emission_1, Ex_Em_spectra_file.iloc[[int(np.where(Ex_Em_spectra_file[0] == int(str(request.form.get('em_1'))))[0][0])]]], axis=0)
                                            Emission_1 = Emission_1.rename({int(np.where(Ex_Em_spectra_file[0] == int(str(request.form.get('em_1'))))[0][0]): str(file_name_without_extension)})
                                        if str(request.form.get('em_2')) != "":
                                            Emission_2 = pd.concat([Emission_2, Ex_Em_spectra_file.iloc[[int(np.where(Ex_Em_spectra_file[0] == int(str(request.form.get('em_2'))))[0][0])]]], axis=0)
                                            Emission_2 = Emission_2.rename({int(np.where(Ex_Em_spectra_file[0] == int(str(request.form.get('em_2'))))[0][0]): str(file_name_without_extension)})
                                        if str(request.form.get('em_3')) != "":
                                            Emission_3 = pd.concat([Emission_3, Ex_Em_spectra_file.iloc[[int(np.where(Ex_Em_spectra_file[0] == int(str(request.form.get('em_3'))))[0][0])]]], axis=0)
                                            Emission_3 = Emission_3.rename({int(np.where(Ex_Em_spectra_file[0] == int(str(request.form.get('em_3'))))[0][0]): str(file_name_without_extension)})
                                        if str(request.form.get('em_4')) != "":
                                            Emission_4 = pd.concat([Emission_4, Ex_Em_spectra_file.iloc[[int(np.where(Ex_Em_spectra_file[0] == int(str(request.form.get('em_4'))))[0][0])]]], axis=0)
                                            Emission_4 = Emission_4.rename({int(np.where(Ex_Em_spectra_file[0] == int(str(request.form.get('em_4'))))[0][0]): str(file_name_without_extension)})
                                        if str(request.form.get('em_5')) != "":
                                            Emission_5 = pd.concat([Emission_5, Ex_Em_spectra_file.iloc[[int(np.where(Ex_Em_spectra_file[0] == int(str(request.form.get('em_5'))))[0][0])]]], axis=0)
                                            Emission_5 = Emission_5.rename({int(np.where(Ex_Em_spectra_file[0] == int(str(request.form.get('em_5'))))[0][0]): str(file_name_without_extension)})
                                        if str(request.form.get('em_6')) != "":
                                            Emission_6 = pd.concat([Emission_6, Ex_Em_spectra_file.iloc[[int(np.where(Ex_Em_spectra_file[0] == int(str(request.form.get('em_6'))))[0][0])]]], axis=0)
                                            Emission_6 = Emission_6.rename({int(np.where(Ex_Em_spectra_file[0] == int(str(request.form.get('em_6'))))[0][0]): str(file_name_without_extension)})
                                    file_number = file_number + 1
                                else:
                                    flash('Please select correct file types for analysis (.csv files for Jasco FP-8050 Series Spectrofluorometers).', category='error')
                            ######################################################################
                            ### Normalize excitations and emissions - only selected wavelngths ###
                            ######################################################################
                            # 1. Add excitation wavelengths as first columns / add emission wavelengths as first rows + transpose emissions
                            # 2. Drop first row with selected wavelength
                            # 3. Save ex / em dataframes to dictionary - to access them later
                            # 4 Normalize excitations / emissions to the selected wavelengths
                            # 5. Add column with wavelengths
                            if set(excitation_wavelengths).issubset(Ex_Em_spectra_file.iloc[0]) and set(emission_wavelengths).issubset(Ex_Em_spectra_file[0]):
                                if Ex_Em_spectra_file.iloc[0].isin([wavelengths_for_norm[0]]).any() and Ex_Em_spectra_file[0].isin([wavelengths_for_norm[1]]).any():
                                    if str(request.form.get('ex_1')) != "":
                                        Excitation_1.insert(loc=0, column='nm', value=Ex_Em_spectra_file[0])
                                        Excitation_1 = Excitation_1.drop([0])
                                        dictionary_ex_em['Excitation_1'] = Excitation_1
                                        index_for_normalization = int(np.where(Excitation_1.iloc[:, 0] == int(wavelengths_for_norm[1]))[0])
                                        fluorecence_at_index = Excitation_1.iloc[index_for_normalization,1:]
                                        Excitation_1_norm = Excitation_1.iloc[:,1:].div(fluorecence_at_index)
                                        Excitation_1_norm.insert(loc=0, column='nm', value=Ex_Em_spectra_file[0])
                                    if str(request.form.get('ex_2')) != "":
                                        Excitation_2.insert(loc=0, column='nm', value=Ex_Em_spectra_file[0])
                                        Excitation_2 = Excitation_2.drop([0])
                                        dictionary_ex_em['Excitation_2'] = Excitation_2
                                        index_for_normalization = int(np.where(Excitation_2.iloc[:, 0] == int(wavelengths_for_norm[1]))[0])
                                        fluorecence_at_index = Excitation_2.iloc[index_for_normalization,1:]
                                        Excitation_2_norm = Excitation_2.iloc[:,1:].div(fluorecence_at_index)
                                        Excitation_2_norm.insert(loc=0, column='nm', value=Ex_Em_spectra_file[0])
                                    if str(request.form.get('ex_3')) != "":
                                        Excitation_3.insert(loc=0, column='nm', value=Ex_Em_spectra_file[0])
                                        Excitation_3 = Excitation_3.drop([0])
                                        dictionary_ex_em['Excitation_3'] = Excitation_3
                                        index_for_normalization = int(np.where(Excitation_3.iloc[:, 0] == int(wavelengths_for_norm[1]))[0])
                                        fluorecence_at_index = Excitation_3.iloc[index_for_normalization,1:]
                                        Excitation_3_norm = Excitation_3.iloc[:,1:].div(fluorecence_at_index)
                                        Excitation_3_norm.insert(loc=0, column='nm', value=Ex_Em_spectra_file[0])
                                    if str(request.form.get('ex_4')) != "":
                                        Excitation_4.insert(loc=0, column='nm', value=Ex_Em_spectra_file[0])
                                        Excitation_4 = Excitation_4.drop([0])
                                        dictionary_ex_em['Excitation_4'] = Excitation_4
                                        index_for_normalization = int(np.where(Excitation_4.iloc[:, 0] == int(wavelengths_for_norm[1]))[0])
                                        fluorecence_at_index = Excitation_4.iloc[index_for_normalization,1:]
                                        Excitation_4_norm = Excitation_4.iloc[:,1:].div(fluorecence_at_index)
                                        Excitation_4_norm.insert(loc=0, column='nm', value=Ex_Em_spectra_file[0])
                                    if str(request.form.get('ex_5')) != "":
                                        Excitation_5.insert(loc=0, column='nm', value=Ex_Em_spectra_file[0])
                                        Excitation_5 = Excitation_5.drop([0])
                                        dictionary_ex_em['Excitation_5'] = Excitation_5
                                        index_for_normalization = int(np.where(Excitation_5.iloc[:, 0] == int(wavelengths_for_norm[1]))[0])
                                        fluorecence_at_index = Excitation_5.iloc[index_for_normalization,1:]
                                        Excitation_5_norm = Excitation_5.iloc[:,1:].div(fluorecence_at_index)
                                        Excitation_5_norm.insert(loc=0, column='nm', value=Ex_Em_spectra_file[0])
                                    if str(request.form.get('ex_6')) != "":
                                        Excitation_6.insert(loc=0, column='nm', value=Ex_Em_spectra_file[0])
                                        Excitation_6 = Excitation_6.drop([0])
                                        dictionary_ex_em['Excitation_6'] = Excitation_6
                                        index_for_normalization = int(np.where(Excitation_6.iloc[:, 0] == int(wavelengths_for_norm[1]))[0])
                                        fluorecence_at_index = Excitation_6.iloc[index_for_normalization,1:]
                                        Excitation_6_norm = Excitation_6.iloc[:,1:].div(fluorecence_at_index)
                                        Excitation_6_norm.insert(loc=0, column='nm', value=Ex_Em_spectra_file[0])
                                    if str(request.form.get('em_1')) != "":
                                        Emission_1.loc['nm'] = Ex_Em_spectra_file.iloc[0]
                                        Emission_1 = Emission_1.T
                                        Emission_1.insert(0, "nm", Emission_1.pop("nm"))
                                        Emission_1 = Emission_1.drop([0])
                                        dictionary_ex_em['Emission_1'] = Emission_1
                                        index_for_normalization = int(np.where(Emission_1.iloc[:, 0] == int(wavelengths_for_norm[0]))[0])
                                        fluorecence_at_index = Emission_1.iloc[index_for_normalization,1:]
                                        Emission_1_norm = Emission_1.iloc[:,1:].div(fluorecence_at_index)
                                        Emission_1_norm.insert(loc=0, column='nm', value=Ex_Em_spectra_file.iloc[0])
                                    if str(request.form.get('em_2')) != "":
                                        Emission_2.loc['nm'] = Ex_Em_spectra_file.iloc[0]
                                        Emission_2 = Emission_2.T
                                        Emission_2.insert(0, "nm", Emission_2.pop("nm"))
                                        Emission_2 = Emission_2.drop([0])
                                        dictionary_ex_em['Emission_2'] = Emission_2
                                        index_for_normalization = int(np.where(Emission_2.iloc[:, 0] == int(wavelengths_for_norm[0]))[0])
                                        fluorecence_at_index = Emission_2.iloc[index_for_normalization,1:]
                                        Emission_2_norm = Emission_2.iloc[:,1:].div(fluorecence_at_index)
                                        Emission_2_norm.insert(loc=0, column='nm', value=Ex_Em_spectra_file.iloc[0])
                                    if str(request.form.get('em_3')) != "":
                                        Emission_3.loc['nm'] = Ex_Em_spectra_file.iloc[0]
                                        Emission_3 = Emission_3.T
                                        Emission_3.insert(0, "nm", Emission_3.pop("nm"))
                                        Emission_3 = Emission_3.drop([0])
                                        dictionary_ex_em['Emission_3'] = Emission_3
                                        index_for_normalization = int(np.where(Emission_3.iloc[:, 0] == int(wavelengths_for_norm[0]))[0])
                                        fluorecence_at_index = Emission_3.iloc[index_for_normalization,1:]
                                        Emission_3_norm = Emission_3.iloc[:,1:].div(fluorecence_at_index)
                                        Emission_3_norm.insert(loc=0, column='nm', value=Ex_Em_spectra_file.iloc[0])
                                    if str(request.form.get('em_4')) != "":
                                        Emission_4.loc['nm'] = Ex_Em_spectra_file.iloc[0]
                                        Emission_4 = Emission_4.T
                                        Emission_4.insert(0, "nm", Emission_4.pop("nm"))
                                        Emission_4 = Emission_4.drop([0])
                                        dictionary_ex_em['Emission_4'] = Emission_4
                                        index_for_normalization = int(np.where(Emission_4.iloc[:, 0] == int(wavelengths_for_norm[0]))[0])
                                        fluorecence_at_index = Emission_4.iloc[index_for_normalization,1:]
                                        Emission_4_norm = Emission_4.iloc[:,1:].div(fluorecence_at_index)
                                        Emission_4_norm.insert(loc=0, column='nm', value=Ex_Em_spectra_file.iloc[0])
                                    if str(request.form.get('em_5')) != "":
                                        Emission_5.loc['nm'] = Ex_Em_spectra_file.iloc[0]
                                        Emission_5 = Emission_5.T
                                        Emission_5.insert(0, "nm", Emission_5.pop("nm"))
                                        Emission_5 = Emission_5.drop([0])
                                        dictionary_ex_em['Emission_5'] = Emission_5
                                        index_for_normalization = int(np.where(Emission_5.iloc[:, 0] == int(wavelengths_for_norm[0]))[0])
                                        fluorecence_at_index = Emission_5.iloc[index_for_normalization,1:]
                                        Emission_5_norm = Emission_5.iloc[:,1:].div(fluorecence_at_index)
                                        Emission_5_norm.insert(loc=0, column='nm', value=Ex_Em_spectra_file.iloc[0])
                                    if str(request.form.get('em_6')) != "":
                                        Emission_6.loc['nm'] = Ex_Em_spectra_file.iloc[0]
                                        Emission_6 = Emission_6.T
                                        Emission_6.insert(0, "nm", Emission_6.pop("nm"))
                                        Emission_6 = Emission_6.drop([0])
                                        dictionary_ex_em['Emission_6'] = Emission_6
                                        index_for_normalization = int(np.where(Emission_6.iloc[:, 0] == int(wavelengths_for_norm[0]))[0])
                                        fluorecence_at_index = Emission_6.iloc[index_for_normalization,1:]
                                        Emission_6_norm = Emission_6.iloc[:,1:].div(fluorecence_at_index)
                                        Emission_6_norm.insert(loc=0, column='nm', value=Ex_Em_spectra_file.iloc[0])
                                    #############################################
                                    ### Calculate PBS-PSII, PBS-PSI, PBS-FREE ###
                                    #############################################
                                    # Get excitations and emissions
                                    if (360 in excitation_wavelengths):
                                        # get dataframe with ex+em wavelengths necessary for the calcualtion of PBS-PSII, PBS-PSI and PBS-free
                                        fluo_ex_360_all = dictionary_ex_em['Excitation_{0}'.format(excitation_wavelengths.index(360)+1)]
                                        # check if correct emissions were measured
                                        if fluo_ex_360_all.iloc[:,0].isin([580]).any():
                                            fluo_ex_360_em_580 = fluo_ex_360_all.iloc[int(np.where(fluo_ex_360_all.iloc[:, 0] == 580)[0]),1:]
                                        if fluo_ex_360_all.iloc[:,0].isin([662]).any():
                                            fluo_ex_360_em_662 = fluo_ex_360_all.iloc[int(np.where(fluo_ex_360_all.iloc[:, 0] == 662)[0]),1:]
                                        if fluo_ex_360_all.iloc[:,0].isin([689]).any():
                                            fluo_ex_360_em_689 = fluo_ex_360_all.iloc[int(np.where(fluo_ex_360_all.iloc[:, 0] == 689)[0]),1:]
                                        if fluo_ex_360_all.iloc[:,0].isin([724]).any():
                                            fluo_ex_360_em_724 = fluo_ex_360_all.iloc[int(np.where(fluo_ex_360_all.iloc[:, 0] == 724)[0]),1:]
                                    if (440 in excitation_wavelengths):
                                        # get dataframe with ex+em wavelengths necessary for the calcualtion of PBS-PSII, PBS-PSI and PBS-free
                                        fluo_ex_440_all = dictionary_ex_em['Excitation_{0}'.format(excitation_wavelengths.index(440)+1)]
                                        # check if correct emissions were measured
                                        if fluo_ex_440_all.iloc[:,0].isin([580]).any():
                                            fluo_ex_440_em_580 = fluo_ex_440_all.iloc[int(np.where(fluo_ex_440_all.iloc[:, 0] == 580)[0]),1:]
                                        if fluo_ex_440_all.iloc[:,0].isin([662]).any():
                                            fluo_ex_440_em_662 = fluo_ex_440_all.iloc[int(np.where(fluo_ex_440_all.iloc[:, 0] == 662)[0]),1:]
                                        if fluo_ex_440_all.iloc[:,0].isin([689]).any():
                                            fluo_ex_440_em_689 = fluo_ex_440_all.iloc[int(np.where(fluo_ex_440_all.iloc[:, 0] == 689)[0]),1:]
                                        if fluo_ex_440_all.iloc[:,0].isin([724]).any():
                                            fluo_ex_440_em_724 = fluo_ex_440_all.iloc[int(np.where(fluo_ex_440_all.iloc[:, 0] == 724)[0]),1:]
                                    if (560 in excitation_wavelengths):
                                        # get dataframe with ex+em wavelengths necessary for the calcualtion of PBS-PSII, PBS-PSI and PBS-free
                                        fluo_ex_560_all = dictionary_ex_em['Excitation_{0}'.format(excitation_wavelengths.index(560)+1)]
                                        # check if correct emissions were measured
                                        if fluo_ex_560_all.iloc[:,0].isin([580]).any():
                                            fluo_ex_560_em_580 = fluo_ex_560_all.iloc[int(np.where(fluo_ex_560_all.iloc[:, 0] == 580)[0]),1:]
                                        if fluo_ex_560_all.iloc[:,0].isin([662]).any():
                                            fluo_ex_560_em_662 = fluo_ex_560_all.iloc[int(np.where(fluo_ex_560_all.iloc[:, 0] == 662)[0]),1:]
                                        if fluo_ex_560_all.iloc[:,0].isin([689]).any():
                                            fluo_ex_560_em_689 = fluo_ex_560_all.iloc[int(np.where(fluo_ex_560_all.iloc[:, 0] == 689)[0]),1:]
                                        if fluo_ex_560_all.iloc[:,0].isin([724]).any():
                                            fluo_ex_560_em_724 = fluo_ex_560_all.iloc[int(np.where(fluo_ex_560_all.iloc[:, 0] == 724)[0]),1:]
                                    if (620 in excitation_wavelengths):
                                        # get dataframe with ex+em wavelengths necessary for the calcualtion of PBS-PSII, PBS-PSI and PBS-free
                                        fluo_ex_620_all = dictionary_ex_em['Excitation_{0}'.format(excitation_wavelengths.index(620)+1)]
                                        # check if correct emissions were measured
                                        if fluo_ex_620_all.iloc[:,0].isin([580]).any():
                                            fluo_ex_620_em_580 = fluo_ex_620_all.iloc[int(np.where(fluo_ex_620_all.iloc[:, 0] == 580)[0]),1:]
                                        if fluo_ex_620_all.iloc[:,0].isin([662]).any():
                                            fluo_ex_620_em_662 = fluo_ex_620_all.iloc[int(np.where(fluo_ex_620_all.iloc[:, 0] == 662)[0]),1:]
                                        if fluo_ex_620_all.iloc[:,0].isin([689]).any():
                                            fluo_ex_620_em_689 = fluo_ex_620_all.iloc[int(np.where(fluo_ex_620_all.iloc[:, 0] == 689)[0]),1:]
                                        if fluo_ex_620_all.iloc[:,0].isin([724]).any():
                                            fluo_ex_620_em_724 = fluo_ex_620_all.iloc[int(np.where(fluo_ex_620_all.iloc[:, 0] == 724)[0]),1:]
                                    # Calculate parameters
                                        Chl_PSII_fluo = fluo_ex_440_em_689
                                        Chl_PSI_fluo = fluo_ex_440_em_724
                                        Chl_tot_fluo = Chl_PSII_fluo + Chl_PSI_fluo
                                        if request.form["checkbox_pigmentation"] == 'checkbox_chl_PC':
                                            PBS_free_fluo = fluo_ex_620_em_662
                                            PBS_PSII_fluo = fluo_ex_620_em_689
                                            PBS_PSI_fluo = fluo_ex_620_em_724
                                        if request.form["checkbox_pigmentation"] == 'checkbox_chl_PE':
                                            PBS_free_fluo = fluo_ex_560_em_662 + fluo_ex_560_em_580
                                            PBS_PSII_fluo = fluo_ex_560_em_689
                                            PBS_PSI_fluo = fluo_ex_560_em_724
                                        if request.form["checkbox_pigmentation"] == 'checkbox_chl_PC_PE':
                                            if (560 in excitation_wavelengths):
                                                PBS_free_fluo = fluo_ex_620_em_662 + fluo_ex_560_em_662 + fluo_ex_560_em_580
                                                PBS_PSII_fluo = fluo_ex_620_em_689 + fluo_ex_560_em_689
                                                PBS_PSI_fluo = fluo_ex_620_em_724 + fluo_ex_560_em_724
                                            else:
                                                PBS_free_fluo = fluo_ex_620_em_662
                                                PBS_PSII_fluo = fluo_ex_620_em_689
                                                PBS_PSI_fluo = fluo_ex_620_em_724
                                    # calculate normalized parameters
                                        Chl_PSII_norm = fluo_ex_440_em_689 / Chl_tot_fluo
                                        Chl_PSI_norm = fluo_ex_440_em_724 / Chl_tot_fluo
                                        PSII_to_PSI = Chl_PSII_norm / Chl_PSI_norm
                                        if request.form["checkbox_pigmentation"] != 'checkbox_chl_only':
                                            PBS_tot_fluo = PBS_free_fluo + PBS_PSII_fluo + PBS_PSI_fluo
                                            PBS_free_norm = PBS_free_fluo / PBS_tot_fluo
                                            PBS_PSII_norm = PBS_PSII_fluo / PBS_tot_fluo
                                            PBS_PSI_norm = PBS_PSI_fluo / PBS_tot_fluo
                                            PBS_PSII_to_PBS_PSI = PBS_PSII_norm / PBS_PSI_norm
                                            if request.form["checkbox_pigmentation"] == 'checkbox_chl_PC_PE':
                                                PC_to_PE = fluo_ex_620_em_662 / (fluo_ex_560_em_662 + fluo_ex_560_em_580)
                                    ###########################################################
                                    ### Plot spectra - EXCITATIONS and EMISSIONS - raw data ###
                                    ###########################################################
                                    ### Check if at least one excitation and emission wavelength was selected
                                    if len(excitation_wavelengths) > 0 and len(emission_wavelengths) > 0:
                                        colors = plt.cm.nipy_spectral(np.linspace(0, 1, file_number+1)) # type: ignore
                                        # Initialise the subplot function using number of rows and columns
                                        fig = plt.figure(figsize=(15,18))
                                        fig.tight_layout() # Shrink to fit the canvas together with legend
                                        fig.subplots_adjust(hspace=0.45) # add horizontal space to read the x-axis and titles well
                                        fig.suptitle("\t\t   Emission spectra \t\t\t\t     Excitation spectra".expandtabs(), x=0.1, y=.91, horizontalalignment='left', verticalalignment='top', fontsize = 15)
                                        ########## Sub-plot ##########
                                        fig_1 = fig.add_subplot(6,3,1)
                                        if str(request.form.get('ex_1')) != "":
                                            for i in range(len(Excitation_1.columns)):
                                                if i > 0:
                                                    fig_1.plot(
                                                        Excitation_1.iloc[:, 0], # x-axis data: 1st column
                                                        Excitation_1.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            fig_1.set_title(f"Excitation {str(request.form.get('ex_1'))} nm: raw data")
                                            fig_1.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_1.set_xlabel('Emission wavelength (nm)')
                                        else:
                                            fig_1.set_title("Excitation 1 not selected")
                                        ########## Sub-plot ##########
                                        fig_2 = fig.add_subplot(6,3,4)
                                        if str(request.form.get('ex_2')) != "":
                                            # Read Excitations_1 curves throughout the dataframe for the plot
                                            for i in range(len(Excitation_2.columns)):
                                                if i > 0:
                                                    fig_2.plot(
                                                        Excitation_2.iloc[:, 0], # x-axis data: 1st column
                                                        Excitation_2.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            fig_2.set_title(f"Excitation {str(request.form.get('ex_2'))} nm: raw data")
                                            fig_2.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_2.set_xlabel('Emission wavelength (nm)')
                                        else:
                                            fig_2.set_title("Excitation 2 not selected")
                                            fig_2.axes.get_xaxis().set_visible(False)
                                            fig_2.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_3 = fig.add_subplot(6,3,7)
                                        fig_3.set_ylabel('Fluorescence (a.u.)')
                                        if str(request.form.get('ex_3')) != "":
                                            for i in range(len(Excitation_3.columns)):
                                                if i > 0:
                                                    fig_3.plot(
                                                        Excitation_3.iloc[:, 0], # x-axis data: 1st column
                                                        Excitation_3.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig 1
                                            fig_3.set_title(f"Excitation {str(request.form.get('ex_3'))} nm: raw data")
                                            fig_3.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_3.set_xlabel('Emission wavelength (nm)')
                                        else:
                                            fig_3.set_title("Excitation 3 not selected")
                                            fig_3.axes.get_xaxis().set_visible(False)
                                            fig_3.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_4 = fig.add_subplot(6,3,10)
                                        if str(request.form.get('ex_4')) != "":
                                            for i in range(len(Excitation_4.columns)):
                                                if i > 0:
                                                    fig_4.plot(
                                                        Excitation_4.iloc[:, 0], # x-axis data: 1st column
                                                        Excitation_4.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            fig_4.set_title(f"Excitation {str(request.form.get('ex_4'))} nm: raw data")
                                            fig_4.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_4.set_xlabel('Emission wavelength (nm)')
                                        else:
                                            fig_4.set_title("Excitation 4 not selected")
                                            fig_4.axes.get_xaxis().set_visible(False)
                                            fig_4.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_5 = fig.add_subplot(6,3,13)
                                        if str(request.form.get('ex_5')) != "":
                                            for i in range(len(Excitation_5.columns)):
                                                if i > 0:
                                                    fig_5.plot(
                                                        Excitation_5.iloc[:, 0], # x-axis data: 1st column
                                                        Excitation_5.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            fig_5.set_title(f"Excitation {str(request.form.get('ex_5'))} nm: raw data")
                                            fig_5.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_5.set_xlabel('Emission wavelength (nm)')
                                        else:
                                            fig_5.set_title("Excitation 5 not selected")
                                            fig_5.axes.get_xaxis().set_visible(False)
                                            fig_5.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_6 = fig.add_subplot(6,3,16)
                                        if str(request.form.get('ex_6')) != "":
                                            for i in range(len(Excitation_6.columns)):
                                                if i > 0:
                                                    fig_6.plot(
                                                        Excitation_6.iloc[:, 0], # x-axis data: 1st column
                                                        Excitation_6.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_6.set_title(f"Excitation {str(request.form.get('ex_6'))} nm: raw data")
                                            fig_6.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_6.set_xlabel('Emission wavelength (nm)')
                                        else:
                                            fig_6.set_title("Excitation 6 not selected")
                                            fig_6.axes.get_xaxis().set_visible(False)
                                            fig_6.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_7 = fig.add_subplot(6,3,2)
                                        if str(request.form.get('em_1')) != "":
                                            for i in range(len(Emission_1.columns)):
                                                if i > 0:
                                                    fig_7.plot(
                                                        Emission_1.iloc[:, 0], # x-axis data: 1st column
                                                        Emission_1.iloc[:, i], # y-axis data
                                                        label = Emission_1.columns[i], # Column names for legend
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_7.set_title(f"Emission {str(request.form.get('em_1'))} nm: raw data")
                                            fig_7.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_7.set_xlabel('Excitation wavelength (nm)')
                                            fig_7.legend(loc='upper left', bbox_to_anchor=(1.1, 1.02))
                                        else:
                                            fig_7.set_title("Emission 1 not selected")
                                            fig_7.axes.get_xaxis().set_visible(False)
                                            fig_7.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_8 = fig.add_subplot(6,3,5)
                                        if str(request.form.get('em_2')) != "":
                                            for i in range(len(Emission_2.columns)):
                                                if i > 0:
                                                    fig_8.plot(
                                                        Emission_2.iloc[:, 0], # x-axis data: 1st column
                                                        Emission_2.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_8.set_title(f"Emission {str(request.form.get('em_2'))} nm: raw data")
                                            fig_8.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_8.set_xlabel('Excitation wavelength (nm)')
                                        else:
                                            fig_8.set_title("Emission 2 not selected")
                                            fig_8.axes.get_xaxis().set_visible(False)
                                            fig_8.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_9 = fig.add_subplot(6,3,8)
                                        if str(request.form.get('em_3')) != "":
                                            for i in range(len(Emission_3.columns)):
                                                if i > 0:
                                                    fig_9.plot(
                                                        Emission_3.iloc[:, 0], # x-axis data: 1st column
                                                        Emission_3.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_9.set_title(f"Emission {str(request.form.get('em_3'))} nm: raw data")
                                            fig_9.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_9.set_xlabel('Excitation wavelength (nm)')
                                        else:
                                            fig_9.set_title("Emission 3 not selected")
                                            fig_9.axes.get_xaxis().set_visible(False)
                                            fig_9.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_10 = fig.add_subplot(6,3,11)
                                        if str(request.form.get('em_4')) != "":
                                            for i in range(len(Emission_4.columns)):
                                                if i > 0:
                                                    fig_10.plot(
                                                        Emission_4.iloc[:, 0], # x-axis data: 1st column
                                                        Emission_4.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_10.set_title(f"Emission {str(request.form.get('em_4'))} nm: raw data")
                                            fig_10.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_10.set_xlabel('Excitation wavelength (nm)')
                                        else:
                                            fig_10.set_title("Emission 4 not selected")
                                            fig_10.axes.get_xaxis().set_visible(False)
                                            fig_10.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_11 = fig.add_subplot(6,3,14)
                                        if str(request.form.get('em_5')) != "":
                                            for i in range(len(Emission_5.columns)):
                                                if i > 0:
                                                    fig_11.plot(
                                                        Emission_5.iloc[:, 0], # x-axis data: 1st column
                                                        Emission_5.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_11.set_title(f"Emission {str(request.form.get('em_5'))} nm: raw data")
                                            fig_11.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_11.set_xlabel('Excitation wavelength (nm)')
                                        else:
                                            fig_11.set_title("Emission 5 not selected")
                                            fig_11.axes.get_xaxis().set_visible(False)
                                            fig_11.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_12 = fig.add_subplot(6,3,17)
                                        if str(request.form.get('em_6')) != "":
                                            for i in range(len(Emission_6.columns)):
                                                if i > 0:
                                                    fig_12.plot(
                                                        Emission_6.iloc[:, 0], # x-axis data: 1st column
                                                        Emission_6.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_12.set_title(f"Emission {str(request.form.get('em_6'))} nm: raw data")
                                            fig_12.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_12.set_xlabel('Excitation wavelength (nm)')
                                        else:
                                            fig_12.set_title("Emission 6 not selected")
                                            fig_12.axes.get_xaxis().set_visible(False)
                                            fig_12.axes.get_yaxis().set_visible(False)
                                        # saving scatter plot to memory
                                        memory_for_spectra_plot = io.BytesIO()
                                        plt.savefig(memory_for_spectra_plot, bbox_inches='tight', format='JPEG')
                                        memory_for_spectra_plot.seek(0)
                                        spectra_plot_in_memory = base64.b64encode(memory_for_spectra_plot.getvalue())
                                        spectra_plot_from_memory = spectra_plot_in_memory.decode('ascii')
                                        # Clearing the plot
                                        plt.clf()
                                        plt.cla()
                                        plt.close()
                                        #############################################################
                                        ### Plot spectra - EXCITATIONS and EMISSIONS - normalized ###
                                        #############################################################
                                        # Initialise the subplot function using number of rows and columns
                                        fig_norm= plt.figure(figsize=(15,18))
                                        fig_norm.tight_layout() # Shrink to fit the canvas together with legend
                                        fig_norm.subplots_adjust(hspace=0.45) # add horizontal space to read the x-axis and titles well
                                        fig_norm.suptitle("\t Normalized emission spectra \t\t   Normalized excitation spectra".expandtabs(), x=0.1, y=.91, horizontalalignment='left', verticalalignment='top', fontsize = 15)
                                        ########## Sub-plot ##########
                                        fig_norm_1 = fig_norm.add_subplot(6,3,1)
                                        if str(request.form.get('ex_1')) != "":
                                            for i in range(len(Excitation_1_norm.columns)):
                                                if i > 0:
                                                    fig_norm_1.plot(
                                                        Excitation_1_norm.iloc[:, 0], # x-axis data: 1st column
                                                        Excitation_1_norm.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            fig_norm_1.set_title(f"Ex. {str(request.form.get('ex_1'))} nm: em. normalized to {str(wavelengths_for_norm[1])} nm")
                                            fig_norm_1.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_norm_1.set_xlabel('Emission wavelength (nm)')
                                        else:
                                            fig_norm_1.set_title("Excitation 1 not selected")
                                            fig_norm_1.axes.get_xaxis().set_visible(False)
                                            fig_norm_1.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_norm_2 = fig_norm.add_subplot(6,3,4)
                                        if str(request.form.get('ex_2')) != "":
                                            # Read Excitations_1 curves throughout the dataframe for the plot
                                            for i in range(len(Excitation_2_norm.columns)):
                                                if i > 0:
                                                    fig_norm_2.plot(
                                                        Excitation_2_norm.iloc[:, 0], # x-axis data: 1st column
                                                        Excitation_2_norm.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            fig_norm_2.set_title(f"Ex.  {str(request.form.get('ex_2'))} nm: em. normalized to {str(wavelengths_for_norm[1])} nm")
                                            fig_norm_2.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_norm_2.set_xlabel('Emission wavelength (nm)')
                                        else:
                                            fig_norm_2.set_title("Excitation 2 not selected")
                                            fig_norm_2.axes.get_xaxis().set_visible(False)
                                            fig_norm_2.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_norm_3 = fig_norm.add_subplot(6,3,7)
                                        fig_norm_3.set_ylabel('Fluorescence (r.u.)')
                                        if str(request.form.get('ex_3')) != "":
                                            for i in range(len(Excitation_3_norm.columns)):
                                                if i > 0:
                                                    fig_norm_3.plot(
                                                        Excitation_3_norm.iloc[:, 0], # x-axis data: 1st column
                                                        Excitation_3_norm.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig 1
                                            fig_norm_3.set_title(f"Ex. {str(request.form.get('ex_3'))} nm: em. normalized to {str(wavelengths_for_norm[1])} nm")
                                            fig_norm_3.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_norm_3.set_xlabel('Emission wavelength (nm)')
                                        else:
                                            fig_norm_3.set_title("Excitation 3 not selected")
                                            fig_norm_3.axes.get_xaxis().set_visible(False)
                                            fig_norm_3.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_norm_4 = fig_norm.add_subplot(6,3,10)
                                        if str(request.form.get('ex_4')) != "":
                                            for i in range(len(Excitation_4_norm.columns)):
                                                if i > 0:
                                                    fig_norm_4.plot(
                                                        Excitation_4_norm.iloc[:, 0], # x-axis data: 1st column
                                                        Excitation_4_norm.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            fig_norm_4.set_title(f"Ex. {str(request.form.get('ex_4'))} nm: em. normalized to {str(wavelengths_for_norm[1])} nm")
                                            fig_norm_4.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_norm_4.set_xlabel('Emission wavelength (nm)')
                                        else:
                                            fig_norm_4.set_title("Excitation 4 not selected")
                                            fig_norm_4.axes.get_xaxis().set_visible(False)
                                            fig_norm_4.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_norm_5 = fig_norm.add_subplot(6,3,13)
                                        if str(request.form.get('ex_5')) != "":
                                            for i in range(len(Excitation_5_norm.columns)):
                                                if i > 0:
                                                    fig_norm_5.plot(
                                                        Excitation_5_norm.iloc[:, 0], # x-axis data: 1st column
                                                        Excitation_5_norm.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            fig_norm_5.set_title(f"Ex. {str(request.form.get('ex_5'))} nm: em. normalized to {str(wavelengths_for_norm[1])} nm")
                                            fig_norm_5.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_norm_5.set_xlabel('Emission wavelength (nm)')
                                        else:
                                            fig_norm_5.set_title("Excitation 5 not selected")
                                            fig_norm_5.axes.get_xaxis().set_visible(False)
                                            fig_norm_5.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_norm_6 = fig_norm.add_subplot(6,3,16)
                                        if str(request.form.get('ex_6')) != "":
                                            for i in range(len(Excitation_6_norm.columns)):
                                                if i > 0:
                                                    fig_norm_6.plot(
                                                        Excitation_6_norm.iloc[:, 0], # x-axis data: 1st column
                                                        Excitation_6_norm.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_norm_6.set_title(f"Ex. {str(request.form.get('ex_6'))} nm: em. normalized to {str(wavelengths_for_norm[1])} nm")
                                            fig_norm_6.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_norm_6.set_xlabel('Emission wavelength (nm)')
                                        else:
                                            fig_norm_6.set_title("Excitation 6 not selected")
                                            fig_norm_6.axes.get_xaxis().set_visible(False)
                                            fig_norm_6.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_norm_7 = fig_norm.add_subplot(6,3,2)
                                        if str(request.form.get('em_1')) != "":
                                            for i in range(len(Emission_1_norm.columns)):
                                                if i > 0:
                                                    fig_norm_7.plot(
                                                        Emission_1_norm.iloc[:, 0], # x-axis data: 1st column
                                                        Emission_1_norm.iloc[:, i], # y-axis data
                                                        label = Emission_1_norm.columns[i], # Column names for legend
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_norm_7.set_title(f"Em. {str(request.form.get('em_1'))} nm: ex. normalized to {str(wavelengths_for_norm[0])} nm")
                                            fig_norm_7.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_norm_7.set_xlabel('Excitation wavelength (nm)')
                                            fig_norm_7.legend(loc='upper left', bbox_to_anchor=(1.1, 1.02))
                                        else:
                                            fig_norm_7.set_title("Emission 1 not selected")
                                            fig_norm_7.axes.get_xaxis().set_visible(False)
                                            fig_norm_7.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_norm_8 = fig_norm.add_subplot(6,3,5)
                                        if str(request.form.get('em_2')) != "":
                                            for i in range(len(Emission_2_norm.columns)):
                                                if i > 0:
                                                    fig_norm_8.plot(
                                                        Emission_2_norm.iloc[:, 0], # x-axis data: 1st column
                                                        Emission_2_norm.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_norm_8.set_title(f"Em. {str(request.form.get('em_2'))} nm: ex. normalized to {str(wavelengths_for_norm[0])} nm")
                                            fig_norm_8.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_norm_8.set_xlabel('Excitation wavelength (nm)')
                                        else:
                                            fig_norm_8.set_title("Emission 2 not selected")
                                            fig_norm_8.axes.get_xaxis().set_visible(False)
                                            fig_norm_8.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_norm_9 = fig_norm.add_subplot(6,3,8)
                                        if str(request.form.get('em_3')) != "":
                                            for i in range(len(Emission_3_norm.columns)):
                                                if i > 0:
                                                    fig_norm_9.plot(
                                                        Emission_3_norm.iloc[:, 0], # x-axis data: 1st column
                                                        Emission_3_norm.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_norm_9.set_title(f"Em. {str(request.form.get('em_3'))} nm: ex. normalized to {str(wavelengths_for_norm[0])} nm")
                                            fig_norm_9.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_norm_9.set_xlabel('Excitation wavelength (nm)')
                                        else:
                                            fig_norm_9.set_title("Emission 3 not selected")
                                            fig_norm_9.axes.get_xaxis().set_visible(False)
                                            fig_norm_9.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_norm_10 = fig_norm.add_subplot(6,3,11)
                                        if str(request.form.get('em_4')) != "":
                                            for i in range(len(Emission_4_norm.columns)):
                                                if i > 0:
                                                    fig_norm_10.plot(
                                                        Emission_4_norm.iloc[:, 0], # x-axis data: 1st column
                                                        Emission_4_norm.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_norm_10.set_title(f"Em. {str(request.form.get('em_4'))} nm: ex. normalized to {str(wavelengths_for_norm[0])} nm")
                                            fig_norm_10.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_norm_10.set_xlabel('Excitation wavelength (nm)')
                                        else:
                                            fig_norm_10.set_title("Emission 4 not selected")
                                            fig_norm_10.axes.get_xaxis().set_visible(False)
                                            fig_norm_10.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_norm_11 = fig_norm.add_subplot(6,3,14)
                                        if str(request.form.get('em_5')) != "":
                                            for i in range(len(Emission_5_norm.columns)):
                                                if i > 0:
                                                    fig_norm_11.plot(
                                                        Emission_5_norm.iloc[:, 0], # x-axis data: 1st column
                                                        Emission_5_norm.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_norm_11.set_title(f"Em. {str(request.form.get('em_5'))} nm: ex. normalized to {str(wavelengths_for_norm[0])} nm")
                                            fig_norm_11.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_norm_11.set_xlabel('Excitation wavelength (nm)')
                                        else:
                                            fig_norm_11.set_title("Emission 5 not selected")
                                            fig_norm_11.axes.get_xaxis().set_visible(False)
                                            fig_norm_11.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_norm_12 = fig_norm.add_subplot(6,3,17)
                                        if str(request.form.get('em_6')) != "":
                                            for i in range(len(Emission_6_norm.columns)):
                                                if i > 0:
                                                    fig_norm_12.plot(
                                                        Emission_6_norm.iloc[:, 0], # x-axis data: 1st column
                                                        Emission_6_norm.iloc[:, i], # y-axis data
                                                        color=colors[i]
                                                        )
                                            # Decorate fig
                                            fig_norm_12.set_title(f"Em. {str(request.form.get('em_6'))} nm: ex. normalized to {str(wavelengths_for_norm[0])} nm")
                                            fig_norm_12.grid(which='both', color='lightgray') # use: which='both' for minor grid
                                            fig_norm_12.set_xlabel('Excitation wavelength (nm)')
                                        else:
                                            fig_norm_12.set_title("Emission 6 not selected")
                                            fig_norm_12.axes.get_xaxis().set_visible(False)
                                            fig_norm_12.axes.get_yaxis().set_visible(False)
                                        # saving scatter plot to memory
                                        memory_for_normalized_spectra_plot = io.BytesIO()
                                        plt.savefig(memory_for_normalized_spectra_plot, bbox_inches='tight', format='JPEG')
                                        memory_for_normalized_spectra_plot.seek(0)
                                        normalized_spectra_plot_in_memory = base64.b64encode(memory_for_normalized_spectra_plot.getvalue())
                                        normalized_spectra_plot_from_memory = normalized_spectra_plot_in_memory.decode('ascii')
                                        # Clearing the plot
                                        plt.clf()
                                        plt.cla()
                                        plt.close()
                                        #######################
                                        ### Plot parameters ###
                                        #######################
                                        # Initialise the subplot function using number of rows and columns
                                        fig_param= plt.figure(figsize=(23,13))
                                        fig_param.tight_layout() # Shrink to fit the canvas together with legend
                                        fig_param.subplots_adjust(hspace=0.4, wspace = 0.4) # add horizontal space to read the x-axis and titles well
                                        fig_param.suptitle("\t\t\t\t\t\t\t\t     Excitaitons+Emissions   \t\t\t\t\t\t\t\t\t  Parameters\t\t Parameters normalized      Parameters total, rel.".expandtabs(), x=0.1, y=.92, horizontalalignment='left', verticalalignment='top', fontsize = 13)
                                        ########## Sub-plot ##########
                                        fig_param_1 = fig_param.add_subplot(5,8,1)
                                        fig_param_1.set_title("Ex 360, Em 580")
                                        if len(fluo_ex_360_em_580) != 0:
                                            fluo_ex_360_em_580.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_1.set_ylabel("a.u.")
                                        else:
                                            fig_param_1.axes.get_xaxis().set_visible(False)
                                            fig_param_1.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_7 = fig_param.add_subplot(5,8,9)
                                        fig_param_7.set_title("Ex 360, Em 662")
                                        if len(fluo_ex_360_em_662) != 0:
                                            fluo_ex_360_em_662.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_7.set_ylabel("a.u.")
                                        else:
                                            fig_param_7.axes.get_xaxis().set_visible(False)
                                            fig_param_7.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_13 = fig_param.add_subplot(5,8,17)
                                        fig_param_13.set_title("Ex 360, Em 689")
                                        if len(fluo_ex_360_em_689) != 0:
                                            fluo_ex_360_em_689.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_13.set_ylabel("a.u.")
                                        else:
                                            fig_param_13.axes.get_xaxis().set_visible(False)
                                            fig_param_13.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_19 = fig_param.add_subplot(5,8,25)
                                        fig_param_19.set_title("Ex 360, Em 724")
                                        if len(fluo_ex_360_em_724) != 0:
                                            fluo_ex_360_em_724.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_19.set_ylabel("a.u.")
                                        else:
                                            fig_param_19.axes.get_xaxis().set_visible(False)
                                            fig_param_19.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_25 = fig_param.add_subplot(5,8,2)
                                        fig_param_25.set_title("Ex 440, Em 580")
                                        if len(fluo_ex_440_em_580) != 0:
                                            fluo_ex_440_em_580.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_25.set_ylabel("a.u.")
                                        else:
                                            fig_param_25.axes.get_xaxis().set_visible(False)
                                            fig_param_25.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_31 = fig_param.add_subplot(5,8,10)
                                        fig_param_31.set_title("Ex 440, Em 662")
                                        if len(fluo_ex_440_em_662) != 0:
                                            fluo_ex_440_em_662.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_31.set_ylabel("a.u.")
                                        else:
                                            fig_param_31.axes.get_xaxis().set_visible(False)
                                            fig_param_31.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_37 = fig_param.add_subplot(5,8,18)
                                        fig_param_37.set_title("Ex 440, Em 689")
                                        if len(fluo_ex_440_em_689) != 0:
                                            fluo_ex_440_em_689.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_37.set_ylabel("a.u.")
                                        else:
                                            fig_param_37.axes.get_xaxis().set_visible(False)
                                            fig_param_37.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_43 = fig_param.add_subplot(5,8,26)
                                        fig_param_43.set_title("Ex 440, Em 724")
                                        if len(fluo_ex_440_em_724) != 0:
                                            fluo_ex_440_em_724.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_43.set_ylabel("a.u.")
                                        else:
                                            fig_param_43.axes.get_xaxis().set_visible(False)
                                            fig_param_43.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_2 = fig_param.add_subplot(5,8,3)
                                        fig_param_2.set_title("Ex 560, Em 580")
                                        if len(fluo_ex_560_em_580) != 0:
                                            fluo_ex_560_em_580.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_2.set_ylabel("a.u.")
                                        else:
                                            fig_param_2.axes.get_xaxis().set_visible(False)
                                            fig_param_2.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_8 = fig_param.add_subplot(5,8,11)
                                        fig_param_8.set_title("Ex 560, Em 662")
                                        if len(fluo_ex_560_em_662) != 0:
                                            fluo_ex_560_em_662.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_8.set_ylabel("a.u.")
                                        else:
                                            fig_param_8.axes.get_xaxis().set_visible(False)
                                            fig_param_8.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_14 = fig_param.add_subplot(5,8,19)
                                        fig_param_14.set_title("Ex 560, Em 689")
                                        if len(fluo_ex_560_em_689) != 0:
                                            fluo_ex_560_em_689.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_14.set_ylabel("a.u.")
                                        else:
                                            fig_param_14.axes.get_xaxis().set_visible(False)
                                            fig_param_14.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_20 = fig_param.add_subplot(5,8,27)
                                        fig_param_20.set_title("Ex 560, Em 724")
                                        if len(fluo_ex_560_em_724) != 0:
                                            fluo_ex_560_em_724.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_20.set_ylabel("a.u.")
                                        else:
                                            fig_param_20.axes.get_xaxis().set_visible(False)
                                            fig_param_20.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_26 = fig_param.add_subplot(5,8,4)
                                        fig_param_26.set_title("Ex 620, Em 580")
                                        if len(fluo_ex_620_em_580) != 0:
                                            fluo_ex_620_em_580.plot.bar(xticks=[], color=colors[1:])
                                        else:
                                            fig_param_26.axes.get_xaxis().set_visible(False)
                                            fig_param_26.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_32 = fig_param.add_subplot(5,8,12)
                                        fig_param_32.set_title("Ex 620, Em 662")
                                        if len(fluo_ex_620_em_662) != 0:
                                            fluo_ex_620_em_662.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_32.set_ylabel("a.u.")
                                        else:
                                            fig_param_32.axes.get_xaxis().set_visible(False)
                                            fig_param_32.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_38 = fig_param.add_subplot(5,8,20)
                                        fig_param_38.set_title("Ex 620, Em 689")
                                        if len(fluo_ex_620_em_689) != 0:
                                            fluo_ex_620_em_689.plot.bar(xticks=[], color=colors[1:])

                                            fig_param_38.set_ylabel("a.u.")
                                        else:
                                            fig_param_38.axes.get_xaxis().set_visible(False)
                                            fig_param_38.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_44 = fig_param.add_subplot(5,8,28)
                                        fig_param_44.set_title("Ex 620, Em 724")
                                        if len(fluo_ex_620_em_724) != 0:
                                            fluo_ex_620_em_724.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_44.set_ylabel("a.u.")
                                        else:
                                            fig_param_44.axes.get_xaxis().set_visible(False)
                                            fig_param_44.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_3 = fig_param.add_subplot(5,8,5)
                                        fig_param_3.set_title("PBS-free")
                                        if len(PBS_free_fluo) != 0:
                                            PBS_free_fluo.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_3.set_ylabel("a.u.")
                                        else:
                                            fig_param_3.axes.get_xaxis().set_visible(False)
                                            fig_param_3.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_9 = fig_param.add_subplot(5,8,13)
                                        fig_param_9.set_title("PBS-PSII")
                                        if len(PBS_PSII_fluo) != 0:
                                            PBS_PSII_fluo.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_9.set_ylabel("a.u.")
                                        else:
                                            fig_param_9.axes.get_xaxis().set_visible(False)
                                            fig_param_9.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_15 = fig_param.add_subplot(5,8,21)
                                        fig_param_15.set_title("PBS-PSI")
                                        if len(PBS_PSI_fluo) != 0:
                                            PBS_PSI_fluo.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_15.set_ylabel("a.u.")
                                        else:
                                            fig_param_15.axes.get_xaxis().set_visible(False)
                                            fig_param_15.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_21 = fig_param.add_subplot(5,8,29)
                                        fig_param_21.set_title("Chl-PSII")
                                        if len(Chl_PSII_fluo) != 0:
                                            Chl_PSII_fluo.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_21.set_ylabel("a.u.")
                                        else:
                                            fig_param_21.axes.get_xaxis().set_visible(False)
                                            fig_param_21.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_27 = fig_param.add_subplot(5,8,37)
                                        fig_param_27.set_title("Chl-PSI")
                                        if len(Chl_PSI_fluo) != 0:
                                            Chl_PSI_fluo.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_27.set_ylabel("a.u.")
                                        else:
                                            fig_param_27.axes.get_xaxis().set_visible(False)
                                            fig_param_27.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_4 = fig_param.add_subplot(5,8,6)
                                        fig_param_4.set_title("PBS-free / PBS-tot")
                                        if len(PBS_free_norm) != 0:
                                            PBS_free_norm.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_4.set_ylabel("r.u.")
                                        else:
                                            fig_param_4.axes.get_xaxis().set_visible(False)
                                            fig_param_4.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_9 = fig_param.add_subplot(5,8,14)
                                        fig_param_9.set_title("PBS-PSII / PBS-tot")
                                        if len(PBS_PSII_norm) != 0:
                                            PBS_PSII_norm.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_9.set_ylabel("r.u.")
                                        else:
                                            fig_param_9.axes.get_xaxis().set_visible(False)
                                            fig_param_9.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_14 = fig_param.add_subplot(5,8,22)
                                        fig_param_14.set_title("PBS-PSI / PBS-tot")
                                        if len(PBS_PSI_norm) != 0:
                                            PBS_PSI_norm.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_14.set_ylabel("r.u.")
                                        else:
                                            fig_param_14.axes.get_xaxis().set_visible(False)
                                            fig_param_14.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_19 = fig_param.add_subplot(5,8,30)
                                        fig_param_19.set_title("Chl-PSII / Chl-tot")
                                        if len(Chl_PSII_norm) != 0:
                                            Chl_PSII_norm.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_19.set_ylabel("r.u.")
                                        else:
                                            fig_param_19.axes.get_xaxis().set_visible(False)
                                            fig_param_19.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_24 = fig_param.add_subplot(5,8,38)
                                        fig_param_24.set_title("Chl-PSI / Chl-tot")
                                        if len(Chl_PSI_norm) != 0:
                                            Chl_PSI_norm.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_24.set_ylabel("r.u.")
                                        else:
                                            fig_param_24.axes.get_xaxis().set_visible(False)
                                            fig_param_24.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_45 = fig_param.add_subplot(5,8,7)
                                        fig_param_45.set_title("Chl-tot")
                                        Chl_tot_fluo_DF = pd.DataFrame([Chl_tot_fluo]) # FM to df, needed for legend
                                        if len(Chl_tot_fluo) != 0:
                                            for i in range(len(Chl_tot_fluo_DF.columns)):
                                                plt.bar(
                                                    Chl_tot_fluo_DF.columns[i], # x-axis data
                                                    Chl_tot_fluo_DF.iloc[:, i], # y-axis data
                                                    label = Chl_tot_fluo_DF.columns[i], # Column names for legend
                                                    color=colors[i+1],
                                                    width = 0.5 # width of the columns
                                                    )
                                            fig_param_45.margins(x=0.42**len(Chl_tot_fluo_DF.columns)) # space between the axes and the first and last bar
                                            fig_param_45.set_ylabel("a.u.")
                                            fig_param_45.set_xticks([]) # no X-axis values
                                            fig_param_45.legend(loc='upper left', bbox_to_anchor=(1.1, 1.02)) # legend
                                        else:
                                            fig_param_45.axes.get_xaxis().set_visible(False)
                                            fig_param_45.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_39 = fig_param.add_subplot(5,8,15)
                                        fig_param_39.set_title("PBS-tot")
                                        if len(PBS_tot_fluo) != 0:
                                            PBS_tot_fluo.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_39.set_ylabel("a.u.")
                                        else:
                                            fig_param_39.axes.get_xaxis().set_visible(False)
                                            fig_param_39.axes.get_yaxis().set_visible(False)

                                        fig_param_29 = fig_param.add_subplot(5,8,23)
                                        fig_param_29.set_title("PBS-PSII / PBS-PSI")
                                        if len(PBS_PSII_to_PBS_PSI) != 0:
                                            PBS_PSII_to_PBS_PSI.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_29.set_ylabel("r.u.")
                                        else:
                                            fig_param_29.axes.get_xaxis().set_visible(False)
                                            fig_param_29.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_34 = fig_param.add_subplot(5,8,31)
                                        fig_param_34.set_title("Chl-PSII / Chl-PSI")
                                        if len(PSII_to_PSI) != 0:
                                            PSII_to_PSI.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_34.set_ylabel("r.u.")
                                        else:
                                            fig_param_34.axes.get_xaxis().set_visible(False)
                                            fig_param_34.axes.get_yaxis().set_visible(False)
                                        ########## Sub-plot ##########
                                        fig_param_39 = fig_param.add_subplot(5,8,39)
                                        fig_param_39.set_title("PC / PE")
                                        if len(PC_to_PE) != 0:
                                            PC_to_PE.plot.bar(xticks=[], color=colors[1:])
                                            fig_param_39.set_ylabel("r.u.")
                                        else:
                                            fig_param_39.axes.get_xaxis().set_visible(False)
                                            fig_param_39.axes.get_yaxis().set_visible(False)
                                        # saving scatter plot to memory
                                        memory_for_bar_plot = io.BytesIO()
                                        plt.savefig(memory_for_bar_plot, bbox_inches='tight', format='JPEG')
                                        memory_for_bar_plot.seek(0)
                                        bar_plot_in_memory = base64.b64encode(memory_for_bar_plot.getvalue())
                                        bar_plot_from_memory = bar_plot_in_memory.decode('ascii')
                                        # Clearing the plot
                                        plt.clf()
                                        plt.cla()
                                        plt.close()
                                        ####################
                                        ### plot 2D maps ###
                                        ####################
                                        if len(spectra_2d_maps) > 0:
                                            # Calculate grid dimensions for subplot
                                            n_files = len(spectra_2d_maps)
                                            n_cols = min(4, n_files)  # Max 4 columns
                                            n_rows = (n_files + n_cols - 1) // n_cols

                                            fig_2d = plt.figure(figsize=(5*n_cols, 4*n_rows))
                                            fig_2d.suptitle("2D Excitation-Emission Maps (Fluorescence Intensity)", fontsize=14, y=1.02)

                                            for idx, (filename, ex_wl, em_wl, intensity) in enumerate(spectra_2d_maps):
                                                ax = fig_2d.add_subplot(n_rows, n_cols, idx + 1)

                                                # Create 2D heatmap using pcolormesh
                                                # X = excitation wavelengths, Y = emission wavelengths, Z = intensity
                                                X, Y = np.meshgrid(ex_wl.astype(float), em_wl.astype(float))

                                                # Plot heatmap with colorbar
                                                c = ax.pcolormesh(X, Y, intensity.astype(float), 
                                                                  shading='auto', 
                                                                  cmap='viridis')  # or 'jet', 'plasma', 'inferno'

                                                ax.set_xlabel('Excitation wavelength (nm)')
                                                ax.set_ylabel('Emission wavelength (nm)')
                                                ax.set_title(filename, fontsize=10)

                                                # Add colorbar
                                                cbar = plt.colorbar(c, ax=ax)
                                                cbar.set_label('Fluorescence (a.u.)')

                                            fig_2d.tight_layout()

                                            # Save 2D map plot to memory
                                            memory_for_2d_plot = io.BytesIO()
                                            plt.savefig(memory_for_2d_plot, bbox_inches='tight', format='JPEG', dpi=150)
                                            memory_for_2d_plot.seek(0)
                                            map_2d_plot_in_memory = base64.b64encode(memory_for_2d_plot.getvalue())
                                            map_2d_plot_from_memory = map_2d_plot_in_memory.decode('ascii')

                                            plt.clf()
                                            plt.cla()
                                            plt.close()
                                        #######################
                                        ### Export to excel ###
                                        #######################
                                        # prepare DF with parameters
                                        param_all = pd.concat([param_all, fluo_ex_360_em_580], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_360_em_662], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_360_em_689], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_360_em_724], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_440_em_580], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_440_em_662], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_440_em_689], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_440_em_724], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_560_em_580], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_560_em_662], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_560_em_689], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_560_em_724], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_620_em_580], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_620_em_662], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_620_em_689], axis = 1)
                                        param_all = pd.concat([param_all, fluo_ex_620_em_724], axis = 1)
                                        param_all = pd.concat([param_all, PBS_free_fluo], axis = 1)
                                        param_all = pd.concat([param_all, PBS_PSII_fluo], axis = 1)
                                        param_all = pd.concat([param_all, PBS_PSI_fluo], axis = 1)
                                        param_all = pd.concat([param_all, Chl_PSII_fluo], axis = 1)
                                        param_all = pd.concat([param_all, Chl_PSI_fluo], axis = 1)
                                        param_all = pd.concat([param_all, PBS_free_norm], axis = 1)
                                        param_all = pd.concat([param_all, PBS_PSII_norm], axis = 1)
                                        param_all = pd.concat([param_all, PBS_PSI_norm], axis = 1)
                                        param_all = pd.concat([param_all, Chl_PSII_norm], axis = 1)
                                        param_all = pd.concat([param_all, Chl_PSI_norm], axis = 1)
                                        param_all = pd.concat([param_all, Chl_tot_fluo], axis = 1)
                                        param_all = pd.concat([param_all, PBS_tot_fluo], axis = 1)
                                        param_all = pd.concat([param_all, PBS_PSII_to_PBS_PSI], axis = 1)
                                        param_all = pd.concat([param_all, PSII_to_PSI], axis = 1)
                                        param_all = pd.concat([param_all, PC_to_PE], axis = 1)
                                        # name columns
                                        param_all.columns = ['Ex360Em580','Ex360Em662', 'Ex360Em689','Ex360Em724',
                                                            'Ex440Em580','Ex440Em662', 'Ex440Em689','Ex440Em724',
                                                            'Ex560Em580','Ex560Em662', 'Ex560Em689','Ex560Em724',
                                                            'Ex620Em580', 'Ex620Em662', 'Ex620Em689','Ex620Em724',
                                                            'PBS-free', 'PBS-PSII','PBS-PSI','Chl-PSII','Chl-PSI',
                                                            'PBS-free_norm','PBS-PSII norm','PBS-PSI_norm','Chl-PSII_norm','Chl-PSI_norm',
                                                            'Chl-tot','PBS-tot','PBS-PSII/PBS-PSI','Chl-PSII/Chl-PSI','PC/PE']

                                        # write all parameters to excel
                                        writer = pd.ExcelWriter(f'{upload_folder}/{file_name_without_extension}.xlsx', engine='openpyxl')
                                        param_all.to_excel(writer, sheet_name = 'Parameters', index=True)
                                        if str(request.form.get('ex_1')) != "":
                                            Excitation_1.to_excel(writer, sheet_name = f"Ex{str(request.form.get('ex_1'))}", index=False)
                                            Excitation_1_norm.to_excel(writer, sheet_name = f"Ex{str(request.form.get('ex_1'))}norm{str(wavelengths_for_norm[1])}", index=False)
                                        if str(request.form.get('ex_2')) != "":
                                            Excitation_2.to_excel(writer, sheet_name = f"Ex{str(request.form.get('ex_2'))}", index=False)
                                            Excitation_2_norm.to_excel(writer, sheet_name = f"Ex{str(request.form.get('ex_2'))}norm{str(wavelengths_for_norm[1])}", index=False)
                                        if str(request.form.get('ex_3')) != "":
                                            Excitation_3.to_excel(writer, sheet_name = f"Ex{str(request.form.get('ex_3'))}", index=False)
                                            Excitation_3_norm.to_excel(writer, sheet_name = f"Ex{str(request.form.get('ex_3'))}norm{str(wavelengths_for_norm[1])}", index=False)
                                        if str(request.form.get('ex_4')) != "":
                                            Excitation_4.to_excel(writer, sheet_name = f"Ex{str(request.form.get('ex_4'))}", index=False)
                                            Excitation_4_norm.to_excel(writer, sheet_name = f"Ex{str(request.form.get('ex_4'))}norm{str(wavelengths_for_norm[1])}", index=False)
                                        if str(request.form.get('ex_5')) != "":
                                            Excitation_5.to_excel(writer, sheet_name = f"Ex{str(request.form.get('ex_5'))}", index=False)
                                            Excitation_5_norm.to_excel(writer, sheet_name = f"Ex{str(request.form.get('ex_5'))}norm{str(wavelengths_for_norm[1])}", index=False)
                                        if str(request.form.get('ex_6')) != "":
                                            Excitation_6.to_excel(writer, sheet_name = f"Ex{str(request.form.get('ex_6'))}", index=False)
                                            Excitation_6_norm.to_excel(writer, sheet_name = f"Ex{str(request.form.get('ex_6'))}norm{str(wavelengths_for_norm[1])}", index=False)
                                        if str(request.form.get('em_1')) != "":
                                            Emission_1.to_excel(writer, sheet_name = f"Em{str(request.form.get('em_1'))}", index=False)
                                            Emission_1_norm.to_excel(writer, sheet_name = f"Em{str(request.form.get('em_1'))}norm{str(wavelengths_for_norm[0])}", index=False)
                                        if str(request.form.get('em_2')) != "":
                                            Emission_2.to_excel(writer, sheet_name = f"Em{str(request.form.get('em_2'))}", index=False)
                                            Emission_2_norm.to_excel(writer, sheet_name = f"Em{str(request.form.get('em_2'))}norm{str(wavelengths_for_norm[0])}", index=False)
                                        if str(request.form.get('em_3')) != "":
                                            Emission_3.to_excel(writer, sheet_name = f"Em{str(request.form.get('em_3'))}", index=False)
                                            Emission_3_norm.to_excel(writer, sheet_name = f"Em{str(request.form.get('em_3'))}norm{str(wavelengths_for_norm[0])}", index=False)
                                        if str(request.form.get('em_4')) != "":
                                            Emission_4.to_excel(writer, sheet_name = f"Em{str(request.form.get('em_4'))}", index=False)
                                            Emission_4_norm.to_excel(writer, sheet_name = f"Em{str(request.form.get('em_4'))}norm{str(wavelengths_for_norm[0])}", index=False)
                                        if str(request.form.get('em_5')) != "":
                                            Emission_5.to_excel(writer, sheet_name = f"Em{str(request.form.get('em_5'))}", index=False)
                                            Emission_5_norm.to_excel(writer, sheet_name = f"Em{str(request.form.get('em_5'))}norm{str(wavelengths_for_norm[0])}", index=False)
                                        if str(request.form.get('em_6')) != "":
                                            Emission_6.to_excel(writer, sheet_name = f"Em{str(request.form.get('em_6'))}", index=False)
                                            Emission_6_norm.to_excel(writer, sheet_name = f"Em{str(request.form.get('em_6'))}norm{str(wavelengths_for_norm[0])}", index=False)
                                        # Close the Pandas Excel writer and output the Excel file.
                                        writer.close()
                                        # Save images
                                        wb = openpyxl.load_workbook(f'{upload_folder}/{file_name_without_extension}.xlsx')
                                        wb.create_sheet(title='Images')
                                        wb.move_sheet('Images', -(len(wb.sheetnames)-1))
                                        ws = wb['Images']
                                        img_parameters = Image(memory_for_bar_plot)
                                        img_data_raw = Image(memory_for_spectra_plot)
                                        img_data_normalized = Image(memory_for_normalized_spectra_plot)
                                        img_map_2d_plot = Image(memory_for_2d_plot)
                                        img_parameters.anchor = 'A1'
                                        img_data_raw.anchor = 'A60'
                                        img_data_normalized.anchor = 'R60'
                                        img_map_2d_plot.anchor = 'AJ1'
                                        ws.add_image(img_parameters)
                                        ws.add_image(img_data_raw)
                                        ws.add_image(img_data_normalized)
                                        ws.add_image(img_map_2d_plot)
                                        wb.save(f'{upload_folder}/{file_name_without_extension}.xlsx')
                                        # save path for html
                                        xlsx_file_path = f'uploads/{file_name_without_extension}.xlsx'
                                        ######################################
                                        ### Delete files older than 20 min ###
                                        ######################################
                                        # List all files
                                        list_of_files_in_upload_folder = os.listdir(upload_folder)
                                        # get number of seconds to reset
                                        seconds = 1200
                                        # scan for old files
                                        for i in list_of_files_in_upload_folder:
                                            # get the location of each file
                                            file_location = os.path.join(upload_folder, str(i)).replace("\\","/")
                                            # get time when the file was modified
                                            file_time = os.stat(file_location).st_mtime
                                            # if a file is modified before 20 min then delete it
                                            if(file_time < current_time - seconds):
                                                os.remove(os.path.join(upload_folder, str(i)).replace("\\","/"))
                                    else:
                                        flash('Please select both excitation and emission wavelengths for the analysis.', category='error')
                                else:
                                    flash('Please select correct excitation and emission wavelengths for normalization (the selected wavelengths are outisde of the measured range).', category='error')
                            else:
                                flash('Please select correct excitation and emission wavelengths for the analysis (some of the selected wavelengths are outisde of the measured range).', category='error')
                        else:
                            flash(f'Please select up to {max_number_of_files} files.', category='error')
                else:
                    flash('Please select .csv files for analysis.', category='error')
            else:
                flash('Please select both excitation and emission wavelengths for normalization.', category='error')
        else:
            flash('Please select at least one excitation and one emission wavelength to analyze.', category='error')

        return render_template("ex_em_spectra_analysis.html",
                        spectra_plot_from_memory = spectra_plot_from_memory,
                        normalized_spectra_plot_from_memory = normalized_spectra_plot_from_memory,
                        bar_plot_from_memory = bar_plot_from_memory,
                        xlsx_file_path = xlsx_file_path,
                        map_2d_plot_from_memory = map_2d_plot_from_memory
                            )

    return render_template("ex_em_spectra_analysis.html")
#    else:
#        flash('Please login', category='error')
#        return redirect("/login")