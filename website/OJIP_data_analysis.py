from flask import Blueprint, render_template, request, jsonify
import os, base64, io
import pandas as pd
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from scipy.interpolate import UnivariateSpline, LSQUnivariateSpline
from scipy.ndimage import gaussian_filter1d
from . import UPLOAD_FOLDER
from werkzeug.utils import secure_filename

OJIP_data_analysis = Blueprint('OJIP_data_analysis', __name__)


# ─── helpers ────────────────────────────────────────────────────────────────

def _ms_factor(fluorometer):
    """Multiplier: native time unit → milliseconds."""
    if fluorometer == 'Aquapen':
        return 0.001        # µs → ms
    if fluorometer == 'FL6000':
        return 1000.0       # s  → ms
    return 1.0              # MULTI-COLOR-PAM already in ms


def _axis_cfg(fluorometer: str) -> tuple[str, str, str, float, set[str], dict[str, tuple[float, float]]]:
    """Return (x_axis_col_name, x_unit_label, y_unit_label, xmin_for_plot, allowed_extensions, search_ranges)."""
    if fluorometer == 'MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)':
        return ('time/ms', 'Time (ms)', 'Fluorescence intensity (V)', 1e-2, {'.csv', '.CSV'},
                dict(FJ=(0.1, 10), FI=(10, 100), FP=(100, 1000)))
    if fluorometer == 'Aquapen':
        return ('time_us', 'Time (μs)', 'Fluorescence intensity (a.u.)', 1e-1, {'.txt'},
                dict(FJ=(100, 10000), FI=(10000, 100000), FP=(100000, 1000000)))
    if fluorometer == 'FL6000':
        return ('time_s', 'Time (s)', 'Fluorescence intensity (a.u.)', 1e-5, {'.txt'},
                dict(FJ=(1e-4, 0.01), FI=(0.01, 0.1), FP=(0.1, 1.0)))
    raise ValueError(f'Unknown fluorometer: {fluorometer}')


def _fit_splines(double_norm_df: pd.DataFrame, x_col: str, kr: int) -> tuple:
    """
    Spline-fit, smooth and compute derivatives for each sample column in double_norm_df.
    Returns (Raw_recon_DF, D1_DF, D2_DF, Resid_DF, Infl_DF, log_time_series).
    """
    dn = double_norm_df
    cols = dn.columns          # [time_col, file1, file2, ...]
    n_files = len(cols) - 1

    log_time = pd.Series(
        np.geomspace(float(dn.iloc[1, 0]), float(dn.iloc[-1, 0]), num=len(dn)),  # type: ignore[arg-type]
        name=cols[0])

    Raw_recon_list, D1_list, D2_list, Infl_list = [], [], [], []

    for i in range(1, n_files + 1):
        fname = cols[i]
        x = dn.iloc[1:, 0].values
        y = dn.iloc[1:, i].values
        knots = UnivariateSpline(x, x, s=0).get_knots()[::kr]
        model = LSQUnivariateSpline(x, y, knots[1:-1], k=3)
        model.set_smoothing_factor(0.5)
        recon = gaussian_filter1d(model(log_time.values), 20)
        d1 = gaussian_filter1d(np.gradient(recon), 20)
        d2 = gaussian_filter1d(np.gradient(d1), 20)
        zc = np.where(np.diff(np.sign(d2)))[0]
        infl = pd.Series(log_time.values[zc]).reset_index(drop=True).rename(fname)
        Raw_recon_list.append(pd.Series(recon, name=fname))
        D1_list.append(pd.Series(d1, name=fname))
        D2_list.append(pd.Series(d2, name=fname))
        Infl_list.append(infl)

    def _assemble(series_list, time_series, col_names):
        df = pd.concat([time_series] + series_list, axis=1)
        df.columns = col_names
        return df

    Raw_recon_DF = _assemble(Raw_recon_list, log_time, cols)
    D1_DF = _assemble(D1_list, log_time, cols)
    D2_DF = _assemble(D2_list, log_time, cols)
    Infl_DF = pd.concat(Infl_list, axis=1)  # columns = sample names

    # Residuals (interpolate reconstructed→raw time axis)
    Resid_list = []
    for i in range(1, n_files + 1):
        interp = np.interp(dn.iloc[:, 0].values, log_time.values, Raw_recon_DF.iloc[:, i].values)
        Resid_list.append(pd.Series(dn.iloc[:, i].values - interp, name=cols[i]))
    Resid_DF = pd.concat([dn.iloc[:, 0].reset_index(drop=True)] + Resid_list, axis=1)
    Resid_DF.columns = cols

    return Raw_recon_DF, D1_DF, D2_DF, Resid_DF, Infl_DF, log_time


def _fit_oj_polynomial(double_norm_df: pd.DataFrame, x_col: str, ms_factor: float,
                       oj_lo_ms: float = 0.5, oj_hi_ms: float = 5.0,
                       n_dense: int = 500) -> dict:
    """
    Fit a 9th-degree polynomial to the O-J region of double_norm data (0.5–5 ms by default).
    Inflection points = roots of d2 where d3 > 0 (Akinyemi et al. 2023).
    Returns {fname: {'poly_oj_time_ms': list, 'poly_oj_d2': list, 'poly_infl_ms': list}}.
    """
    dn = double_norm_df
    cols = dn.columns
    n_files = len(cols) - 1

    oj_lo = oj_lo_ms / ms_factor
    oj_hi = oj_hi_ms / ms_factor
    x_all = np.asarray(dn.iloc[:, 0].values, dtype=float)

    result: dict[str, dict] = {}
    for i in range(1, n_files + 1):
        fname = cols[i]
        y_all = pd.to_numeric(dn.iloc[:, i], errors='coerce').values

        mask  = (x_all >= oj_lo) & (x_all <= oj_hi) & np.isfinite(y_all)
        x_oj  = x_all[mask]
        y_oj  = y_all[mask]

        empty = {'poly_oj_time_ms': [], 'poly_oj_d2': [], 'poly_infl_ms': []}
        if len(x_oj) < 12:
            result[fname] = empty
            continue

        # Work in ms; normalise to [-1, 1] for degree-9 numerical stability
        x_ms = x_oj * ms_factor
        xc   = (x_ms[0] + x_ms[-1]) / 2.0
        xs   = max((x_ms[-1] - x_ms[0]) / 2.0, 1e-12)
        x_n  = (x_ms - xc) / xs

        try:
            coeffs = np.polyfit(x_n, y_oj, 9)
        except Exception:
            result[fname] = empty
            continue

        p   = np.poly1d(coeffs)
        pd2 = p.deriv(2)
        pd3 = p.deriv(3)

        # Dense evaluation grid for smooth display
        x_dense_ms = np.linspace(x_ms[0], x_ms[-1], n_dense)
        x_dense_n  = (x_dense_ms - xc) / xs
        d2_dense   = pd2(x_dense_n)
        d3_dense   = pd3(x_dense_n)

        # Inflection points: zero-crossings of d2 where d3 > 0
        zc = np.where(np.diff(np.sign(d2_dense)))[0]
        inflections: list[float] = []
        for idx in zc:
            nxt = min(idx + 1, len(d3_dense) - 1)
            if (d3_dense[idx] + d3_dense[nxt]) / 2 > 0:
                t_infl = float(np.interp(
                    0,
                    [d2_dense[idx], d2_dense[min(idx + 1, len(d2_dense) - 1)]],
                    [x_dense_ms[idx], x_dense_ms[min(idx + 1, len(x_dense_ms) - 1)]],
                ))
                inflections.append(round(t_infl, 6))

        result[fname] = {
            'poly_oj_time_ms': [round(float(v), 6) for v in x_dense_ms],
            'poly_oj_d2':      [_safe(v) for v in d2_dense],
            'poly_infl_ms':    inflections,
        }
    return result


def _find_fjfifp(D2_DF, x_col, ranges, file_names, Infl_DF):
    """
    Identify FJ/FI/FP derivative minima and inflection points per sample.
    Returns six pd.Series (FJ_deriv, FI_deriv, FP_deriv, FJ_infl, FI_infl, FP_infl)
    indexed by file_names, or raises ValueError on bad data.
    """
    t = D2_DF[x_col]

    def range_idx(lo, hi):
        return t.sub(lo).abs().idxmin(), t.sub(hi).abs().idxmin()

    FJ_lo, FJ_hi = range_idx(*ranges['FJ'])
    FI_lo, FI_hi = range_idx(*ranges['FI'])
    FP_lo, FP_hi = range_idx(*ranges['FP'])

    results = {k: [] for k in ('FJ', 'FI', 'FP')}
    for i in range(1, len(D2_DF.columns)):
        col = D2_DF.iloc[:, i]
        for key, lo, hi in [('FJ', FJ_lo, FJ_hi), ('FI', FI_lo, FI_hi), ('FP', FP_lo, FP_hi)]:
            mn = col.loc[lo:hi].min()
            idx = col.sub(mn).abs().idxmin()
            if pd.isna(idx):
                raise ValueError('Could not identify phase timing — check data integrity.')
            results[key].append(t.iloc[int(idx)])

    FJ_deriv = pd.Series(results['FJ'], index=file_names)
    FI_deriv = pd.Series(results['FI'], index=file_names)
    FP_deriv = pd.Series(results['FP'], index=file_names)

    def nearest_inflect(deriv_ser):
        return pd.Series({
            col: Infl_DF[col][Infl_DF[col] > val].min()
            for col, val in deriv_ser.items()
        })

    return (FJ_deriv, FI_deriv, FP_deriv,
            nearest_inflect(FJ_deriv), nearest_inflect(FI_deriv), nearest_inflect(FP_deriv))


def _calc_areas_fm_timing(Summary_file, data_cols, FJ_idx, FI_idx, ms_factor,
                          F50ms_idx, F100ms_idx, F200ms_idx, F300ms_idx):
    """
    Compute complementary areas (OJ, JI, IP, OP) and FM timing per sample.
    Returns (AREAOJ, AREAJI, AREAIP, AREAOP, FM_timings_series) as pd.Series indexed by data_cols.
    """
    aoj, aji, aip, aop, fm_t = [], [], [], [], {}

    for i, col in enumerate(data_cols, start=1):
        F50ms_v = Summary_file.iloc[F50ms_idx, i]
        F100ms_v = Summary_file.iloc[F100ms_idx, i]
        F200ms_v = Summary_file.iloc[F200ms_idx, i]
        F300ms_v = Summary_file.iloc[F300ms_idx, i]

        if float(F100ms_v) < float(F50ms_v):
            Fm_val = Summary_file.iloc[F100ms_idx:, i].max()
            Fm_idx = Summary_file.iloc[F100ms_idx:, i].sub(Fm_val).abs().idxmin()
            if float(F200ms_v) < float(F100ms_v):
                Fm_val = Summary_file.iloc[F200ms_idx:, i].max()
                Fm_idx = Summary_file.iloc[F200ms_idx:, i].sub(Fm_val).abs().idxmin()
                if float(F300ms_v) < float(F200ms_v):
                    Fm_val = Summary_file.iloc[F300ms_idx:, i].max()
                    Fm_idx = Summary_file.iloc[F300ms_idx:, i].sub(Fm_val).abs().idxmin()
        else:
            Fm_val = Summary_file.iloc[F50ms_idx:, i].max()
            Fm_idx = Summary_file.iloc[F50ms_idx:, i].sub(Fm_val).abs().idxmin()

        fm_t[col] = float(Summary_file.iloc[Fm_idx, 0]) * ms_factor

        x = Summary_file.iloc[:, 0]
        y = Summary_file.iloc[:, i]
        FM_raw = y.iloc[:Fm_idx].max()

        aoj.append(max(x.iloc[:FJ_idx]) * FM_raw - np.trapezoid(y.iloc[:FJ_idx], x.iloc[:FJ_idx]))
        aji.append((max(x.iloc[FJ_idx:FI_idx]) - min(x.iloc[FJ_idx:FI_idx])) * FM_raw
                   - np.trapezoid(y.iloc[FJ_idx:FI_idx], x.iloc[FJ_idx:FI_idx]))
        aip.append((max(x.iloc[FI_idx:Fm_idx]) - min(x.iloc[FI_idx:Fm_idx])) * FM_raw
                   - np.trapezoid(y.iloc[FI_idx:Fm_idx], x.iloc[FI_idx:Fm_idx]))
        aop.append(max(x.iloc[:Fm_idx]) * FM_raw - np.trapezoid(y.iloc[:Fm_idx], x.iloc[:Fm_idx]))

    return (pd.Series(aoj, index=data_cols), pd.Series(aji, index=data_cols),
            pd.Series(aip, index=data_cols), pd.Series(aop, index=data_cols),
            pd.Series(fm_t))


def _safe(v):
    """Convert to float, returning None for NaN/None."""
    try:
        f = float(v)
        return None if np.isnan(f) else round(f, 8)
    except Exception:
        return None


def _t_safe(v, ms_factor):
    """Convert native-unit time value to ms, None on NaN."""
    try:
        f = float(v)
        return None if np.isnan(f) else f * ms_factor
    except Exception:
        return None


# ─── routes ─────────────────────────────────────────────────────────────────

@OJIP_data_analysis.route('/OJIP_data_analysis', methods=['GET'])
def ojip_page():
    return render_template('OJIP_analysis.html')


@OJIP_data_analysis.route('/api/ojip_process', methods=['POST'])
def ojip_process():
    upload_folder = UPLOAD_FOLDER
    if not os.path.isdir(upload_folder):
        os.mkdir(upload_folder)

    if 'OJIP_files' not in request.files:
        return jsonify({'status': 'error', 'message': 'No files received.'}), 400

    files = request.files.getlist('OJIP_files')
    if not files or secure_filename(files[0].filename or '') == '':
        return jsonify({'status': 'error', 'message': 'Please select one or more files.'}), 400

    fluorometer = request.form.get('fluorometer', '')
    kr = int(request.form.get('knots_reduction_factor', 10))
    reduce_size = request.form.get('checkbox_reduce_file_size') == 'checked'
    FJ_time_ms = float(request.form.get('FJ_time', 2.0))
    FI_time_ms = float(request.form.get('FI_time', 30.0))

    if FJ_time_ms >= FI_time_ms:
        return jsonify({'status': 'error', 'message': 'FJ time must be less than FI time.'}), 400
    if len(files) > 100:
        return jsonify({'status': 'error', 'message': 'Maximum 100 files allowed.'}), 400

    try:
        x_col, x_unit, y_unit, _, allowed_ext, ranges = _axis_cfg(fluorometer)
    except ValueError as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400

    ms = _ms_factor(fluorometer)
    FJ_time = FJ_time_ms / ms   # native units
    FI_time = FI_time_ms / ms

    # ── parse files ──────────────────────────────────────────────────────────
    Summary_file = pd.DataFrame()
    file_names_list = []
    last_fname = 'ojip'

    for file_number, file in enumerate(files):
        fname_no_ext = str.lower(os.path.splitext(file.filename or '')[0])
        ext = str.lower(os.path.splitext(file.filename or '')[1])
        fname_full = secure_filename(file.filename or '')
        last_fname = fname_no_ext

        if ext not in allowed_ext:
            return jsonify({'status': 'error',
                            'message': f'Wrong file type for {fname_full}. Expected: {allowed_ext}'}), 400

        if fluorometer == 'MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)':
            df = pd.read_csv(file.stream, sep=';', engine='python')
            if str(df.columns[0]) != 'time/ms':
                return jsonify({'status': 'error',
                                'message': f'{fname_full}: first column must be "time/ms".'}), 400
            if file_number == 0:
                Summary_file = df.iloc[:, 0:2].rename(columns={df.columns[1]: fname_no_ext})
            else:
                Summary_file[fname_no_ext] = df.iloc[:, 1].values

        elif fluorometer in ('Aquapen', 'FL6000'):
            raw = file.read().decode('utf-8', errors='replace').splitlines(keepends=True)
            df = pd.DataFrame(raw)
            df = df[0].str.split('\t', expand=True).iloc[:, :2]

            if fluorometer == 'Aquapen':
                if not df[0].astype(str).str.strip().str.contains('FluorPen|AquaPen', case=False).any():
                    return jsonify({'status': 'error',
                                    'message': f'{fname_full}: not a valid AquaPen/FluorPen file.'}), 400
                if file_number == 0:
                    Summary_file = df.iloc[:, :].rename(columns={df.columns[0]: 'time_us', df.columns[1]: fname_no_ext})
                else:
                    Summary_file[fname_no_ext] = df.iloc[:, 1].values
            else:  # FL6000
                if not df[0].astype(str).str.strip().str.contains('Fluorometer', case=False).any():
                    return jsonify({'status': 'error',
                                    'message': f'{fname_full}: not a valid FL6000 file.'}), 400
                start = df[df[0].str.strip() == 'Time'].index[0] + 1
                df = df.iloc[start:].reset_index(drop=True)
                if file_number == 0:
                    Summary_file = df.rename(columns={df.columns[0]: 'time_s', df.columns[1]: fname_no_ext})
                else:
                    Summary_file[fname_no_ext] = df.iloc[:, 1].values

        file_names_list.append(fname_no_ext)

    # ── clean data ───────────────────────────────────────────────────────────
    if fluorometer == 'Aquapen':
        good = Summary_file['time_us'].astype(str).str.isnumeric()
        Summary_file = Summary_file[good].iloc[1:].astype(int)
    elif fluorometer == 'FL6000':
        Summary_file = Summary_file.rename(columns={Summary_file.columns[0]: 'time_s'})
        for col in Summary_file.columns:
            Summary_file[col] = pd.to_numeric(Summary_file[col], errors='coerce')
        Summary_file = Summary_file.dropna(subset=['time_s'])

    Summary_file = Summary_file.reset_index(drop=True)

    # ── reduce MC-PAM data ───────────────────────────────────────────────────
    # MC-PAM records at 0.01 ms resolution (~73 000 pts/file). For OJIP analysis,
    # 0.1 ms in the fast region and ~1 ms in the slow region is more than sufficient.
    # Target: ≤ 2 000 points per file regardless of original length.
    if fluorometer == 'MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)' and reduce_size:
        # Clip pre-illumination baseline (keep from t ≈ 0.01 ms onward)
        F0_i = int(Summary_file['time/ms'].sub(0.01).abs().idxmin())
        Summary_file = Summary_file.iloc[F0_i:].reset_index(drop=True)
        # Three-zone downsampling preserving the initial steep rise:
        #   0 – 0.5 ms : full 0.01 ms resolution  (~50 pts)  — initial O→J rise intact
        #   0.5 – 30 ms: every 5th pt → 0.05 ms   (~590 pts) — J phase well sampled
        #   30 ms+      : ~200 pts                            — I→P slow phase
        t05_i   = int(Summary_file['time/ms'].sub(0.5).abs().idxmin())
        FI30_i  = int(Summary_file['time/ms'].sub(30).abs().idxmin())
        n_slow  = max(1, len(Summary_file) - FI30_i)
        s_factor = max(1, n_slow // 200)
        very_fast = Summary_file.iloc[:t05_i]
        mid       = Summary_file.iloc[t05_i:FI30_i:5]
        slow      = Summary_file.iloc[FI30_i::s_factor]
        Summary_file = pd.concat([very_fast, mid, slow]).reset_index(drop=True)

    data_cols = list(Summary_file.columns[1:])
    n_files = len(data_cols)

    # ── normalize ────────────────────────────────────────────────────────────
    if fluorometer == 'MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)':
        F0_index = Summary_file[x_col].sub(0.01).abs().idxmin()
    else:
        F0_index = Summary_file[x_col].sub(0).abs().idxmin()

    F0 = Summary_file[data_cols].loc[F0_index]
    FM = Summary_file[data_cols].max()

    OJIP_shifted_to_zero = pd.concat([
        Summary_file.iloc[:, 0],
        Summary_file[data_cols].subtract(F0, axis=1)
    ], axis=1)
    OJIP_shifted_to_max = pd.concat([
        Summary_file.iloc[:, 0],
        Summary_file[data_cols].add(abs(FM - FM.max()), axis=1)
    ], axis=1)
    FMFORNORM = OJIP_shifted_to_zero[data_cols].max()
    OJIP_double_normalized = pd.concat([
        Summary_file.iloc[:, 0],
        OJIP_shifted_to_zero[data_cols].div(FMFORNORM, axis=1)
    ], axis=1)

    # ── spline fitting ───────────────────────────────────────────────────────
    try:
        Raw_recon_DF, D1_DF, D2_DF, Resid_DF, Infl_DF, log_time = _fit_splines(
            OJIP_double_normalized, x_col, kr)
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Spline fitting failed: {e}'}), 400

    poly_oj = _fit_oj_polynomial(OJIP_double_normalized, x_col, ms)                          # FJ window 0.5–5 ms
    poly_oi = _fit_oj_polynomial(OJIP_double_normalized, x_col, ms, oj_lo_ms=10.0, oj_hi_ms=100.0)  # FI window 10–100 ms

    # ── find FJ/FI/FP ────────────────────────────────────────────────────────
    try:
        FJ_deriv, FI_deriv, FP_deriv, FJ_infl, FI_infl, FP_infl = _find_fjfifp(
            D2_DF, x_col, ranges, data_cols, Infl_DF)
    except ValueError as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400

    # ── reference time indexes ───────────────────────────────────────────────
    def tidx(t): return Summary_file[x_col].sub(t).abs().idxmin()

    if fluorometer == 'MULTI-COLOR-PAM / Dual PAM (Heinz Walz GmbH)':
        F50us_idx  = tidx(0.05);  FK_idx     = tidx(0.3)
        F50ms_idx  = tidx(50);    F100ms_idx = tidx(100)
        F200ms_idx = tidx(200);   F300ms_idx = tidx(300)
    elif fluorometer == 'Aquapen':
        F50us_idx  = tidx(50);    FK_idx     = tidx(300)
        F50ms_idx  = tidx(50000); F100ms_idx = tidx(100000)
        F200ms_idx = tidx(200000);F300ms_idx = tidx(300000)
    elif fluorometer == 'FL6000':
        F50us_idx  = tidx(5e-5);  FK_idx     = tidx(3e-4)
        F50ms_idx  = tidx(0.05);  F100ms_idx = tidx(0.1)
        F200ms_idx = tidx(0.2);   F300ms_idx = tidx(0.3)

    FJ_idx = tidx(FJ_time)
    FI_idx = tidx(FI_time)

    F50 = Summary_file[data_cols].loc[F50us_idx]
    FK  = Summary_file[data_cols].loc[FK_idx]
    FJ  = Summary_file[data_cols].loc[FJ_idx]
    FI  = Summary_file[data_cols].loc[FI_idx]

    # ── JIP parameters ───────────────────────────────────────────────────────
    FV      = FM - F0
    FVFM    = FV / FM
    M0      = 4 * (FK - F50) / FV
    VJ      = (FJ - F0) / FV
    VI      = (FI - F0) / FV
    OJ      = FJ - F0
    JI      = FI - FJ
    IP      = FM - FI
    PSIE0   = 1 - VJ
    PSIR0   = 1 - VI
    DELTAR0 = PSIR0 / PSIE0
    PHIE0   = FVFM * PSIE0
    PHIR0   = FVFM * PSIR0
    TR0RC   = M0 / VJ
    ABSRC   = TR0RC / FVFM
    ET0RC   = TR0RC * PSIE0
    RE0RC   = TR0RC * PSIR0
    DI0RC   = ABSRC - TR0RC

    # ── areas + FM timing ────────────────────────────────────────────────────
    AREAOJ, AREAJI, AREAIP, AREAOP, FM_timings_series = _calc_areas_fm_timing(
        Summary_file, data_cols, FJ_idx, FI_idx, ms,
        F50ms_idx, F100ms_idx, F200ms_idx, F300ms_idx)
    SM = AREAOP / FV
    N  = SM * M0 * (1 / VJ)

    # ── build xlsx ───────────────────────────────────────────────────────────
    params_to_concat = [
        F0, FK, FJ, FI, FM, OJ, JI, IP, VJ, VI, M0, PSIE0, PSIR0, DELTAR0, FVFM,
        PHIE0, PHIR0, ABSRC, TR0RC, ET0RC, RE0RC, DI0RC,
        pd.Series(AREAOJ), pd.Series(AREAJI), pd.Series(AREAIP), pd.Series(AREAOP),
        SM, N, FJ_infl, FI_infl, FP_infl, FM_timings_series, FJ_deriv, FI_deriv, FP_deriv
    ]
    OJIP_param_all = pd.concat(params_to_concat, axis=1)
    OJIP_param_all.columns = [
        'Fin', 'FK', 'FJ', 'FI', 'Fmax', 'Amplitude(0-J)', 'Amplitude(J-I)', 'Amplitude(I-P)',
        'VJ', 'VI', 'M0', 'ψE0', 'ψR0', 'δR0', 'ψP0 (Fv/Fm)', 'φE0', 'φR0',
        'ABS/RC', 'TR0/RC', 'ET0/RC', 'RE0/RC', 'DI0/RC',
        'Complementary area O-J', 'Complementary area J-I',
        'Complementary area I-P', 'Complementary area (O-P)',
        'Normalized complementary area Sm', 'N (turn-over number QA)',
        'Time FJ', 'Time FI', 'Time FP', 'Time FM',
        'Time Min 2nd Deriv Pre-FJ', 'Time Min 2nd Deriv Pre-FI', 'Time Min 2nd Deriv Pre-FP'
    ]

    # ── build JSON payload ───────────────────────────────────────────────────
    time_raw_ms = (Summary_file.iloc[:, 0].astype(float) * ms).tolist()
    time_log_ms = (log_time.astype(float) * ms).tolist()

    curves = {}
    for i, fname in enumerate(data_cols, start=1):
        curves[fname] = {
            'raw':          [_safe(v) for v in Summary_file.iloc[:, i]],
            'shifted_F0':   [_safe(v) for v in OJIP_shifted_to_zero.iloc[:, i]],
            'shifted_FM':   [_safe(v) for v in OJIP_shifted_to_max.iloc[:, i]],
            'double_norm':  [_safe(v) for v in OJIP_double_normalized.iloc[:, i]],
            'residuals':    [_safe(v) for v in Resid_DF.iloc[:, i]],
            'reconstructed':[_safe(v) for v in Raw_recon_DF.iloc[:, i]],
            'd1':              [_safe(v) for v in D1_DF.iloc[:, i]],
            'd2':              [_safe(v) for v in D2_DF.iloc[:, i]],
            'poly_oj_time_ms': poly_oj[fname]['poly_oj_time_ms'],
            'poly_oj_d2':      poly_oj[fname]['poly_oj_d2'],
            'poly_oi_time_ms': poly_oi[fname]['poly_oj_time_ms'],
            'poly_oi_d2':      poly_oi[fname]['poly_oj_d2'],
        }

    key_values = {}
    for fname in data_cols:
        key_values[fname] = {
            'F0':  _safe(F0[fname]),  'FM': _safe(FM[fname]),
            'FK':  _safe(FK[fname]),  'F50': _safe(F50[fname]),
            'FJ':  _safe(FJ[fname]),  'FI':  _safe(FI[fname]),
            'FJ_time_user_ms':    FJ_time_ms,
            'FI_time_user_ms':    FI_time_ms,
            'FJ_time_deriv_ms':   _t_safe(FJ_deriv.get(fname), ms),
            'FI_time_deriv_ms':   _t_safe(FI_deriv.get(fname), ms),
            'FP_time_deriv_ms':   _t_safe(FP_deriv.get(fname), ms),
            'FJ_time_inflect_ms': _t_safe(FJ_infl.get(fname),  ms),
            'FI_time_inflect_ms': _t_safe(FI_infl.get(fname),  ms),
            'FP_time_inflect_ms': _t_safe(FP_infl.get(fname),  ms),
            'FM_time_ms':  _safe(FM_timings_series.get(fname)),
            'Area_OJ': _safe(AREAOJ[fname]), 'Area_JI': _safe(AREAJI[fname]),
            'Area_IP': _safe(AREAIP[fname]), 'Area_OP': _safe(AREAOP[fname]),
            'poly_infl_ms':    poly_oj[fname]['poly_infl_ms'],
            'poly_fi_infl_ms': poly_oi[fname]['poly_infl_ms'],
        }

    return jsonify({
        'status':      'success',
        'fluorometer': fluorometer,
        'kr':          kr,
        'fj_time_ms':  FJ_time_ms,
        'fi_time_ms':  FI_time_ms,
        'files':       data_cols,
        'file_stem':   last_fname,
        'time_raw_ms': time_raw_ms,
        'time_log_ms': time_log_ms,
        'curves':      curves,
        'key_values':  key_values,
    })


@OJIP_data_analysis.route('/api/ojip_refit', methods=['POST'])
def ojip_refit():
    """
    Re-fit splines with a new kr value.
    Receives JSON: {fluorometer, kr, fj_time_ms, fi_time_ms, time_raw_ms, double_norm: {file:[...]}}
    Returns: {curves: {file:{reconstructed,d1,d2}}, key_timings: {file:{...}}}
    """
    data = request.get_json(force=True)
    fluorometer = data.get('fluorometer', '')
    kr = int(data.get('kr', 10))
    FJ_time_ms = float(data.get('fj_time_ms', 2.0))
    FI_time_ms = float(data.get('fi_time_ms', 30.0))
    time_raw_ms = data['time_raw_ms']
    double_norm_dict = data['double_norm']  # {file: [y values]}
    file_names = list(double_norm_dict.keys())

    try:
        x_col, _, _, _, _, ranges = _axis_cfg(fluorometer)
    except ValueError as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400

    ms = _ms_factor(fluorometer)
    time_native = [t / ms for t in time_raw_ms]

    # Reconstruct double_norm DataFrame in native time units
    dn_df = pd.DataFrame({x_col: time_native})
    for fname, vals in double_norm_dict.items():
        dn_df[fname] = vals

    try:
        Raw_recon_DF, D1_DF, D2_DF, Resid_DF, Infl_DF, log_time = _fit_splines(dn_df, x_col, kr)
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Refit failed: {e}'}), 400

    poly_oj = _fit_oj_polynomial(dn_df, x_col, ms)
    poly_oi = _fit_oj_polynomial(dn_df, x_col, ms, oj_lo_ms=10.0, oj_hi_ms=100.0)

    try:
        FJ_deriv, FI_deriv, FP_deriv, FJ_infl, FI_infl, FP_infl = _find_fjfifp(
            D2_DF, x_col, ranges, file_names, Infl_DF)
    except ValueError as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400

    time_log_ms = (log_time.astype(float) * ms).tolist()
    updated_curves = {}
    for i, fname in enumerate(file_names, start=1):
        updated_curves[fname] = {
            'reconstructed': [_safe(v) for v in Raw_recon_DF.iloc[:, i]],
            'd1':            [_safe(v) for v in D1_DF.iloc[:, i]],
            'd2':            [_safe(v) for v in D2_DF.iloc[:, i]],
            'residuals':       [_safe(v) for v in Resid_DF.iloc[:, i]],
            'poly_oj_time_ms': poly_oj[fname]['poly_oj_time_ms'],
            'poly_oj_d2':      poly_oj[fname]['poly_oj_d2'],
            'poly_oi_time_ms': poly_oi[fname]['poly_oj_time_ms'],
            'poly_oi_d2':      poly_oi[fname]['poly_oj_d2'],
        }

    key_timings = {}
    for fname in file_names:
        key_timings[fname] = {
            'FJ_time_deriv_ms':   _t_safe(FJ_deriv.get(fname), ms),
            'FI_time_deriv_ms':   _t_safe(FI_deriv.get(fname), ms),
            'FP_time_deriv_ms':   _t_safe(FP_deriv.get(fname), ms),
            'FJ_time_inflect_ms': _t_safe(FJ_infl.get(fname),  ms),
            'FI_time_inflect_ms': _t_safe(FI_infl.get(fname),  ms),
            'FP_time_inflect_ms': _t_safe(FP_infl.get(fname),  ms),
            'poly_infl_ms':    poly_oj[fname]['poly_infl_ms'],
            'poly_fi_infl_ms': poly_oi[fname]['poly_infl_ms'],
        }

    return jsonify({
        'status':       'success',
        'time_log_ms':  time_log_ms,
        'curves':       updated_curves,
        'key_timings':  key_timings,
    })


@OJIP_data_analysis.route('/api/ojip_add_charts', methods=['POST'])
def ojip_add_charts():
    """
    Create a compact summary xlsx: Parameters + Charts + Group Statistics.
    All data is received from the client — no server-side xlsx file is read.
    """
    data         = request.get_json(force=True)
    charts       = data.get('charts', [])
    file_stem    = secure_filename(data.get('file_stem', 'ojip'))
    group_export = data.get('group_export')
    params_table = data.get('params_table')   # {header: [...], rows: [[...]]}

    summary_fname  = f'{file_stem}_summary.xlsx'
    summary_full   = os.path.join(UPLOAD_FOLDER, summary_fname)
    summary_static = f'uploads/{summary_fname}'

    try:
        wb = Workbook()

        # ── 1. Parameters sheet (data passed from client) ─────────────────────
        ws_params = wb.worksheets[0]
        ws_params.title = 'Parameters'
        if params_table:
            ws_params.append(params_table.get('header', []))
            for r in params_table.get('rows', []):
                ws_params.append(r)

        # ── 2. Charts sheet ────────────────────────────────────────────────────
        ws_charts = wb.create_sheet('Charts')
        row = 1
        for c in charts:
            url = c.get('data_url', '')
            if not url or ',' not in url:
                continue
            b64 = url.split(',', 1)[1]
            if not b64:
                continue
            try:
                img_bytes = base64.b64decode(b64)
            except Exception:
                continue
            if len(img_bytes) < 8:
                continue
            img_buf = io.BytesIO(img_bytes)
            try:
                xl_img = Image(img_buf)
            except Exception:
                continue
            img_buf.seek(0)
            TARGET_W = 700
            orig_w, orig_h = xl_img.width, xl_img.height
            if orig_w > 0:
                scale = TARGET_W / orig_w
                xl_img.width  = TARGET_W
                xl_img.height = round(orig_h * scale)
            else:
                xl_img.width, xl_img.height = TARGET_W, 400

            title = c.get('title', '')
            if title:
                ws_charts.cell(row=row, column=1, value=title)
                row += 1

            xl_img.anchor = f'A{row}'
            ws_charts.add_image(xl_img)
            row += round(xl_img.height / 20) + 2

        # ── 3. Group statistics sheets ─────────────────────────────────────────
        if group_export:
            grp_stats    = group_export.get('stats', {})
            samples      = group_export.get('samples', [])
            param_order  = group_export.get('param_order', [])
            param_labels = group_export.get('param_labels', {})
            grp_names    = list(grp_stats.keys())

            if grp_stats and param_order:
                ws_st = wb.create_sheet('Group_Statistics')
                hdr = ['Parameter']
                for g in grp_names:
                    hdr += [f'{g} mean', f'{g} SD', f'{g} N']
                ws_st.append(hdr)
                for p in param_order:
                    stat_row = [param_labels.get(p, p)]
                    for g in grp_names:
                        s = grp_stats.get(g, {}).get('params', {}).get(p)
                        if s:
                            stat_row += [round(s['mean'], 6), round(s['sd'], 6), s.get('n')]
                        else:
                            stat_row += [None, None, None]
                    ws_st.append(stat_row)

            if samples and param_order:
                ws_sp = wb.create_sheet('Group_Samples')
                ws_sp.append(['Sample', 'Group'] + [param_labels.get(p, p) for p in param_order])
                for sr in samples:
                    sample_row = [sr.get('sample'), sr.get('group')]
                    for p in param_order:
                        v = sr.get(p)
                        sample_row.append(round(v, 6) if v is not None else None)
                    ws_sp.append(sample_row)

        wb.save(summary_full)
        return jsonify({'status': 'success', 'xlsx_path': summary_static})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
